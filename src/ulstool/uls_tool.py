#!/usr/bin/env python3
""" ULS download and manipulation tool """

# pylint: disable=wrong-import-order, logging-fstring-interpolation

import argparse
from collections.abc import Iterator
import copy
import csv
import datetime
import enum
import inspect
import locale
import logging
import math
try:
    import openpyxl
except ImportError:
    pass
import operator
import os
import re
import shutil
import socket
import sqlalchemy as sa
import ssl
import sys
import tempfile
import time
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union
import urllib.error
import urllib.parse
import urllib.request
from xml.dom import minidom
import xml.etree.ElementTree as etree
import zipfile

# Script version
VERSION = "2.0"

# Type for row data for exchange with database
ROW_VALUE_TYPE = Optional[Union[str, int, float, bool, datetime.date]]

# Type for field data for database
ROW_DATA_TYPE = \
    Dict[str, Optional[Union[str, int, float, bool, datetime.date]]]

# Minimum height of antenna
MIN_ANT_HEIGHT_M = 3

# Valid range for RX antenna gains
RX_GAIN_RANGE = (10, 80)


def dp(*args, **kwargs):
    """Print debug message

    Arguments:
    args   -- Format and positional arguments. If latter present - formatted
              with %
    kwargs -- Keyword arguments. If present formatted with format()
    """
    msg = args[0] if args else ""
    if len(args) > 1:
        msg = msg % args[1:]
    if args and kwargs:
        msg = msg.format(**kwargs)
    fi = inspect.getframeinfo(inspect.currentframe().f_back)
    print("DP %s()@%d: %s" % (fi.function, fi.lineno, msg))


def ob(v) -> Optional[bool]:
    """ MyPy pacifier. Ensures that argument is either bool or None and returns
    it """
    if v is None:
        return None
    assert isinstance(v, bool)
    return v


def of(v) -> Optional[float]:
    """ MyPy pacifier. Ensures that argument is either float or None and
    returns it """
    if v is None:
        return None
    assert isinstance(v, float)
    return v


def oi(v) -> Optional[int]:
    """ MyPy pacifier. Ensures that argument is either integer or None and
    returns it """
    if v is None:
        return None
    assert isinstance(v, int)
    return v


def ost(v) -> Optional[str]:
    """ MyPy pacifier. Ensures that argument is either string or None and
    returns it """
    if v is None:
        return None
    assert isinstance(v, str)
    return v


def od(v) -> Optional[datetime.datetime]:
    """ MyPy pacifier. Ensures that argument is either datetime or None and
    returns it """
    if v is None:
        return None
    assert isinstance(v, datetime.datetime)
    return v


def mb(v) -> bool:
    """ MyPy pacifier. Ensures that argument is bool and returns it """
    assert isinstance(v, bool)
    return v


def mf(v) -> float:
    """ MyPy pacifier. Ensures that argument is float and returns it """
    assert isinstance(v, float)
    return v


def mi(v) -> int:
    """ MyPy pacifier. Ensures that argument is integer and returns it """
    assert isinstance(v, int)
    return v


def mst(v) -> str:
    """ MyPy pacifier. Ensures that argument is string and returns it """
    assert isinstance(v, str)
    return v


def md(v) -> datetime.datetime:
    """ MyPy pacifier. Ensures that argument is datetime and returns it """
    assert isinstance(v, datetime.datetime)
    return v


def error(msg: str) -> None:
    """ Prints given msg as error message and exit abnormally """
    logging.error(msg)
    sys.exit(1)


def error_if(cond: Any, msg: str) -> None:
    """ If condition evaluates to true prints given msg as error message and
    exits abnormally """
    if cond:
        error(msg)


def spreadsheet_col_name(col_num: int) -> str:
    """ Returns spreadsheet column name for given 1-based column number) """
    ret = ""
    while col_num > 0:
        col_num -= 1
        ret = chr(ord("A") + (col_num % 26)) + ret
        col_num //= 26
    return ret


class Progressor:
    """ Prints progress information (if enabled), erasing previous progress
    message """
    # True if progress printing enabled
    enabled: bool = False

    # Length of previously printed progress line
    _prev_len: int = 0

    @classmethod
    def print_progress(cls, msg: str = "") -> None:
        """ Print given line in progress mode (erasing the previous) """
        if not cls.enabled:
            return
        print(msg + " " * max(0, cls._prev_len - len(msg)) + "\r", end="")
        cls._prev_len = len(msg)

    @classmethod
    def print(cls, msg: str) -> None:
        """ Prints given message in normal mode (still, erasing the previous
        progress message) """
        if not cls.enabled:
            return
        print(msg + " " * max(0, cls._prev_len - len(msg)))
        cls._prev_len = 0


class Timing:
    """ Timing context manager

    Private attributes:
    _start_time -- datetime with operation start time
    _epilogue   -- Operation name to print in the bottom line
    """
    def __init__(self, msg: Optional[str] = None,
                 epilogue: str = "Operation") -> None:
        if msg:
            Progressor.print(msg)
        self._start_time = datetime.datetime.now()
        self._epilogue = epilogue

    def __enter__(self) -> None:
        """ Called on 'with' entry """
        pass

    def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
        """ Called on with exit (normal or by exception) """
        total_seconds = \
            int((datetime.datetime.now() - self._start_time).total_seconds())
        seconds = total_seconds % 60
        total_minutes = total_seconds // 60
        minutes = total_minutes % 60
        hours = total_minutes // 60
        duration = ""
        if hours:
            duration += f"{hours}h "
        if minutes or hours:
            duration += f"{minutes}m "
        duration += f"{seconds}s"
        Progressor.print(f"{self._epilogue} took {duration}")


class FreqRange:
    """ Transmission frequency range

    Public attributes:
    emission_code -- Emission code or None
    start_mhz     -- Emission range lower bound in MHz or None
    end_mhz       -- Emission range upper bound in MHz or None
    """
    # Lower 6GHz subband
    LOW_6G_START_MHZ: int = 5925
    LOW_6G_END_MHZ: int = 6425

    # Upper 6GHz subband
    HIGH_6G_START_MHZ: int = 6525
    HIGH_6G_END_MHZ: int = 6875

    # Rounding before equality comparison
    ROUNDING = 5

    # Maps frequency scale letters to MHz multipliers
    SCALE_MHZ: Dict[str, float] = {"H": 0.000001, "K": 0.001, "M": 1.,
                                   "G": 1000., "T": 1000000.}

    def __init__(
            self, emission_code: Optional[str],
            center_mhz: Optional[float] = None,
            start_mhz: Optional[float] = None,
            end_mhz: Optional[float] = None) -> None:
        """ Constructor. May be constructed from emission designator and center
        frequency (e.g. from ULS database or table) or from emission designator
        and lower/upper frequency bounds (e.g. from AFC database)

        Arguments:
        emission_code -- Emission code or None
        center_mhz    -- Center frequency in MHz or None
        start_mhz     -- Emission range lower bound in MHz or None
        end_mhz       -- Emission range upper bound in MHz or None
        """
        self.emission_code = emission_code
        self.start_mhz = None
        self.end_mhz = None
        if center_mhz is not None:
            bw_mhz = self.emission_bw_mhz(emission_code)
            if bw_mhz is not None:
                self.start_mhz = center_mhz - bw_mhz / 2
                self.end_mhz = center_mhz + bw_mhz / 2
        else:
            self.start_mhz = start_mhz
            self.end_mhz = end_mhz

    def is_6ghz(self):
        """ True if frequency range is valid and falls into 6GHz category """
        if (self.start_mhz is None) or (self.end_mhz is None):
            return False
        for start_6g_mhz, end_6g_mhz in \
                [(self.LOW_6G_START_MHZ, self.LOW_6G_END_MHZ),
                 (self.HIGH_6G_START_MHZ, self.HIGH_6G_END_MHZ)]:
            if (self.end_mhz > start_6g_mhz) and \
                    (self.start_mhz < end_6g_mhz):
                return True
        return False

    def contains(self, other: "FreqRange") -> bool:
        """ True if this range contains given one """
        if not self:
            return False
        if not other:
            return True
        assert (self.start_mhz is not None) and (self.end_mhz is not None)
        assert (other.start_mhz is not None) and (other.end_mhz is not None)
        return (self.start_mhz < other.start_mhz) and \
            (self.end_mhz > other.end_mhz)

    def __bool__(self) -> bool:
        """ True if frequency range is valid (has both bounds defined) """
        return (self.start_mhz is not None) and (self.end_mhz is not None)

    def __eq__(self, other: Any) -> bool:
        """ Equality comparison. Only ends are compared, emission descriptors
        are ignored """
        return isinstance(other, self.__class__) and \
            all((round(self.__dict__[attr], self.ROUNDING) ==
                 round(other.__dict__[attr], self.ROUNDING))
                for attr in ("start_mhz", "end_mhz"))

    def __hash__(self) -> int:
        """ Hash, computed on significant parts of frequency bounds """
        return hash(tuple([round(f, self.ROUNDING)
                           for f in (self.start_mhz, self.end_mhz)
                           if f is not None]))

    @classmethod
    def emission_bw_mhz(cls, emission_code: Optional[str]) -> Optional[float]:
        """ Returns bandwidth in MHz for given emission descriptor

        Arguments:
        emission_code -- Emission code or None
        Returns For valid emission descriptor returns its bandwidth in MHz,
        otherwise returns None
        """
        emission_code = emission_code or ""
        m = re.match(r"^(\d+)([HKMGT])(\d*)$", emission_code[: 4])
        return None if (len(emission_code) < 4) or (m is None) \
            else float(m.group(1) + "." + (m.group(3) or "")) * \
            cls.SCALE_MHZ[m.group(2)]

    def __str__(self):
        """ String representation (for debug purposes) """
        return f"{self.start_mhz}-{self.end_mhz}, {self.emission_code}"

    def __repr__(self):
        """ Debugger representation (for debug purposes) """
        return f"<{self}>"


class ArgContainer:
    """ Container of arguments (e.g. for sa.Field creation)

    Private attributes:
    args   -- List of positional arguments
    kwargs -- Dictionary of named arguments
    """
    def __init__(self, *args, **kwargs):
        self.args = args
        self.kwargs = kwargs

    def apply(self, function: Callable) -> Any:
        """ Calls given function with contained arguments """
        return function(*self.args, **self.kwargs)

    def __getitem__(self, idx: Union[int, str]) -> Any:
        """ Retrieves argument by name/index

        Arguments:
        idx -- Numeric index in positional arguments or named index in named
               arguments
        Returns value of stored argument. None if index is out of range """
        if isinstance(idx, int):
            return self.args[idx] if 0 <= idx <= len(self.args) else None
        return self.kwargs.get(idx)


class UlsTable:
    """ FCC ULS table

    Attributes:
    name             -- Table name
    fields           -- List of UlsTable.Field objects
    id_fields        -- List of names of fields that in a perfect world
                        comprise a table primary key
    indices          -- List of SQLAlchemy composite indices
    make_primary_key -- In ULS database generate primary key from 'id_fields'
    """
    class Field:
        """ Field (column) descriptor

        Attributes:
        dsc      -- Field description (from FCC ULS document)
        col_args -- Arguments for sa.Column constructor. None for fields that
                    do not go to ULS database
        default  -- None or default value (in string representation) to use if
                    field in ULS FCC file is empty
        col_type -- None or column type (sa.Integer(), sa.Unicode(), etc.)
        col_name -- None or column name
        nullable -- True for columns not stored in ULS database or if ULS
                    database column marked as nullable
        """
        def __init__(self, dsc: str,
                     col_args: Optional[ArgContainer] = None,
                     default: str = None) -> None:
            self.dsc = dsc
            self.col_args = col_args
            self.default = default
            self.col_type: Any = None if col_args is None else col_args[1]
            self.col_name: Optional[str] = \
                None if col_args is None else col_args[0]
            self.nullable: bool = \
                True if (col_args is None) or (col_args["nullable"] is None) \
                else col_args["nullable"]

    # All relevant tables
    _all: List["UlsTable"] = []

    # UlsTable objects by name
    _by_name: Dict[str, "UlsTable"] = {}

    def __init__(self, name: str, num_fields: int,
                 fields: List["UlsTable.Field"], id_fields: List[str],
                 indices: Optional[List[sa.Index]] = None,
                 make_primary_key: bool = True) -> None:
        """ Constructor

        Arguments:
        name             -- Table name
        num_fields       -- Number of fields (for self-check purposes)
        fields           -- List of Field objects
        id_fields        -- List of names of fields that in the perfect world
                            completely identify the row (e.g. comprise a
                            primary key)
        indices          -- None or list of SQLAlchemy composite indices
        make_primary_key -- In ULS database generate primary key from
                            'id_fields'
        """
        self.name = name
        assert num_fields == len(fields)
        self.fields: List["UlsTable.Field"] = fields
        self.id_fields: List[str] = id_fields
        self.indices: List[sa.Index] = indices or []
        self.make_primary_key = make_primary_key

    @classmethod
    def all(cls) -> List["UlsTable"]:
        """ List of all ULS table object """
        cls._prepare_tables()
        return cls._all

    @classmethod
    def by_name(cls, table_name: str) -> "UlsTable":
        """ Returns ULS table by name """
        cls._prepare_tables()
        return cls._by_name[table_name]

    @classmethod
    def try_by_name(cls, table_name: str) -> Optional["UlsTable"]:
        """ Returns ULS table by name (None for wrong name)"""
        cls._prepare_tables()
        return cls._by_name.get(table_name)

    @classmethod
    def _prepare_tables(cls) -> None:
        """ Prepares _all and _by_name if not yet prepared """
        if cls._all:
            return
        F = cls.Field
        AC = ArgContainer
        cls._all = [
            cls("AN", 38,   # Antenna
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("Antenna Action Performed",
                   AC("antenna_action_performed", sa.Unicode(1),
                      nullable=True)),
                 F("Antenna Number",
                   AC("antenna_number", sa.Integer(), nullable=False),
                   default="1"),
                 F("Location Number",
                   AC("location_number", sa.Integer(), nullable=False)),
                 F("Receive Zone Code",
                   AC("receive_zone_code", sa.Unicode(6), nullable=True)),
                 F("Antenna Type Code",
                   AC("antenna_type_code", sa.Unicode(1), nullable=True)),
                 F("Height to Tip",
                   AC("height_to_tip", sa.Float(), nullable=True)),
                 F("Height to Center RAAT",
                   AC("height_to_center_raat", sa.Float(), nullable=True)),
                 F("Antenna Make",
                   AC("antenna_make", sa.Unicode(25), nullable=True)),
                 F("Antenna Model",
                   AC("antenna_model", sa.Unicode(25), nullable=True)),
                 F("Tilt",
                   AC("tilt", sa.Float(), nullable=True)),
                 F("Polarization Code",
                   AC("polarization_code", sa.Unicode(5), nullable=True)),
                 F("Beamwidth",
                   AC("beamwidth", sa.Float(), nullable=True)),
                 F("Gain",
                   AC("gain", sa.Float(), nullable=True)),
                 F("Azimuth",
                   AC("azimuth", sa.Float(), nullable=True)),
                 F("Height Above AvgTerrain",
                   AC("height_above_avg_terrain", sa.Float(), nullable=True)),
                 F("Diversity Height",
                   AC("diversity_height", sa.Float(), nullable=True)),
                 F("Diversity Gain",
                   AC("diversity_gain", sa.Float(), nullable=True)),
                 F("Diversity Beam",
                   AC("diversity_beam", sa.Float(), nullable=True)),
                 F("Reflector Height",
                   AC("reflector_height", sa.Float(), nullable=True)),
                 F("Reflector Width",
                   AC("reflector_width", sa.Float(), nullable=True)),
                 F("Reflector Separation",
                   AC("reflector_separation ", sa.Float(), nullable=True)),
                 F("Passive Repeater Number",
                   AC("repeater_seq_num", sa.Integer(), nullable=True)),
                 F("Back-to-Back Tx Dish Gain",
                   AC("back_to_back_tx_dish_gain", sa.Float(), nullable=True)),
                 F("Back-to-Back Rx Dish Gain",
                   AC("back_to_back_rx_dish_gain", sa.Float(), nullable=True)),
                 F("Location Name",
                   AC("location_name", sa.Unicode(20), nullable=True)),
                 F("Passive Repeater Sequence ID",
                   AC("passive_repeater_id", sa.Integer(), nullable=True)),
                 F("Alternative CGSA Method",
                   AC("alternative_cgsa_method", sa.Unicode(1),
                      nullable=True)),
                 F("Path Number",
                   AC("path_number", sa.Integer(), nullable=False),
                   default="1"),
                 F("Line loss",
                   AC("line_loss", sa.Float(), nullable=True)),
                 F("Status Code",
                   AC("status_code", sa.Unicode(1), nullable=True)),
                 F("Status Date",
                   AC("status_date", sa.DateTime(), nullable=True)),
                 F("PSD/Non-PSD Methodology",
                   AC("psd_nonpsd_methodology", sa.Unicode(10),
                      nullable=True)),
                 F("Maximum ERP",
                   AC("maximum_erp", sa.Float(), nullable=True))],
                id_fields=["call_sign", "location_number", "antenna_number",
                           "path_number"],
                indices=[sa.Index("AN_ix", "call_sign", "location_number",
                                  "antenna_number", "path_number",
                                  unique=True)]),
            cls("CP", 14,   # Control point
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("Control Point Action Performed",
                   AC("control_point_action_performed ", sa.Unicode(1),
                      nullable=True)),
                 F("Control Point Number",
                   AC("control_point_number", sa.Integer(), nullable=False)),
                 F("Control Address",
                   AC("control_address", sa.Unicode(80), nullable=True)),
                 F("Control City",
                   AC("control_city", sa.Unicode(20), nullable=True)),
                 F("State Code",
                   AC("state_code", sa.Unicode(2), nullable=True)),
                 F("Control Phone",
                   AC("control_phone", sa.Unicode(10), nullable=True)),
                 F("Control County",
                   AC("control_county", sa.Unicode(60), nullable=True)),
                 F("Status Code",
                   AC("status_code", sa.Unicode(1), nullable=True)),
                 F("Status Date",
                   AC("status_date", sa.DateTime(), nullable=True))],
                id_fields=["call_sign", "control_point_number"],
                indices=[sa.Index("CP_ix", "call_sign", "control_point_number",
                                  unique=True)]),
            cls("EM", 16,   # Emission
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("Location Number",
                   AC("location_number", sa.Integer(), nullable=False)),
                 F("Antenna Number",
                   AC("antenna_number", sa.Integer(), nullable=False)),
                 F("Frequency Assigned/Channel Center",
                   AC("frequency_assigned", sa.Float(), nullable=False)),
                 F("Emission Action Performed",
                   AC("emission_action_performed", sa.Unicode(1),
                      nullable=True)),
                 F("Emission Code",
                   AC("emission_code", sa.Unicode(10), nullable=True)),
                 F("Digital Mod Rate",
                   AC("digital_mod_rate", sa.Float(), nullable=True)),
                 F("Digital Mod Type",
                   AC("digital_mod_type", sa.Unicode(255), nullable=True)),
                 F("Frequency Number",
                   AC("frequency_number", sa.Integer(), nullable=False),
                   default="1"),
                 F("Status Code",
                   AC("status_code", sa.Unicode(1), nullable=True)),
                 F("Status Date",
                   AC("status_date", sa.DateTime(), nullable=True)),
                 F("Emission Sequence Id",
                   AC("emission_sequence_id", sa.Integer(), nullable=False))],
                id_fields=["call_sign", "location_number", "antenna_number",
                           "frequency_number",
                           "emission_sequence_id"],
                indices=[sa.Index("EM_ix", "call_sign", "location_number",
                                  "antenna_number", unique=False),
                         sa.Index("EM_id_ix", "call_sign", "location_number",
                                  "antenna_number", "frequency_number",
                                  "emission_sequence_id", unique=False),
                         sa.Index("EM_ant_idx", "call_sign", "location_number",
                                  "antenna_number", unique=False)],
                make_primary_key=False),
            cls("EN", 30,   # Entity
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("Entity Type",
                   AC("entity_type", sa.Unicode(2), nullable=False)),
                 F("Licensee ID",
                   AC("licensee_id", sa.Unicode(9), nullable=True)),
                 F("Entity Name",
                   AC("entity_name", sa.Unicode(200), nullable=True)),
                 F("First Name",
                   AC("first_name", sa.Unicode(20), nullable=True)),
                 F("MI",
                   AC("mi", sa.Unicode(1), nullable=True)),
                 F("Last Name",
                   AC("last_name", sa.Unicode(20), nullable=True)),
                 F("Suffix",
                   AC("suffix", sa.Unicode(3), nullable=True)),
                 F("Phone",
                   AC("phone", sa.Unicode(10), nullable=True)),
                 F("Fax",
                   AC("fax", sa.Unicode(10), nullable=True)),
                 F("Email",
                   AC("email", sa.Unicode(50), nullable=True)),
                 F("Street Address",
                   AC("street_address", sa.Unicode(60), nullable=True)),
                 F("City",
                   AC("city", sa.Unicode(20), nullable=True)),
                 F("State",
                   AC("state", sa.Unicode(2), nullable=True)),
                 F("Zip Code",
                   AC("zip_code", sa.Unicode(9), nullable=True)),
                 F("PO Box",
                   AC("po_box", sa.Unicode(20), nullable=True)),
                 F("Attention Line",
                   AC("attention_line", sa.Unicode(35), nullable=True)),
                 F("SGIN",
                   AC("sgin", sa.Unicode(3), nullable=True)),
                 F("FCC Registration Number (FRN)",
                   AC("frn", sa.Unicode(10), nullable=True)),
                 F("Applicant Type Code",
                   AC("applicant_type_code", sa.Unicode(1), nullable=True)),
                 F("Applicant Type Code Other",
                   AC("applicant_type_other", sa.Unicode(40), nullable=True)),
                 F("Status Code",
                   AC("status_code", sa.Unicode(1), nullable=True)),
                 F("Status Date",
                   AC("status_date", sa.DateTime(), nullable=True)),
                 F("3.7 GHz License Type",
                   AC("lic_category_code", sa.Unicode(1), nullable=True)),
                 F("Linked Unique System Identifier",
                   AC("linked_license_id", sa.Integer(), nullable=True)),
                 F("Linked Call Sign",
                   AC("linked_callsign", sa.Unicode(10), nullable=True))],
                id_fields=["call_sign", "entity_type"],
                indices=[sa.Index("EN_ix", "call_sign", unique=False),
                         sa.Index("EN_id_ix", "call_sign", "entity_type")]),
            cls("FR", 30,   # Frequency
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("Frequency Action Performed",
                   AC("frequency_action_performed", sa.Unicode(1),
                      nullable=True)),
                 F("Location Number",
                   AC("location_number", sa.Integer(), nullable=False)),
                 F("Antenna Number",
                   AC("antenna_number", sa.Integer(), nullable=False)),
                 F("Class Station Code",
                   AC("class_station_code", sa.Unicode(4), nullable=True)),
                 F("Op Altitude Code",
                   AC("op_altitude_code", sa.Unicode(1), nullable=True)),
                 F("Frequency Assigned",
                   AC("frequency_assigned", sa.Float(), nullable=True)),
                 F("Frequency Upper Band",
                   AC("frequency_upper_band", sa.Float(), nullable=True)),
                 F("Frequency Carrier",
                   AC("frequency_carrier", sa.Float(), nullable=True)),
                 F("Time Begin Operations",
                   AC("time_begin_operations", sa.Integer(), nullable=True)),
                 F("Time End Operations",
                   AC("time_end_operations", sa.Integer(), nullable=True)),
                 F("Power Output",
                   AC("power_output", sa.Float(), nullable=True)),
                 F("Power ERP",
                   AC("power_erp", sa.Float(), nullable=True)),
                 F("Tolerance",
                   AC("tolerance", sa.Float(), nullable=True)),
                 F("Frequency Indicator",
                   AC("frequency_ind", sa.Unicode(1), nullable=True)),
                 F("Status",
                   AC("status", sa.Unicode(1), nullable=True)),
                 F("EIRP",
                   AC("eirp", sa.Float(), nullable=True)),
                 F("Transmitter Make",
                   AC("transmitter_make", sa.Unicode(25), nullable=True)),
                 F("Transmitter Model",
                   AC("transmitter_model", sa.Unicode(25), nullable=True)),
                 F("Auto Transmitter Power Control",
                   AC("auto_transmitter_power_control", sa.Unicode(1),
                      nullable=True)),
                 F("Number of Units",
                   AC("cnt_mobile_units", sa.Integer(), nullable=True)),
                 F("Number of Paging Receivers",
                   AC("cnt_mob_pagers", sa.Integer(), nullable=True)),
                 F("Frequency Number",
                   AC("freq_seq_id", sa.Integer(), nullable=False),
                   default="1"),
                 F("Status Code",
                   AC("status_code", sa.Unicode(1), nullable=True)),
                 F("Status Date",
                   AC("status_date", sa.DateTime(), nullable=True)),
                 F("Date First Used",
                   AC("date_first_used", sa.DateTime(), nullable=True))],
                id_fields=["call_sign", "location_number", "antenna_number",
                           "freq_seq_id"],
                indices=[sa.Index("FR_ix", "call_sign", "location_number",
                                  "antenna_number", "freq_seq_id",
                                  unique=False)],
                make_primary_key=False),
            cls("HD", 59,   # Application/License
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("License Status",
                   AC("license_status", sa.Unicode(1), nullable=False)),
                 F("Radio Service Code",
                   AC("radio_service_code", sa.Unicode(2), nullable=False)),
                 F("Grant Date",
                   AC("grant_date", sa.Unicode(10), nullable=True)),
                 F("Expired Date",
                   AC("expired_date", sa.Unicode(10), nullable=True)),
                 F("Cancellation Date",
                   AC("cancellation_date", sa.Unicode(10), nullable=True)),
                 F("Eligibility Rule Num",
                   AC("eligibility_rule_num", sa.Unicode(10), nullable=True)),
                 F("Reserved",
                   AC("applicant_type_code_reserved", sa.Unicode(1),
                      nullable=True)),
                 F("Alien",
                   AC("alien", sa.Unicode(1), nullable=True)),
                 F("Alien Government",
                   AC("alien_government", sa.Unicode(1), nullable=True)),
                 F("Alien Corporation",
                   AC("alien_corporation", sa.Unicode(1), nullable=True)),
                 F("Alien Officer",
                   AC("alien_officer", sa.Unicode(1), nullable=True)),
                 F("Alien Control",
                   AC("alien_control", sa.Unicode(1), nullable=True)),
                 F("Revoked",
                   AC("revoked", sa.Unicode(1), nullable=True)),
                 F("Convicted",
                   AC("convicted", sa.Unicode(1), nullable=True)),
                 F("Adjudged",
                   AC("adjudged", sa.Unicode(1), nullable=True)),
                 F("Reserved",
                   AC("involved_reserved", sa.Unicode(1), nullable=True)),
                 F("Common Carrier",
                   AC("common_carrier", sa.Unicode(1), nullable=True)),
                 F("Non Common Carrier",
                   AC("non_common_carrier", sa.Unicode(1), nullable=True)),
                 F("Private Comm",
                   AC("private_comm", sa.Unicode(1), nullable=True)),
                 F("Fixed",
                   AC("fixed", sa.Unicode(1), nullable=True)),
                 F("Mobile",
                   AC("mobile", sa.Unicode(1), nullable=True)),
                 F("Radiolocation",
                   AC("radiolocation", sa.Unicode(1), nullable=True)),
                 F("Satellite",
                   AC("satellite", sa.Unicode(1), nullable=True)),
                 F("Developmental or STA or Demonstration",
                   AC("developmental_or_sta", sa.Unicode(1), nullable=True)),
                 F("InterconnectedService",
                   AC("interconnected_service", sa.Unicode(1), nullable=True)),
                 F("Certifier First Name",
                   AC("certifier_first_name", sa.Unicode(20), nullable=True)),
                 F("Certifier MI",
                   AC("certifier_mi", sa.Unicode(1), nullable=True)),
                 F("Certifier Last Name",
                   AC("certifier_last_name", sa.Unicode(20), nullable=True)),
                 F("Certifier Suffix",
                   AC("certifier_suffix", sa.Unicode(3), nullable=True)),
                 F("Certifier Title",
                   AC("certifier_title", sa.Unicode(40), nullable=True)),
                 F("Female",
                   AC("gender", sa.Unicode(1), nullable=True)),  # Sic!
                 F("Black or African-American",
                   AC("african_american", sa.Unicode(1), nullable=True)),
                 F("Native American",
                   AC("native_american", sa.Unicode(1), nullable=True)),
                 F("Hawaiian",
                   AC("hawaiian", sa.Unicode(1), nullable=True)),
                 F("Asian",
                   AC("asian", sa.Unicode(1), nullable=True)),
                 F("White",
                   AC("white", sa.Unicode(1), nullable=True)),
                 F("Hispanic",
                   AC("ethnicity", sa.Unicode(1), nullable=True)),  # Sic!
                 F("Effective Date",
                   AC("effective_date", sa.Unicode(10), nullable=True)),
                 F("Last Action Date",
                   AC("last_action_date", sa.Unicode(1), nullable=False)),
                 F("Auction ID",
                   AC("auction_id", sa.Integer(), nullable=True)),
                 F("Broadcast Services - Regulatory Status",
                   AC("reg_stat_broad_serv", sa.Unicode(1), nullable=True)),
                 F("Band Manager - Regulatory Status",
                   AC("band_manager", sa.Unicode(1), nullable=True)),
                 F("Broadcast Services - Type of Radio Service",
                   AC("type_serv_broad_serv", sa.Unicode(1), nullable=True)),
                 F("Alien Ruling",
                   AC("alien_ruling", sa.Unicode(1), nullable=True)),
                 F("Licensee Name Change",
                   AC("licensee_name_change", sa.Unicode(1), nullable=True)),
                 F("Whitespace Indicator",
                   AC("whitespace_ind", sa.Unicode(1), nullable=True)),
                 F("Operation/Performance Requirement Choice",
                   AC("additional_cert_choice", sa.Unicode(1), nullable=True)),
                 F("Operation/Performance Requirement Answer",
                   AC("additional_cert_answer", sa.Unicode(1), nullable=True)),
                 F("Discontinuation of Service",
                   AC("discontinuation_ind", sa.Unicode(1), nullable=True)),
                 F("Regulatory Compliance",
                   AC("regulatory_compliance_ind", sa.Unicode(1),
                      nullable=True)),
                 F("900 MHz Eligibility Certification",
                   AC("eligibility_cert_900", sa.Unicode(1), nullable=True)),
                 F("900 MHz Transition Plan Certification",
                   AC("transition_plan_cert_900", sa.Unicode(1),
                      nullable=True)),
                 F("900 MHz Return Spectrum Certification",
                   AC("return_spectrum_cert_900", sa.Unicode(1),
                      nullable=True)),
                 F("900 MHz Payment Certification",
                   AC("payment_cert_900", sa.Unicode(1), nullable=True))],
                id_fields=["call_sign"],
                indices=[sa.Index("HD_ix", "call_sign", unique=True)]),
            cls("LO", 51,   # Location
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("Location Action Performed",
                   AC("location_action_performed", sa.Unicode(1),
                      nullable=True)),
                 F("Location Type Code",
                   AC("location_type_code", sa.Unicode(1), nullable=True)),
                 F("Location Class Code",
                   AC("loc_class_code", sa.Unicode(1), nullable=False)),
                 F("Location Number",
                   AC("location_number", sa.Integer(), nullable=False)),
                 F("Site Status",
                   AC("site_status", sa.Unicode(1), nullable=True)),
                 F("Corresponding Fixed Location",
                   AC("corresponding_fixed_location", sa.Integer(),
                      nullable=True)),
                 F("Location Address",
                   AC("location_address", sa.Unicode(80), nullable=True)),
                 F("Location City",
                   AC("location_city", sa.Unicode(20), nullable=True)),
                 F("Location County/Borough/Parish",
                   AC("location_county", sa.Unicode(60), nullable=True)),
                 F("Location State",
                   AC("location_state", sa.Unicode(2), nullable=True)),
                 F("Radius of Operation",
                   AC("radius_of_operation", sa.Float(), nullable=True)),
                 F("Area of Operation Code",
                   AC("area_of_operation_code", sa.Unicode(1), nullable=True)),
                 F("Clearance Indicator",
                   AC("clearance_indicator", sa.Unicode(1), nullable=True)),
                 F("Ground Elevation",
                   AC("ground_elevation", sa.Float(), nullable=True)),
                 F("Latitude Degrees",
                   AC("lat_degrees", sa.Integer(), nullable=True)),
                 F("Latitude Minutes",
                   AC("lat_minutes", sa.Integer(), nullable=True)),
                 F("Latitude Seconds",
                   AC("lat_seconds", sa.Float(), nullable=True)),
                 F("Latitude Direction",
                   AC("lat_direction", sa.Unicode(1), nullable=True)),
                 F("Longitude Degrees",
                   AC("long_degrees", sa.Integer(), nullable=True)),
                 F("Longitude Minutes",
                   AC("long_minutes", sa.Integer(), nullable=True)),
                 F("LongitudeSeconds",
                   AC("long_seconds", sa.Float(), nullable=True)),
                 F("Longitude Direction",
                   AC("long_direction", sa.Unicode(1), nullable=True)),
                 F("Max Latitude Degrees",
                   AC("max_lat_degrees", sa.Integer(), nullable=True)),
                 F("Max Latitude Minutes",
                   AC("max_lat_minutes", sa.Integer(), nullable=True)),
                 F("Max Latitude Seconds",
                   AC("max_lat_seconds", sa.Float(), nullable=True)),
                 F("Max Latitude Direction",
                   AC("max_lat_direction", sa.Unicode(1), nullable=True)),
                 F("Max Longitude Degrees",
                   AC("max_long_degrees", sa.Integer(), nullable=True)),
                 F("Max Longitude Minutes",
                   AC("max_long_minutes", sa.Integer(), nullable=True)),
                 F("Max Longitude Seconds",
                   AC("max_long_seconds", sa.Float(), nullable=True)),
                 F("Max Longitude Direction",
                   AC("max_long_direction", sa.Unicode(1), nullable=True)),
                 F("Nepa",
                   AC("nepa", sa.Unicode(1), nullable=True)),
                 F("Quiet Zone Notification Date",
                   AC("quiet_zone_notification_date", sa.Unicode(10),
                      nullable=True)),
                 F("Tower Registration Number",
                   AC("tower_registration_number", sa.Unicode(10),
                      nullable=True)),
                 F("Height of Support Structure",
                   AC("height_of_support_structure", sa.Float(),
                      nullable=True)),
                 F("Overall Height of Structure",
                   AC("overall_height_of_structure", sa.Float(),
                      nullable=True)),
                 F("Structure Type",
                   AC("structure_type", sa.Unicode(7), nullable=True)),
                 F("Airport ID",
                   AC("airport_id", sa.Unicode(4), nullable=True)),
                 F("Location Name",
                   AC("location_name", sa.Unicode(20), nullable=True)),
                 F("Units Hand Held",
                   AC("units_hand_held", sa.Integer(), nullable=True)),
                 F("Units Mobile",
                   AC("units_mobile", sa.Integer(), nullable=True)),
                 F("Units Temp Fixed",
                   AC("units_temp_fixed", sa.Integer(), nullable=True)),
                 F("Units Aircraft",
                   AC("units_aircraft", sa.Integer(), nullable=True)),
                 F("Units Itinerant",
                   AC("units_itinerant", sa.Integer(), nullable=True)),
                 F("Status Code",
                   AC("status_code", sa.Unicode(1), nullable=True)),
                 F("Status Date",
                   AC("status_date", sa.DateTime(), nullable=True)),
                 F("Earth Station Agreement",
                   AC("earth_agree", sa.Unicode(1), nullable=True))],
                id_fields=["call_sign", "location_number"],
                indices=[sa.Index("LO_ix", "call_sign", "location_number",
                                  unique=True)]),
            cls("PA", 22,   # Path
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("Action Performed",
                   AC("path_action_performed", sa.Unicode(1), nullable=True)),
                 F("Path Number / Link Number",
                   AC("path_number", sa.Integer(), nullable=False)),
                 F("Transmit Location Number",
                   AC("transmit_location_number", sa.Integer(),
                      nullable=False)),
                 F("Transmit Antenna Number",
                   AC("transmit_antenna_number", sa.Integer(),
                      nullable=False)),
                 F("Receiver Location Number",
                   AC("receiver_location_number", sa.Integer(),
                      nullable=False)),
                 F("Receiver Antenna Number",
                   AC("receiver_antenna_number", sa.Integer(), nullable=False),
                   default="1"),
                 F("MAS/DEMS SubType of Operation",
                   AC("mas_dems_subtype", sa.Unicode(2), nullable=True)),
                 F("Path Type Code",
                   AC("path_type_desc", sa.Unicode(20), nullable=True)),
                 F("Passive Receiver Indicator",
                   AC("passive_receiver_indicator", sa.Unicode(1),
                      nullable=True)),
                 F("Country Code",
                   AC("country_code", sa.Unicode(3), nullable=True)),
                 F("Interference to GSO",
                   AC("interference_to_gso", sa.Unicode(1), nullable=True)),
                 F("Receiver Call Sign",
                   AC("receiver_callsign", sa.Unicode(10), nullable=True)),
                 F("AngularSeparation",
                   AC("angular_sep", sa.Float(), nullable=True)),
                 F("Cert No Alternative",
                   AC("cert_no_alternative", sa.Unicode(1), nullable=True)),
                 F("Cert No Interference",
                   AC("cert_no_interference", sa.Unicode(1), nullable=True)),
                 F("Status Code",
                   AC("status_code", sa.Unicode(1), nullable=True)),
                 F("Status Date",
                   AC("status_date", sa.DateTime(), nullable=True))],
                id_fields=["call_sign", "path_number"],
                indices=[sa.Index("PA_ix", "call_sign", "path_number",
                                  unique=True),
                         sa.Index("PA_tx_ix", "call_sign",
                                  "transmit_location_number",
                                  "transmit_antenna_number", unique=False)]),
            cls("SG", 15,   # Segment
                [F("Record Type",
                   AC("record_type", sa.Unicode(2), nullable=False)),
                 F("Unique System Identifier",
                   AC("unique_system_identifier", sa.Integer(),
                      nullable=False)),
                 F("ULS File Number",
                   AC("uls_file_number", sa.Unicode(14), nullable=True)),
                 F("EBF Number",
                   AC("ebf_number", sa.Unicode(30), nullable=True)),
                 F("Call Sign",
                   AC("call_sign", sa.Unicode(10), nullable=False)),
                 F("Action Performed",
                   AC("segment_action_performed", sa.Unicode(1),
                      nullable=True)),
                 F("Path Number",
                   AC("path_number", sa.Integer(), nullable=False)),
                 F("Transmit Location Number",
                   AC("transmit_location", sa.Integer(), nullable=False)),
                 F("Transmit Antenna Number",
                   AC("transmit_antenna", sa.Integer(), nullable=False),
                   default="1"),
                 F("Receiver Location Number",
                   AC("receiver_location", sa.Integer(), nullable=False)),
                 F("Receiver Antenna Number",
                   AC("receiver_antenna", sa.Integer(), nullable=False),
                   default="1"),
                 F("Segment Number",
                   AC("segment_number", sa.Integer(), nullable=False),
                   default="1"),
                 F("Segment Length",
                   AC("segment_length", sa.Float(), nullable=True)),
                 F("Status Code",
                   AC("status_code", sa.Unicode(1), nullable=True)),
                 F("Status Date",
                   AC("status_date", sa.DateTime(), nullable=True))],
                id_fields=["call_sign", "path_number", "segment_number"],
                indices=[sa.Index("SG_ix", "call_sign", "path_number",
                                  unique=False),
                         sa.Index("SG_id_ix", "call_sign", "path_number",
                                  "segment_number", unique=True)]),
            ]
        cls._by_name = {t.name: t for t in cls._all}


class Complainer:
    """ Manages printing complaints to log and report file
    This class is intended to be used as content manager (with 'with' clause)

    Private attributes:
    self._filename   -- Name of file to write report to. None if report not
                        needed
    self._complaints -- List of complaints made
    """
    class ComplaintBase:
        """ Base class for complaints

        Public attributes:
        category       -- Category to print in report file
        sort_key       -- Sort key for complaints within category
        report_message -- Message for report
        priority       -- Category priority that define the order in which
                          category appear in report (0 is lowest and default)
        """
        def __init__(self, category: str, sort_key: Any, rm: str,
                     lm: Optional[str] = None, priority: int = 0) -> None:
            """ Constructor
            Prints log message to log, adds complaint to complaint list, if it
            is defined

            Attributes:
            category -- Category to print in report file
            sort_key -- Sort key for complaints within category
            rm       -- Message for report. Empty to not report
            lm       -- Message for log. None to use message for report. Empty
                        to not log message
            priority -- Category priority that define the order in which
                        category appear in report (0 is lowest and default)
            """
            self.category = category
            self.priority = priority
            self.sort_key = sort_key
            self.report_message = rm
            if rm and (Complainer.instance is not None):
                Complainer.instance.add_complaint(self)
            if lm is None:
                lm = rm
            if lm:
                logging.warning(lm)

    # Instance of this type, used by ComplaintBase to add complaints
    instance: Optional["Complainer"] = None

    def __init__(self, filename: Optional[str]) -> None:
        """ Constructor

        Arguments:
        filename -- None or name of file for reporting
        """
        self._filename = filename
        self._complaints: List["Complainer.ComplaintBase"] = []

    def __enter__(self) -> None:
        """ Context manager entry """
        assert self.__class__.instance is None
        if self._filename:
            self.__class__.instance = self

    def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
        """ Context manager exit """
        assert (self.__class__.instance is None) or \
            (self.__class__.instance == self)
        self.__class__.instance = None
        if not self._filename:
            return
        grouped_complaints: Dict[Tuple[int, str],
                                 List["Complainer.ComplaintBase"]] = {}
        for complaint in self._complaints:
            group_key = (-complaint.priority, complaint.category)
            grouped_complaints.setdefault(group_key, []).append(complaint)
        with open(self._filename, mode="w", encoding="ascii") as f:
            first_group = True
            for group_key in sorted(grouped_complaints.keys()):
                group_complaints = grouped_complaints[group_key]
                print(("" if first_group else "\n\n") +
                      f"{group_complaints[0].category}", file=f)
                first_group = False
                for complaint in sorted(group_complaints,
                                        key=operator.attrgetter("sort_key")):
                    print(f"  {complaint.report_message}", file=f)

    def add_complaint(self, complaint: "Complainer.ComplaintBase") -> None:
        """ Adds complaint to complaint list """
        self._complaints.append(complaint)


class ComplaintReplacedMandatoryField(Complainer.ComplaintBase):
    """ Complaint on missing mandatory field """
    def __init__(self, table: str, line_num, column: str,
                 replacement: str) -> None:
        """ Constructor

        Arguments:
        table       -- Table name
        line_num    -- Line number
        column      -- Column name
        replacement -- Replacement
        """
        super().__init__(
            category="Replacement of empty mandatory field",
            sort_key=(table, line_num, column),
            rm=(f"'{table}' table missing value for column '{column}' at line "
                f"{line_num}. Column value replaced by {replacement}"),
            lm="")


class ComplaintDataDuplication(Complainer.ComplaintBase):
    """ Complaint on data duplication in ULS table """
    def __init__(self, table: str, line_num: int, row_data: ROW_DATA_TYPE) \
            -> None:
        """ Constructor

        Arguments:
        table     -- Table name
        line_num  -- Line number
        row_data  -- Dictionary that represents row data
        """
        key_fields_str = \
            ", ".join(f"'{k}': {row_data[k]}'"
                      for k in UlsTable.by_name(table).id_fields)
        super().__init__(
            category="Duplication of identifying fields (of primary key)",
            sort_key=(table, line_num),
            rm=(f"Table {table} has data duplication at line {line_num}. "
                f"Values of key_fields: {key_fields_str}. Duplicate dropped"),
            priority=100)


class ComplaintEmissionCode(Complainer.ComplaintBase):
    """ Complaint on unsupported emission code """
    def __init__(self, callsign: str, emission_code: str, line_num: int) \
            -> None:
        """ Constructor

        Arguments:
        callsign      -- Callsign
        emission_code -- Offending emissions descriptor
        line_num      -- Line number in EM table
        """
        super().__init__(
            category="Unsupported emission code structure",
            sort_key=line_num,
            rm=(f"'EM' table, callsign {callsign} has unsupported emission "
                f"code {emission_code} at line {line_num}"))


class ComplaintMissingEmission(Complainer.ComplaintBase):
    """ Complaint on missing emission record """
    def __init__(self, callsign: str, loc_num: int, ant_num: int,
                 freq_num: int, line_num: int,
                 center_freq: Optional[float] = None) -> None:
        """ Constructor

        Arguments:
        callsign    -- Callsign
        loc_num     -- Location number
        ant_num     -- Antenna number
        freq_num    -- Frequency number
        line_num    -- Line number
        center_freq -- Center frequency in MHz (if appropriate) or None
        """
        cf = f"center frequency {center_freq}" if center_freq is not None \
            else ""
        super().__init__(
            category="Missing emission code",
            sort_key=(callsign, loc_num, ant_num),
            rm=(f"Callsign {callsign} missing emission descriptor for "
                f"location {loc_num}, antenna {ant_num}, frequency "
                f"{freq_num}{(', ' + cf) if cf else ''}, required for 'FR' "
                f"table row {line_num}. Row dropped"))


class ComplaintMissingSegment(Complainer.ComplaintBase):
    """ Complaint on missing segment definition """
    def __init__(self, callsign: str, path_num: int, seg_num: int) -> None:
        """ Constructor

        Arguments:
        callsign -- Callsign
        path_num -- Path number
        seg_num  -- Missing segment number
        """
        super().__init__(
            category="Missing segment definition",
            sort_key=(callsign, path_num, seg_num),
            rm=(f"Path {callsign}/{path_num} lacks definition of segment "
                f"{seg_num}. Path dropped"))


class ComplaintWrongRepeaterMark(Complainer.ComplaintBase):
    """ Complaint on path whose PR mark differs from repeaters' absence """
    def __init__(self, callsign: str, path_num: int, has_repeaters: bool) \
            -> None:
        """ Constructor

        Arguments:
        callsign      -- Path callsign
        path_num      -- Path number
        marked_pr     -- True if marked as having passive repeaters, False if
                         marked as not having them
        has_repeaters -- True if has repeaters, False if not
        """
        if has_repeaters:
            super().__init__(
                category="Multiple segments in path marked as having no " +
                "repeaters",
                sort_key=(callsign, path_num),
                rm=(f"Path {callsign}/{path_num} in 'PA' "
                    f"table marked as having no repeaters  but 'SG' table has "
                    f"more than one  segment for this path. Path considered "
                    f"to having  repeaters"))
        else:
            super().__init__(
                category="One segment in path marked as having repeaters",
                sort_key=(callsign, path_num),
                rm=(f"Path {callsign}/{path_num} in 'PA' "
                    f"table marked as having passive repeaters but 'SG' table "
                    f"has just one segment for this path. Path considered to "
                    f"having no repeaters"))


class ComplaintMissingCoordinates(Complainer.ComplaintBase):
    """ Complaint on missing coordinates """
    def __init__(self, callsign: str, path_num: int, role: str) -> None:
        """ Constructor

        Arguments:
        callsign -- Callsign
        path_num -- Path number
        role     -- Location role
        """
        super().__init__(
            category="Missing location coordinates",
            sort_key=(callsign),
            rm=(f"Path {callsign}/{path_num} lacks {role} coordinates. Path "
                f"dropped"))


class ComplaintMissingAntennaHeight(Complainer.ComplaintBase):
    """ Complaint on missing antenna height """
    def __init__(self, callsign: str, path_num: int, role: str) -> None:
        """ Constructor

        Arguments:
        callsign -- Callsign
        path_num -- Path number
        role     -- Location role
        """
        super().__init__(
            category="Missing antenna height",
            sort_key=(callsign, path_num),
            rm=(f"Path {callsign}/{path_num} lacks {role} antenna height. "
                f"Path dropped"))


class ComplaintLowAntennaHeight(Complainer.ComplaintBase):
    """ Complaint on missing antenna height """
    def __init__(self, callsign: str, path_num: int, role: str,
                 raat_height_m: float, drop: bool) -> None:
        """ Constructor

        Arguments:
        callsign -- Callsign
        path_num -- Path number
        role     -- Location role
        drop     -- True if path will be dropped
        """
        super().__init__(
            category=f"Antenna height to center RAAT < {MIN_ANT_HEIGHT_M}",
            sort_key=(callsign, path_num),
            rm=(f"Path {callsign}/{path_num} {role} antenna RAAT height is "
                f"{raat_height_m}, less than {MIN_ANT_HEIGHT_M}"
                f"{'. Path dropped' if drop else ''}"))


class ComplaintExpiredActiveLicense(Complainer.ComplaintBase):
    """ Complaint expired license still marked as active """
    def __init__(self, callsign: str, license_status: str,
                 expired_date: Optional[datetime.date],
                 cancellation_date: Optional[datetime.date]) -> None:
        """ Constructor

        Arguments:
        callsign          -- License callsign
        license_status    -- License status
        expired_date      -- Optional expired date
        cancellation_date -- Optional cancellation date
        """
        explanation = ""
        for date, date_type in [(expired_date, "expiration date"),
                                (cancellation_date, "cancellation date")]:
            if date is None:
                continue
            if explanation:
                explanation += ", "
            explanation += f"its {date_type} is {date}"
        super().__init__(
            category="Expired or cancelled active license",
            sort_key=callsign,
            rm=(f"'HD' table marks '{callsign}' as active "
                f"('{license_status}'), but {explanation}"))


class ComplaintBadRxGain(Complainer.ComplaintBase):
    """ Complaint on missing antenna height """
    def __init__(self, callsign: str, path_num: int, rx_gain: Optional[float],
                 drop: bool) -> None:
        """ Constructor

        Arguments:
        callsign -- Callsign
        path_num -- Path number
        gain     -- RX gain or None
        drop     -- True if path will be dropped
        """
        if rx_gain is None:
            super().__init__(
                category="Receiver antenna gain is undefined",
                sort_key=(callsign, path_num),
                rm=(f"Path {callsign}/{path_num} has undefined RX gain"
                    f"{'. Path dropped' if drop else ''}"))
        else:
            super().__init__(
                category=f"Receiver antenna gain not in "
                f"[{RX_GAIN_RANGE[0]} - {RX_GAIN_RANGE[1]}] range",
                sort_key=(callsign, path_num),
                rm=(f"Path {callsign}/{path_num} has RX antenna gain of "
                    f"{rx_gain}, which is out of "
                    f"[{RX_GAIN_RANGE[0]} - {RX_GAIN_RANGE[1]}] range"
                    f"{'. Path dropped' if drop else ''}"))


class ComplaintInconsistentRepeater(Complainer.ComplaintBase):
    """ Complaint on repeater with inconsistent parameters """
    def __init__(self, callsign: str, path_num: int, repeater_idx: int) \
            -> None:
        """ Constructor

        Arguments:
        callsign     -- Callsign
        path_num     -- Path number
        repeater_idx -- 1-based repeater index
        """
        super().__init__(
            category="Repeater has different RX and TX antenna parameters " +
            "(not ULS error, but AFC database format limitation)",
            sort_key=(callsign, path_num, repeater_idx),
            rm=(f"Path {callsign}/{path_num}, repeater {repeater_idx} has "
                f"different antenna parameters on RX and TX sides"))


class GeoCalc:
    """ Simplified geodetic computations """

    # Approximate number of miles per latitude degree
    MILES_PER_DEGREE = 69

    @classmethod
    def segment_within_range(
            cls, center_lat: float, center_lon: float, range_miles: float,
            p1_lat: float, p1_lon: float, p2_lat: float, p2_lon: float) \
            -> bool:
        """ Checks if given segment intersects with or lies within given circle

        Arguments:
        center_lat   -- Latitude of center in degrees
        center_lon   -- Longitude of center in degrees
        radius_miles -- Radius in miles
        p1_lat       -- Latitude of segment begin in degrees
        p1_lon       -- Longitude of segment begin in degrees
        p2_lat       -- Latitude of segment end in degrees
        p2_lon       -- Longitude of segment end in degrees
        True if intersects
        """
        # Switch to Cartesian coordinates in mile units ...
        aspect = math.cos(math.radians(center_lat))
        center_x = center_lon * cls.MILES_PER_DEGREE * aspect
        center_y = center_lat * cls.MILES_PER_DEGREE
        # ... and also move center to (0, 0)
        x1 = p1_lon * cls.MILES_PER_DEGREE * aspect - center_x
        y1 = p1_lat * cls.MILES_PER_DEGREE - center_y
        x2 = p2_lon * cls.MILES_PER_DEGREE * aspect - center_x
        y2 = p2_lat * cls.MILES_PER_DEGREE - center_y

        # Below is a black magic. Writing segment in parametric form
        # (0 - (x1, y1), 1 - (x2, y2)), substituting to circle equation,
        # finding solutions of resulting quadratic equation

        dx = x2 - x1
        dy = y2 - y1

        # Coefficients of quadratic equation
        a = dx * dx + dy * dy
        b = 2 * (x1 * dx + y1 * dy)
        c = x1 * x1 + y1 * y1 - range_miles * range_miles

        # Corner case - points coincide, this is not a parabola)
        if a == 0:
            return (x1 * x1 + y2 * y2) <= (range_miles * range_miles)

        # This is horns-up (a>0) parabola, vertex at -b/2a. Looking for roots
        # such that segment between them intersect [0, 1]. Roots might be
        # graphically analyzed without solving the equation

        if -b <= 0:
            # Vertex to the left of 0. Looking for root to the right of 0 (as
            # if equation solvable, it has root < 0 for sure). There is on if
            # parabola crosses ordinates axis at or below abscissas axis
            return c <= 0

        if -b < (2 * a):
            # Vertex between 0 and 1. Equation must have roots - checking
            # discriminant
            return (b * b - 4 * a * c) >= 0

        # Vertex to the right of 1. If solvable - there is one root to the
        # right of 1. To have root to the left of 1 parabola must intersect
        # x=1 below abscissas axis
        return (a + b + c) <= 0

    @classmethod
    def circle(cls, center_lat: float, center_lon: float, radius_miles: float,
               num_vertex: int, close: bool = False) -> Iterator:
        """ Polygonal approximation of circle

        Arguments:
        center_lat   -- Center latitude in degrees
        center_lon   -- Center longitude in degrees
        radius_miles -- Radius in miles
        num_vertex   -- Number of vertexes in polygon
        close        -- True to yield starting point also as the last
        Returns sequence of (lat, lon) tuples of required polygon """
        # Local scale factors between geodetic and Cartesian coordinates
        lon_scale = cls.MILES_PER_DEGREE * math.cos(math.radians(center_lat))
        lat_scale = cls.MILES_PER_DEGREE
        angle_step = 2 * math.pi / num_vertex
        for i in range(num_vertex + (1 if close else 0)):
            yield (center_lat +
                   radius_miles * math.sin(i * angle_step) / lat_scale,
                   center_lon +
                   radius_miles * math.cos(i * angle_step) / lon_scale)

    @classmethod
    def arrow_head(cls, begin_lat: float, begin_lon: float, end_lat: float,
                   end_lon: float, angle: float, size_miles: float) \
            -> Tuple[float, float, float, float]:
        """ Computes coordinates of the ends of arrow head

        Arguments:
        begin_lat  -- Arrow stem begin latitude
        begin_lon  -- Arrow stem begin longitude
        end_lat    -- Arrow stem end latitude
        end_lon    -- Arrow stem end longitude
        angle      -- Angle between arrowhead side and stem - in degrees
        size_miles -- Size of arrowhead side in degrees
        Returns (lat1, lon1, lat2, lon2) tuple containing coordinates of ends
        of arrow head
        """
        lon_scale = cls.MILES_PER_DEGREE * math.cos(math.radians(end_lat))
        lat_scale = cls.MILES_PER_DEGREE
        # Direction is from end to begin
        direction = \
            math.atan2((begin_lat - end_lat) * lat_scale,
                       (begin_lon - end_lon) * lon_scale)
        ang_rad = math.radians(angle)
        return \
            (end_lat + size_miles * math.sin(direction + ang_rad) / lat_scale,
             end_lon + size_miles * math.cos(direction + ang_rad) / lon_scale,
             end_lat + size_miles * math.sin(direction - ang_rad) / lat_scale,
             end_lon + size_miles * math.cos(direction - ang_rad) / lon_scale)


class PathInfo:
    """ Full set of path data

    Attributes:
    callsign          -- Callsign
    path_num          -- Path number or None
    status            -- License status
    radio_service     -- Radio service
    name              -- Entity name or None
    common_carrier    -- Common carrier - bool or None
    rx_callsign       -- RX callsign
    rx_antenna_num    -- RX antenna number or None
    freq_range        -- FreqRange object
    tx_eirp           -- Maximum TX EIRP or None
    tx_ant            -- AntCoord for transmitter
    tx_polarization   -- TX antenna polarization or None
    tx_gain           -- TX antenna gain
    rx_ant            -- AntCoord for receiver
    rx_ant_model      -- RX antenna model name or None
    rx_gain           -- RX gain or None
    repeaters         -- List of Repeater objects
    grant_date        -- License grant date - date or None
    expired_date      -- License expired date - date or None
    cancellation_date -- License cancellation date - date or None
    p_rx_indicator    -- Passive receiver indicator (from PA)
    fsid              -- 'fsid' field from AFC database or None
    """
    class AntCoord:
        """ Full information about RX/TX/repeater antenna location

        Attributes:
        lat                     -- Latitude in signed degrees or None
        lon                     -- Longitude in signed degrees or None
        ground_elev_m           -- Ground elevation in meters or None
        height_to_center_raat_m -- Height of antenna center above ground in
                                   meters or None
        loc_type_code           -- Location type_code from LO
        """
        def __init__(
                self, lat: Optional[float] = None, lon: Optional[float] = None,
                lat_deg: Optional[int] = None, lat_min: Optional[int] = None,
                lat_sec: Optional[float] = None, lat_dir: Optional[str] = None,
                lon_deg: Optional[int] = None, lon_min: Optional[int] = None,
                lon_sec: Optional[float] = None, lon_dir: Optional[str] = None,
                ground_elev_m: Optional[float] = None,
                height_to_center_raat_m: Optional[float] = None,
                loc_type_code: Optional[str] = None) -> None:
            """ Constructor (constructed either from full latitude/longitude,
            or from degrees/minutes/seconds)

            Arguments:
            lat                     -- Latitude in degrees - float or None
            lon                     -- Longitude in degrees - float or None
            lat_deg                 -- Degree part of latitude or None
            lat_min                 -- Minute part of latitude or None
            lat_sec                 -- Second part of latitude - float or None
            lat_dir                 -- Latitude direction - N, S or None
            lon_deg                 -- Degree part of longitude or None
            lon_min                 -- Minute part of longitude or None
            lon_sec                 -- Second part of longitude - float or None
            lon_dir                 -- Longitude direction - E, W or None
            ground_elev_m           -- Ground elevation in meters or None
            height_to_center_raat_m -- Height of antenna center above ground in
                                       meters or None
            loc_type_code           -- Location type code from LO
            """
            self.lat: Optional[float] = None
            self.lon: Optional[float] = None
            if (lat is not None) and (lon is not None):
                self.lat = lat
                self.lon = lon
            elif (lat_deg is not None) and (lat_min is not None) and \
                    (lat_sec is not None) and (lat_dir is not None) and \
                    (lon_deg is not None) and (lon_min is not None) and \
                    (lon_sec is not None) and (lon_dir is not None):
                for d, m, s, di, attr in \
                        [(lat_deg, lat_min, lat_sec, lat_dir, "lat"),
                         (lon_deg, lon_min, lon_sec, lon_dir, "lon")]:
                    setattr(
                        self, attr,
                        (d + m / 60 + s / 3600) * (1 if di in "NE" else -1))
            self.ground_elev_m = ground_elev_m
            self.height_to_center_raat_m = height_to_center_raat_m
            self.loc_type_code = loc_type_code

        def __bool__(self) -> bool:
            """ True if coordinates and antenna center height above ground
            defined """
            return all(v is not None for v in
                       [self.lat, self.lon, self.height_to_center_raat_m])

        def __hash__(self) -> int:
            """ Hash over essential parts of coordinates """
            ret = 0
            for value, digits in [(self.lat, 6), (self.lon, 6),
                                  # (self.ground_elev_m, 1),
                                  (self.height_to_center_raat_m, 1)]:
                if value is not None:
                    ret += hash(round(value, digits))
            return ret

        def __eq__(self, other):
            """ Equality comparison """
            return isinstance(other, self.__class__) and \
                all(getattr(self, attr) == getattr(other, attr)
                    for attr in self.__dict__)

        @classmethod
        def combine(cls, v1: Optional["PathInfo.AntCoord"],
                    v2: Optional["PathInfo.AntCoord"]) \
                -> Tuple[bool, Optional["PathInfo.AntCoord"]]:
            """ Combine two presumably equal values into one

            Arguments:
            v1 -- First value to combine (may be None)
            v2 -- Second value to combine (may be None)
            Returns (consistent, combination) value. 'consistent' is true if
            values do not have different non-None atributes. 'combination' is
            None if both components are None, otherwise value made of non-None
            attributes of both values
            """
            if v1 is None:
                return (True, v2)
            if v2 is None:
                return (True, v1)
            if v1 == v2:
                return (True, v1)
            if v1 == v2:
                return (True, v1)
            consistent = True
            ret = copy.copy(v1)
            for attr in ret.__dict__:
                a1 = getattr(v1, attr)
                a2 = getattr(v2, attr)
                if a1 is not None:
                    if (a2 is not None) and (a2 != a1):
                        consistent = False
                elif a2 is not None:
                    setattr(ret, attr, a2)
            return (consistent, ret)

    class Repeater:
        """ Information about repeater

        Attributes:
        ant_coord            -- AntCoord or None
        reflector_height     -- Reflector height or None
        reflector_width      -- Reflector width or None
        back_to_back_gain_tx -- Back to back gain TX or None
        back_to_back_gain_rx -- Back to back gain RX or None
        """
        def __init__(self, ant_coord: Optional["PathInfo.AntCoord"],
                     reflector_height: Optional[float],
                     reflector_width: Optional[float],
                     back_to_back_gain_tx: Optional[float],
                     back_to_back_gain_rx: Optional[float]) -> None:
            """ Constructor

            Arguments:
            ant_coord            -- AntCoord or None
            reflector_height     -- Reflector height or None
            reflector_width      -- Reflector width or None
            back_to_back_gain_tx -- Back to back gain TX or None
            back_to_back_gain_rx -- Back to back gain RX or None
            """
            self.ant_coord = ant_coord
            self.reflector_height = reflector_height
            self.reflector_width = reflector_width
            self.back_to_back_gain_tx = back_to_back_gain_tx
            self.back_to_back_gain_rx = back_to_back_gain_rx

        def __bool__(self) -> bool:
            """ True if antenna coordinates defined """
            return bool(self.ant_coord)

        def __hash__(self) -> int:
            """ Hash over antenna coordinates """
            return hash(self.ant_coord)

        def __eq__(self, other) -> bool:
            """ Equality comparison """
            return isinstance(other, self.__class__) and \
                all((getattr(self, attr) == getattr(other, attr))
                    for attr in self.__dict__)

        @classmethod
        def combine(cls, v1: Optional["PathInfo.Repeater"],
                    v2: Optional["PathInfo.Repeater"]) \
                -> Tuple[bool, Optional["PathInfo.Repeater"]]:
            """ Combine two optional Repeater objects into one

            Arguments:
            v1 -- First combined objects
            v2 -- Second combined object
            Return (consistent, combination) tuple. 'consistent' is True if
            combined objects do not have non-None attributes of different
            values. 'combined' is object, combined of non-None attributes of
            source objects
            """
            if v1 is None:
                return (True, v2)
            if v2 is None:
                return (True, v1)
            if v1 == v2:
                return (True, v1)
            ret = copy.deepcopy(v1)
            consistent = True
            for attr in ret.__dict__:
                a1 = getattr(ret, attr)
                a2 = getattr(v2, attr)
                if (a1 != a2) and (a2 is not None):
                    if a1 is None:
                        setattr(ret, attr, a2)
                    elif isinstance(a1, float):
                        consistent = False
                    elif isinstance(a1, PathInfo.AntCoord):
                        c, v = PathInfo.AntCoord.combine(a1, a2)
                        if c:
                            setattr(ret, attr, v)
                        else:
                            consistent = False
                    else:
                        error("Internal error. Unsupported attribute type")
            return (consistent, ret)

    def __init__(self, callsign: str, status: str, radio_service: str,
                 name: Optional[str],
                 common_carrier: Optional[Union[bool, str]],
                 mobile: Optional[Union[bool, str]],
                 rx_callsign: Optional[str], rx_antenna_num: int,
                 freq_range: FreqRange, tx_eirp: Optional[float],
                 tx_ant: "PathInfo.AntCoord", tx_polarization: Optional[str],
                 tx_gain: Optional[float], rx_ant: "PathInfo.AntCoord",
                 rx_ant_model: Optional[str], rx_gain: Optional[float],
                 repeaters: List["PathInfo.Repeater"],
                 p_rx_indicator: Optional[str], path_num: Optional[int] = None,
                 grant_date: Optional[Union[str, datetime.date]] = None,
                 expired_date: Optional[Union[str, datetime.date]] = None,
                 cancellation_date: Optional[Union[str, datetime.date]] = None,
                 fsid: Optional[int] = None) -> None:
        """ Constructor

        Arguments:
        callsign          -- Path callsign
        status            -- License status
        radio_service     -- Radio service
        name              -- Licensed entity name or None
        common_carrier    -- Common carrier - bool or None
        mobile            -- Mobile - bool or None
        rx_callsign       -- RX callsign or None (in latter case path callsign
                             is used)
        rx_antenna_num    -- RX antenna number
        freq_range        -- FreqRange object
        tx_eirp           -- Maximum TX EIRP
        tx_ant            -- AntCoord for receiver
        tx_polarization   -- TX antenna polarization or None
        tx_gain           -- TX gain or None
        rx_ant            -- AntCoord for receiver
        rx_ant_model      -- RX antenna model or None
        rx_gain           -- RX gain or None
        repeaters         -- List or Repeater objects
        p_rx_indicator    -- Passive receiver indicator (from PA)
        path_num          -- Path number or None
        grant_date        -- License grant date or None
        expired_date      -- License expired date or None
        cancellation_date -- License expired date or None
        fsid              -- 'fsid' field from AFC database or None
        """
        self.callsign = callsign
        self.status = status
        self.radio_service = radio_service
        self.name = name
        self.common_carrier = self._yn_to_bool(common_carrier)
        self.mobile = self._yn_to_bool(mobile)
        self.rx_callsign = rx_callsign
        self.rx_antenna_num = rx_antenna_num
        self.freq_range = freq_range
        self.tx_eirp = tx_eirp
        self.tx_ant = tx_ant
        self.tx_polarization = tx_polarization
        self.tx_gain = tx_gain
        self.rx_ant = rx_ant
        self.rx_ant_model = rx_ant_model
        self.rx_gain = rx_gain
        self.repeaters = repeaters
        self.p_rx_indicator = p_rx_indicator
        self.path_num = path_num
        self.fsid = fsid

        make_optional_date: Callable[[Optional[Union[datetime.date, str]]],
                                     Optional[datetime.date]] = \
            lambda d: d if (d is None) or isinstance(d, datetime.date) \
            else datetime.datetime.strptime(d, "%m/%d/%Y").date()
        self.grant_date = make_optional_date(grant_date)
        self.expired_date = make_optional_date(expired_date)
        self.cancellation_date = make_optional_date(cancellation_date)

    def is_afc(self) -> bool:
        """ True if path fit for AFC database: is 6GHz, license active """
        return (self.status in "AL") and self.freq_range.is_6ghz()

    def is_expired(self) -> bool:
        """ True for expired active path """
        return (self.status == "A") and \
            any(((d is not None) and (d < datetime.datetime.now().date()))
                for d in (self.expired_date, self.cancellation_date))

    def path_hash(self, with_path_num: bool = False,
                  with_all_repeaters: bool = True) -> int:
        """ Hash over essential parts of path

        Arguments:
        with_path_num      -- Include path number in hash computation
        with_all_repeaters -- True to include all repeaters, False to include
                              first repeater only (if any)
        Returns hash over essential parts of path: callsign, coordinates of
        receiver and transmitter antennas, coordinates of first repeater's
        antennas, frequency range, possibly path number, possibly antennas of
        remaining repeaters
        """
        return \
            hash(tuple([hash(v) for v in
                 ([self.callsign, self.tx_ant, self.rx_ant, self.freq_range] +
                  ([self.path_num] if with_path_num else []) +
                  self.repeaters[: None if with_all_repeaters else 1])]))

    def path_name(self) -> str:
        """ Returns path name (callsign and path number if available) """
        return self.callsign if self.path_num is None \
            else f"{self.callsign}/{self.path_num}"

    def within_range(self, center_lat: float, center_lon: float,
                     range_miles: float) -> bool:
        """ True if path is within given range of given point """
        within: Callable[["PathInfo.AntCoord", "PathInfo.AntCoord"], bool] = \
            lambda ac1, ac2: \
            (ac1.lat is not None) and (ac1.lon is not None) and \
            (ac2.lat is not None) and (ac2.lon is not None) and \
            GeoCalc.segment_within_range(
                center_lat=center_lat, center_lon=center_lon,
                range_miles=range_miles, p1_lat=ac1.lat, p1_lon=ac1.lon,
                p2_lat=ac2.lat, p2_lon=ac2.lon)
        ac = self.tx_ant
        for repeater in self.repeaters:
            assert repeater.ant_coord is not None
            if within(ac, repeater.ant_coord):
                return True
            ac = repeater.ant_coord
        return within(ac, self.rx_ant)

    def __bool__(self) -> bool:
        """ True if all coordinates and heights are defined """
        return all(bool(c)
                   for c in ([self.tx_ant, self.rx_ant] + self.repeaters))

    def _yn_to_bool(self, v: Optional[Union[bool, str]]) -> Optional[bool]:
        """ Converts optional Y/N string to optional bool """
        if (v is None) or isinstance(v, bool):
            return v
        assert isinstance(v, str)
        return v.lower() == "y"


class PathFreqBucket:
    """ Collection of variants of same path with different frequency ranges
    reduces paths with same frequency range

    Private attributes:
    _paths -- List of PathInfo objects
    """
    def __init__(self) -> None:
        """ Constructor """
        self._paths: List[PathInfo] = []

    def try_add(self, path: PathInfo) -> bool:
        """ Tries to add given path to collection

        Arguments:
        path -- PathInfo object to add
        Returns False on failure (bucket already contains paths with other
        callsign/path number) """
        if self._paths and \
                ((self._paths[0].callsign != path.callsign) or
                 (self._paths[0].path_num != path.path_num)):
            return False
        # idx = 0
        # while idx < len(self._paths):
        #     if path.freq_range.contains(self._paths[idx].freq_range):
        #         self._paths.pop(idx)
        #     elif self._paths[idx].freq_range.contains(path.freq_range):
        #         return True
        #     idx += 1
        if not any((p.freq_range == path.freq_range) for p in self._paths):
            self._paths.append(path)
        return True

    def __iter__(self) -> Iterator:
        """ Iterator over contained ranges """
        for path in sorted(self._paths,
                           key=operator.attrgetter("freq_range.start_mhz")):
            yield path

    def clear(self) -> None:
        """ Clears the bucket """
        self._paths.clear()


class PathValidator:
    """ Path validator
    Also does most of complaining, yet certain complaints are in scattered over
    code for performance reasons

    Private attributes:
    _pass_all          -- True to allow all paths without checks
    _allow_low_antenna -- True to allow low antennas
    _allow_bad_gain    -- True to allow undefined or out of range RX gain
    _expired_callsigns -- Set of callsigns with reported expiration
    """
    def __init__(self, pass_all: bool = False,
                 _allow_low_antenna: bool = False,
                 allow_bad_gain: bool = False) -> None:
        """ Constructor

        Arguments:
        pass_all          -- True to allow all paths without checks
        allow_low_antenna -- True to allow low antennas
        allow_bad_gain    -- True to allow undefined or out of range RX gain
        expired_callsigns -- Set of callsigns with reported expiration
        """
        self._pass_all = pass_all
        self._allow_low_antenna = _allow_low_antenna
        self._allow_bad_gain = allow_bad_gain
        self._expired_callsigns: Set[str] = set()

    def validate(self, path: PathInfo) -> bool:
        """ Validate given path for inclusion into AFC database, prints what's
        wrong with path

        Arguments:
        path -- PathInfo object to check
        Returns True if path is eligible
        """
        if self._pass_all:
            return True
        if (path.status not in "AL") or (not path.freq_range.is_6ghz()):
            return False
        drop = False
        if (path.p_rx_indicator == "Y") != (len(path.repeaters) > 0):
            ComplaintWrongRepeaterMark(
                callsign=path.callsign, path_num=mi(path.path_num),
                has_repeaters=len(path.repeaters) > 0)
        for ac, role in [(path.tx_ant, "transmitter"),
                         (path.rx_ant, "receiver")] + \
                [(r.ant_coord, f"repeater {idx + 1}")
                 for idx, r in enumerate(path.repeaters)]:
            if (ac.lat is None) or (ac.lon is None):
                ComplaintMissingCoordinates(
                    callsign=path.callsign, path_num=mi(path.path_num),
                    role=role)
                drop = True
            if ac.height_to_center_raat_m is None:
                ComplaintMissingAntennaHeight(
                    callsign=path.callsign, path_num=mi(path.path_num),
                    role=role)
                drop = True
            elif mf(ac.height_to_center_raat_m) < MIN_ANT_HEIGHT_M:
                ComplaintLowAntennaHeight(
                    callsign=path.callsign, path_num=mi(path.path_num),
                    role=role, raat_height_m=ac.height_to_center_raat_m,
                    drop=not self._allow_low_antenna)
                if not self._allow_low_antenna:
                    drop = True
        if (path.rx_gain is None) or \
                (not (RX_GAIN_RANGE[0] <= mf(path.rx_gain) <=
                      RX_GAIN_RANGE[1])):
            ComplaintBadRxGain(
                callsign=path.callsign, path_num=mi(path.path_num),
                rx_gain=path.rx_gain, drop=not self._allow_bad_gain)
            if not self._allow_bad_gain:
                drop = True
        if path.is_expired() and \
                (path.callsign not in self._expired_callsigns):
            ComplaintExpiredActiveLicense(
                callsign=path.callsign, license_status=path.status,
                expired_date=path.expired_date,
                cancellation_date=path.cancellation_date)
            self._expired_callsigns.add(path.callsign)
        return not drop


class AfcDb:
    """ Information about AFC database """
    class Format(enum.Enum):
        """ AFC database format variations and their properties

        Attributes:
        has_path_num       -- Has path_number field
        multirepeater      -- Supports more than one repeater
        has_repeater_hwg   -- Has information about repeater antenna height,
                              width, gain
        separate_repeaters -- Repeaters in separate table
        """
        ORIGINAL: Dict[Any, Any] = {}
        WITH_PATH_NUM = {"has_path_num": True}
        MULTIREPEATER = {"has_path_num": True, "multirepeater": True}
        SEP_REP = {"has_path_num": True, "multirepeater": True,
                   "separate_repeaters": True}

        def __init__(self, args: Dict[str, bool]) -> None:
            """ Constructor

            Arguments:
            args -- Dictionary, assigned to enum value """
            self.has_path_num: bool = args.get("has_path_num", False)
            self.multirepeater: bool = args.get("multirepeater", False)
            self.has_repeater_hwg: bool = args.get("multirepeater", False)
            self.separate_repeaters: bool = \
                args.get("separate_repeaters", False)

        @classmethod
        def latest(cls) -> "AfcDb.Format":
            """ Returns latest table format """
            return list(cls)[-1]

    class ColInfo:
        """ AFC database column descriptor

        Attributes:
        col_arg -- Arguments for sa.Column constructor
        name    -- Database column name
        formats -- None or list of formats to which column can be written
        """
        def __init__(self, col_arg: ArgContainer,
                     formats: Optional[List["AfcDb.Format"]] = None) -> None:
            """ Constructor

            Arguments:
            col_arg -- Arguments for sa.Column constructor
            formats -- None or list of formats to which column can be written
            """
            self.col_arg = col_arg
            self.formats: Optional[Set["AfcDb.Format"]] = \
                set(formats) if formats else None
            self.name: str = mst(col_arg[0])

    # Path table name for AFC database
    PATH_TABLE_NAME = "uls"

    # Repeater table ame for AFC database
    REP_TABLE_NAME = "pr"

    # Maximum number of repeaters
    MAX_REPEATERS = 3

    # AFC database column descriptors
    _COLUMNS: Dict[str, List["AfcDb.ColInfo"]] = {
        PATH_TABLE_NAME: [
            ColInfo(ArgContainer("id", sa.Integer(), nullable=False,
                                 primary_key=True),
                    formats=[Format.ORIGINAL, Format.WITH_PATH_NUM,
                             Format.MULTIREPEATER]),
            ColInfo(ArgContainer("fsid", sa.Integer(), nullable=False,
                                 index=True)),
            ColInfo(ArgContainer("callsign", sa.Unicode(16), nullable=False)),
            ColInfo(ArgContainer("status", sa.Unicode(1), nullable=False)),
            ColInfo(ArgContainer("radio_service", sa.Unicode(4),
                                 nullable=False)),
            ColInfo(ArgContainer("name", sa.Unicode(256), nullable=False)),
            ColInfo(ArgContainer("common_carrier", sa.Boolean(),
                                 nullable=True)),
            ColInfo(ArgContainer("mobile", sa.Boolean(), nullable=True)),
            ColInfo(ArgContainer("rx_callsign", sa.Unicode(16),
                                 nullable=False)),
            ColInfo(ArgContainer("rx_antenna_num", sa.Integer(),
                                 nullable=False)),
            ColInfo(ArgContainer("freq_assigned_start_mhz", sa.Float(),
                                 nullable=False)),
            ColInfo(ArgContainer("freq_assigned_end_mhz", sa.Float(),
                                 nullable=False)),
            ColInfo(ArgContainer("tx_eirp", sa.Float(), nullable=True)),
            ColInfo(ArgContainer("emissions_des", sa.Unicode(64),
                                 nullable=False)),
            ColInfo(ArgContainer("tx_lat_deg", sa.Float(), nullable=False,
                                 index=True)),
            ColInfo(ArgContainer("tx_long_deg", sa.Float(), nullable=False,
                                 index=True)),
            ColInfo(ArgContainer("tx_ground_elev_m", sa.Float(),
                                 nullable=True)),
            ColInfo(ArgContainer("tx_polarization", sa.Unicode(5),
                                 nullable=True)),
            ColInfo(ArgContainer("tx_height_to_center_raat_m", sa.Float(),
                                 nullable=True)),
            ColInfo(ArgContainer("tx_gain", sa.Float(), nullable=True)),
            ColInfo(ArgContainer("rx_lat_deg", sa.Float(), nullable=False,
                                 index=True)),
            ColInfo(ArgContainer("rx_long_deg", sa.Float(), nullable=False,
                                 index=True)),
            ColInfo(ArgContainer("rx_ground_elev_m", sa.Float(),
                                 nullable=True)),
            ColInfo(ArgContainer("rx_ant_model", sa.Unicode(64),
                                 nullable=True)),
            ColInfo(ArgContainer("rx_height_to_center_raat_m", sa.Float(),
                                 nullable=True)),
            ColInfo(ArgContainer("rx_gain", sa.Float(), nullable=True)),
            ColInfo(ArgContainer("p_rx_indicator", sa.Unicode(1),
                                 nullable=True),
                    formats=[Format.ORIGINAL, Format.WITH_PATH_NUM,
                             Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_num", sa.Integer(), nullable=False),
                    formats=[Format.MULTIREPEATER, Format.SEP_REP]),
            ColInfo(ArgContainer("p_rp_lat_degs", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.ORIGINAL, Format.WITH_PATH_NUM]),
            ColInfo(ArgContainer("p_rp_lon_degs", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.ORIGINAL, Format.WITH_PATH_NUM]),
            ColInfo(ArgContainer("p_rp_height_to_center_raat_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.ORIGINAL, Format.WITH_PATH_NUM]),
            ColInfo(ArgContainer("p_rp_1_lat_degs", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_1_lon_degs", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_1_height_to_center_raat_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_1_reflector_height_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_1_reflector_width_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_1_back_to_back_gain_tx", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_1_back_to_back_gain_rx", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_2_lat_degs", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_2_lon_degs", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_2_height_to_center_raat_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_2_reflector_height_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_2_reflector_width_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_2_back_to_back_gain_tx", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_2_back_to_back_gain_rx", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_3_lat_degs", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_3_lon_degs", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_3_height_to_center_raat_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_3_reflector_height_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_3_reflector_width_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_3_back_to_back_gain_tx", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("p_rp_3_back_to_back_gain_rx", sa.Float(),
                                 nullable=True),
                    formats=[Format.MULTIREPEATER]),
            ColInfo(ArgContainer("path_number", sa.Integer(), nullable=False),
                    formats=[Format.WITH_PATH_NUM, Format.MULTIREPEATER,
                             Format.SEP_REP])],
        REP_TABLE_NAME: [
            ColInfo(ArgContainer("id", sa.Integer(), primary_key=True,
                                 nullable=False),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("fsid", sa.Integer(), nullable=False,
                                 index=True),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("prSeq", sa.Integer(), nullable=False),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("pr_lat_deg", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("pr_lon_deg", sa.Float(), nullable=True,
                                 index=True),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("pr_height_to_center_raat_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("pr_reflector_height_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("pr_reflector_width_m", sa.Float(),
                                 nullable=True),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("pr_back_to_back_gain_tx", sa.Float(),
                                 nullable=True),
                    formats=[Format.SEP_REP]),
            ColInfo(ArgContainer("pr_back_to_back_gain_rx", sa.Float(),
                                 nullable=True),
                    formats=[Format.SEP_REP])]}

    @classmethod
    def columns(cls, db_format: Optional["AfcDb.Format"], table_name: str) -> \
            Iterator:
        """ Generates sequence of columns

        Arguments:
        db_format  -- None or one of AfcDb.Format members (in this case only
                      columns pertinent to this format should be included
        table_name -- cls.PATH_TABLE_NAME or cls.REP_TABLE_NAME
        Return Sequence of AfcDb.ColInfo
        """
        for column in cls._COLUMNS[table_name]:
            if (db_format is not None) and (column.formats is not None) and \
                    (db_format not in column.formats):
                continue
            yield column

    @classmethod
    def is_afc_db(cls, metadata: sa.MetaData) -> bool:
        """ True if given metadata is of AFC database """
        return cls.PATH_TABLE_NAME in metadata

    @classmethod
    def db_format(cls, metadata: sa.MetaData) -> "AfcDb.Format":
        """ Format of AFC database of given metadata """
        if cls.REP_TABLE_NAME in metadata:
            return cls.Format.SEP_REP
        table = metadata.tables[cls.PATH_TABLE_NAME]
        if not hasattr(table.c, "path_number"):
            return cls.Format.ORIGINAL
        if not hasattr(table.c, "p_rp_1_lat_degs"):
            return cls.Format.WITH_PATH_NUM
        return cls.Format.MULTIREPEATER


class DbPathIterator:
    """ Iterator over paths stored in ULS or AFC database
    Must be used as context manager (i.e. with 'with' clause)

    Private attributes:
    _filename -- Database filename
    _metadata -- Database metadata
    _conn     -- Database connection
    """
    def __init__(self, db_filename: str) -> None:
        """ Constructor

        Arguments:
        db_filename -- SQLite database file name. Caller must ensure it exists
        """
        assert os.path.isfile(db_filename)
        self._filename = db_filename
        engine = sa.create_engine("sqlite:///" + db_filename)
        self._metadata = sa.MetaData()
        self._metadata.reflect(bind=engine)
        self._conn = engine.connect()

    def has_path_num(self) -> bool:
        """ True if data source has path numbers """
        return (not AfcDb.is_afc_db(self._metadata)) or \
            AfcDb.db_format(self._metadata).has_path_num

    def __enter__(self) -> "DbPathIterator":
        """ Context manager entry. Returns self """
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
        """ Context manager exit. Closes database connection """
        self._conn.close()

    def __iter__(self) -> Iterator:
        """ Iterate over paths in database """
        return self._afc_paths() if AfcDb.is_afc_db(self._metadata) \
            else self._uls_paths()

    def _afc_paths(self) -> Iterator:
        """ Iterate over AFC database """
        db_format = AfcDb.db_format(self._metadata)
        repeaters_by_fsid: Dict[int, List[PathInfo.Repeater]] = {}
        if db_format.separate_repeaters:
            pr = self._metadata.tables[AfcDb.REP_TABLE_NAME]
            s = sa.select([pr]).order_by(pr.c.fsid, pr.c.prSeq)
            for r in self._conn.execute(s):
                path_repeaters: List[PathInfo.Repeater] = \
                    repeaters_by_fsid.setdefault(r.fsid, [])
                repeater = \
                    self._get_afc_repeater(
                        record=r, prefix="pr", coordinate_suffix="",
                        hwg=db_format.has_repeater_hwg)
                error_if(repeater is None,
                         (f"AFC database '{os.path.basename(self._filename)}' "
                          f"is corrupted: path {r.fsid} has incomplete "
                          f"information about repeater {r.prSeq}"))
                assert repeater is not None
                path_repeaters.append(repeater)
                error_if(
                    r.prSeq != len(path_repeaters),
                    (f"AFC database '{os.path.basename(self._filename)}' is "
                     f"corrupted: path {r.fsid} has out of sequence repeater "
                     f"sequence number {r.prSeq}"))
        uls = self._metadata.tables[AfcDb.PATH_TABLE_NAME]
        s = sa.select([uls])
        row_num = 0
        for r in self._conn.execute(s):
            repeaters: List[PathInfo.Repeater] = []
            if db_format.separate_repeaters:
                repeaters = repeaters_by_fsid.get(r.fsid, [])
                error_if(
                    r.p_rp_num != len(repeaters),
                    (f"Path {r.callsign}/{r.path_number} (FSID {r.fsid}) of "
                     f"AFC database '{os.path.basename(self._filename)}' has "
                     f"inconsistent information about repeaters: {r.p_rp_num} "
                     f"repeaters according to {AfcDb.PATH_TABLE_NAME} but "
                     f"{len(repeaters)} according to {AfcDb.REP_TABLE_NAME} "
                     f"table"))
            elif db_format.multirepeater:
                for idx in range(1, AfcDb.MAX_REPEATERS + 1):
                    repeater = \
                        self._get_afc_repeater(
                            record=r, prefix=f"p_rp_{idx}",
                            coordinate_suffix="s",
                            hwg=db_format.has_repeater_hwg)
                    if not repeater:
                        break
                    repeaters.append(repeater)
            else:
                repeater = \
                    self._get_afc_repeater(
                        record=r, prefix="p_rp", coordinate_suffix="s",
                        hwg=db_format.has_repeater_hwg)
                if repeater:
                    repeaters.append(repeater)
            path = \
                PathInfo(
                    callsign=r.callsign, status=r.status,
                    radio_service=r.radio_service, name=r.name,
                    common_carrier=r.common_carrier,
                    mobile=r.mobile, rx_callsign=r.rx_callsign,
                    rx_antenna_num=r.rx_antenna_num,
                    freq_range=FreqRange(
                        emission_code=r.emissions_des,
                        start_mhz=r.freq_assigned_start_mhz,
                        end_mhz=r.freq_assigned_end_mhz),
                    tx_eirp=r.tx_eirp,
                    tx_ant=PathInfo.AntCoord(
                        lat=r.tx_lat_deg, lon=r.tx_long_deg,
                        ground_elev_m=r.tx_ground_elev_m,
                        height_to_center_raat_m=r.tx_height_to_center_raat_m),
                    tx_polarization=r.tx_polarization, tx_gain=r.tx_gain,
                    rx_ant=PathInfo.AntCoord(
                        lat=r.rx_lat_deg, lon=r.rx_long_deg,
                        ground_elev_m=r.rx_ground_elev_m,
                        height_to_center_raat_m=r.rx_height_to_center_raat_m),
                    rx_ant_model=r.rx_ant_model, rx_gain=r.rx_gain,
                    repeaters=repeaters,
                    p_rx_indicator=('Y' if repeaters else 'N')
                    if db_format.separate_repeaters
                    else r.p_rx_indicator,
                    fsid=r.fsid,
                    path_num=r.path_number if db_format.has_path_num else None)
            row_num += 1
            if (row_num % 10000) == 0:
                Progressor.print_progress(f"{row_num}")
            yield path
        Progressor.print_progress()

    def _get_afc_repeater(self, record: Any, prefix: str,
                          coordinate_suffix: str,
                          hwg: bool) -> Optional[PathInfo.Repeater]:
        """ Read repeater from record of AFC database

        Arguments:
        record            -- Record containing repeater data
        prefix            -- Prefix of field names
        coordinate_suffix -- Suffix of coordinate fields' names
        hwg               -- True if information about repeater antenna heifht,
                             width and gains present in the record
        Return Repeater object if it is present else None """
        lat = record[f"{prefix}_lat_deg{coordinate_suffix}"]
        lon = record[f"{prefix}_lon_deg{coordinate_suffix}"]
        raat = record[f"{prefix}_height_to_center_raat_m"]
        if (lat is None) or (lon is None) or (raat is None):
            return None
        return \
            PathInfo.Repeater(
                ant_coord=PathInfo.AntCoord(
                    lat=lat, lon=lon, height_to_center_raat_m=raat),
                reflector_height=record[
                    f"{prefix}_reflector_height_m"] if hwg else None,
                reflector_width=record[
                    f"{prefix}_reflector_width_m"] if hwg else None,
                back_to_back_gain_tx=record[
                    f"{prefix}_back_to_back_gain_tx"] if hwg else None,
                back_to_back_gain_rx=record[
                    f"{prefix}_back_to_back_gain_rx"] if hwg else None)

    def _uls_paths(self) -> Iterator:
        """ Iterate over ULS database """
        an = self._metadata.tables["AN"]
        an_rx = an.alias("an_rx")
        an_r1rx = an.alias("an_r1rx")
        an_r1tx = an.alias("an_r1tx")
        an_r2rx = an.alias("an_r2rx")
        an_r2tx = an.alias("an_r2tx")
        an_r3rx = an.alias("an_r3rx")
        an_r3tx = an.alias("an_r3tx")
        an_r4rx = an.alias("an_r4rx")
        an_r4tx = an.alias("an_r4tx")
        em = self._metadata.tables["EM"]
        en = self._metadata.tables["EN"]
        fr = self._metadata.tables["FR"]
        hd = self._metadata.tables["HD"]
        lo = self._metadata.tables["LO"]
        lo_rx = lo.alias("lo_rx")
        lo_r1rx = lo.alias("lo_r1rx")
        lo_r1tx = lo.alias("lo_r1tx")
        lo_r2rx = lo.alias("lo_r2rx")
        lo_r2tx = lo.alias("lo_r2tx")
        lo_r3rx = lo.alias("lo_r3rx")
        lo_r3tx = lo.alias("lo_r3tx")
        lo_r4rx = lo.alias("lo_r4rx")
        lo_r4tx = lo.alias("lo_r4tx")
        pa = self._metadata.tables["PA"]
        sg1 = self._metadata.tables["SG"]
        sg2 = sg1.alias("sg2")
        sg3 = sg1.alias("sg3")
        sg4 = sg1.alias("sg4")
        sg5 = sg1.alias("sg5")
        sg6 = sg1.alias("sg6")
        s = sa.select(
            [pa.c.call_sign, pa.c.path_number,
             hd.c.license_status, hd.c.radio_service_code, en.c.entity_name,
             hd.c.common_carrier, hd.c.mobile, hd.c.grant_date,
             hd.c.expired_date, hd.c.cancellation_date, pa.c.receiver_callsign,
             pa.c.receiver_antenna_number,
             fr.c.frequency_assigned, fr.c.frequency_upper_band,
             sa.func.max(fr.c.eirp),  em.c.emission_code,
             lo.c.lat_degrees, lo.c.lat_minutes, lo.c.lat_seconds,
             lo.c.lat_direction, lo.c.long_degrees, lo.c.long_minutes,
             lo.c.long_seconds, lo.c.long_direction, lo.c.ground_elevation,
             lo.c.location_type_code, an.c.polarization_code,
             an.c.height_to_center_raat, an.c.gain,
             lo_rx.c.lat_degrees, lo_rx.c.lat_minutes, lo_rx.c.lat_seconds,
             lo_rx.c.lat_direction, lo_rx.c.long_degrees, lo_rx.c.long_minutes,
             lo_rx.c.long_seconds, lo_rx.c.long_direction,
             lo_rx.c.ground_elevation, lo_rx.c.location_type_code,
             an_rx.c.antenna_model, an_rx.c.height_to_center_raat,
             an_rx.c.gain, pa.c.passive_receiver_indicator,

             lo_r1rx.c.lat_degrees, lo_r1rx.c.lat_minutes,
             lo_r1rx.c.lat_seconds, lo_r1rx.c.lat_direction,
             lo_r1rx.c.long_degrees, lo_r1rx.c.long_minutes,
             lo_r1rx.c.long_seconds, lo_r1rx.c.long_direction,
             lo_r1rx.c.ground_elevation, lo_r1rx.c.location_type_code,
             an_r1rx.c.height_to_center_raat, an_r1rx.c.reflector_height,
             an_r1rx.c.reflector_width, an_r1rx.c.back_to_back_tx_dish_gain,
             an_r1rx.c.back_to_back_rx_dish_gain,
             lo_r1tx.c.lat_degrees, lo_r1tx.c.lat_minutes,
             lo_r1tx.c.lat_seconds, lo_r1tx.c.lat_direction,
             lo_r1tx.c.long_degrees, lo_r1tx.c.long_minutes,
             lo_r1tx.c.long_seconds, lo_r1tx.c.long_direction,
             lo_r1tx.c.ground_elevation, lo_r1tx.c.location_type_code,
             an_r1tx.c.height_to_center_raat, an_r1tx.c.reflector_height,
             an_r1tx.c.reflector_width, an_r1tx.c.back_to_back_tx_dish_gain,
             an_r1tx.c.back_to_back_rx_dish_gain,

             lo_r2rx.c.lat_degrees, lo_r2rx.c.lat_minutes,
             lo_r2rx.c.lat_seconds, lo_r2rx.c.lat_direction,
             lo_r2rx.c.long_degrees, lo_r2rx.c.long_minutes,
             lo_r2rx.c.long_seconds, lo_r2rx.c.long_direction,
             lo_r2rx.c.ground_elevation, lo_r2rx.c.location_type_code,
             an_r2rx.c.height_to_center_raat, an_r2rx.c.reflector_height,
             an_r2rx.c.reflector_width, an_r2rx.c.back_to_back_tx_dish_gain,
             an_r2rx.c.back_to_back_rx_dish_gain,
             lo_r2tx.c.lat_degrees, lo_r2tx.c.lat_minutes,
             lo_r2tx.c.lat_seconds, lo_r2tx.c.lat_direction,
             lo_r2tx.c.long_degrees, lo_r2tx.c.long_minutes,
             lo_r2tx.c.long_seconds, lo_r2tx.c.long_direction,
             lo_r2tx.c.ground_elevation, lo_r2tx.c.location_type_code,
             an_r2tx.c.height_to_center_raat, an_r2tx.c.reflector_height,
             an_r2tx.c.reflector_width, an_r2tx.c.back_to_back_tx_dish_gain,
             an_r2tx.c.back_to_back_rx_dish_gain,

             lo_r3rx.c.lat_degrees, lo_r3rx.c.lat_minutes,
             lo_r3rx.c.lat_seconds, lo_r3rx.c.lat_direction,
             lo_r3rx.c.long_degrees, lo_r3rx.c.long_minutes,
             lo_r3rx.c.long_seconds, lo_r3rx.c.long_direction,
             lo_r3rx.c.ground_elevation, lo_r3rx.c.location_type_code,
             an_r3rx.c.height_to_center_raat, an_r3rx.c.reflector_height,
             an_r3rx.c.reflector_width, an_r3rx.c.back_to_back_tx_dish_gain,
             an_r3rx.c.back_to_back_rx_dish_gain,
             lo_r3tx.c.lat_degrees, lo_r3tx.c.lat_minutes,
             lo_r3tx.c.lat_seconds, lo_r3tx.c.lat_direction,
             lo_r3tx.c.long_degrees, lo_r3tx.c.long_minutes,
             lo_r3tx.c.long_seconds, lo_r3tx.c.long_direction,
             lo_r3tx.c.ground_elevation, lo_r3tx.c.location_type_code,
             an_r3tx.c.height_to_center_raat, an_r3tx.c.reflector_height,
             an_r3tx.c.reflector_width, an_r3tx.c.back_to_back_tx_dish_gain,
             an_r3tx.c.back_to_back_rx_dish_gain,

             lo_r4rx.c.lat_degrees, lo_r4rx.c.lat_minutes,
             lo_r4rx.c.lat_seconds, lo_r4rx.c.lat_direction,
             lo_r4rx.c.long_degrees, lo_r4rx.c.long_minutes,
             lo_r4rx.c.long_seconds, lo_r4rx.c.long_direction,
             lo_r4rx.c.ground_elevation, lo_r4rx.c.location_type_code,
             an_r4rx.c.height_to_center_raat, an_r4rx.c.reflector_height,
             an_r4rx.c.reflector_width, an_r4rx.c.back_to_back_tx_dish_gain,
             an_r4rx.c.back_to_back_rx_dish_gain,
             lo_r4tx.c.lat_degrees, lo_r4tx.c.lat_minutes,
             lo_r4tx.c.lat_seconds, lo_r4tx.c.lat_direction,
             lo_r4tx.c.long_degrees, lo_r4tx.c.long_minutes,
             lo_r4tx.c.long_seconds, lo_r4tx.c.long_direction,
             lo_r4tx.c.ground_elevation, lo_r4tx.c.location_type_code,
             an_r4tx.c.height_to_center_raat, an_r4tx.c.reflector_height,
             an_r4tx.c.reflector_width, an_r4tx.c.back_to_back_tx_dish_gain,
             an_r4tx.c.back_to_back_rx_dish_gain,

             sg6.c.call_sign]).select_from(
                pa.join(
                    en, (en.c.call_sign == pa.c.call_sign) &
                    (en.c.entity_type == "L"), isouter=True).
                join(hd, hd.c.call_sign == pa.c.call_sign, isouter=True).
                join(fr, (fr.c.call_sign == pa.c.call_sign) &
                     (fr.c.location_number == pa.c.transmit_location_number) &
                     (fr.c.antenna_number == pa.c.transmit_antenna_number),
                     isouter=True).
                join(em, (em.c.call_sign == pa.c.call_sign) &
                     (em.c.location_number == pa.c.transmit_location_number) &
                     (em.c.antenna_number == pa.c.transmit_antenna_number) &
                     (em.c.frequency_number == fr.c.freq_seq_id)).
                join(lo, (lo.c.call_sign == pa.c.call_sign) &
                     (lo.c.location_number == pa.c.transmit_location_number)).
                join(lo_rx, (lo_rx.c.call_sign == pa.c.call_sign) &
                     (lo_rx.c.location_number ==
                      pa.c.receiver_location_number)).
                join(an, (an.c.call_sign == pa.c.call_sign) &
                     (an.c.path_number == pa.c.path_number) &
                     (an.c.location_number == pa.c.transmit_location_number) &
                     (an.c.antenna_number == pa.c.transmit_antenna_number)).
                join(an_rx, (an_rx.c.call_sign == pa.c.call_sign) &
                     (an_rx.c.path_number == pa.c.path_number) &
                     (an_rx.c.location_number ==
                      pa.c.receiver_location_number) &
                     (an_rx.c.antenna_number == pa.c.receiver_antenna_number)).

                join(sg1, (sg1.c.call_sign == pa.c.call_sign) &
                     (sg1.c.path_number == pa.c.path_number) &
                     (sg1.c.segment_number == 1), isouter=True).

                join(lo_r1rx, (lo_r1rx.c.call_sign == pa.c.call_sign) &
                     (lo_r1rx.c.location_number == sg1.c.receiver_location),
                     isouter=True).
                join(an_r1rx, (an_r1rx.c.call_sign == pa.c.call_sign) &
                     (an_r1rx.c.path_number == pa.c.path_number) &
                     (an_r1rx.c.location_number == sg1.c.receiver_location) &
                     (an_r1rx.c.antenna_number == sg1.c.receiver_antenna),
                     isouter=True).
                join(sg2, (sg2.c.call_sign == pa.c.call_sign) &
                     (sg2.c.path_number == pa.c.path_number) &
                     (sg2.c.segment_number == 2), isouter=True).
                join(lo_r1tx, (lo_r1tx.c.call_sign == pa.c.call_sign) &
                     (lo_r1tx.c.location_number == sg2.c.transmit_location),
                     isouter=True).
                join(an_r1tx, (an_r1tx.c.call_sign == pa.c.call_sign) &
                     (an_r1tx.c.path_number == pa.c.path_number) &
                     (an_r1tx.c.location_number == sg2.c.transmit_location) &
                     (an_r1tx.c.antenna_number == sg2.c.transmit_antenna),
                     isouter=True).

                join(lo_r2rx, (lo_r2rx.c.call_sign == pa.c.call_sign) &
                     (lo_r2rx.c.location_number == sg2.c.receiver_location),
                     isouter=True).
                join(an_r2rx, (an_r2rx.c.call_sign == pa.c.call_sign) &
                     (an_r2rx.c.path_number == pa.c.path_number) &
                     (an_r2rx.c.location_number == sg2.c.receiver_location) &
                     (an_r2rx.c.antenna_number == sg2.c.receiver_antenna),
                     isouter=True).
                join(sg3, (sg3.c.call_sign == pa.c.call_sign) &
                     (sg3.c.path_number == pa.c.path_number) &
                     (sg3.c.segment_number == 3), isouter=True).
                join(lo_r2tx, (lo_r2tx.c.call_sign == pa.c.call_sign) &
                     (lo_r2tx.c.location_number == sg3.c.transmit_location),
                     isouter=True).
                join(an_r2tx, (an_r2tx.c.call_sign == pa.c.call_sign) &
                     (an_r2tx.c.path_number == pa.c.path_number) &
                     (an_r2tx.c.location_number == sg3.c.transmit_location) &
                     (an_r2tx.c.antenna_number == sg3.c.transmit_antenna),
                     isouter=True).

                join(lo_r3rx, (lo_r3rx.c.call_sign == pa.c.call_sign) &
                     (lo_r3rx.c.location_number == sg3.c.receiver_location),
                     isouter=True).
                join(an_r3rx, (an_r3rx.c.call_sign == pa.c.call_sign) &
                     (an_r3rx.c.path_number == pa.c.path_number) &
                     (an_r3rx.c.location_number == sg3.c.receiver_location) &
                     (an_r3rx.c.antenna_number == sg3.c.receiver_antenna),
                     isouter=True).
                join(sg4, (sg4.c.call_sign == pa.c.call_sign) &
                     (sg4.c.path_number == pa.c.path_number) &
                     (sg4.c.segment_number == 4), isouter=True).
                join(lo_r3tx, (lo_r3tx.c.call_sign == pa.c.call_sign) &
                     (lo_r3tx.c.location_number == sg4.c.transmit_location),
                     isouter=True).
                join(an_r3tx, (an_r3tx.c.call_sign == pa.c.call_sign) &
                     (an_r3tx.c.path_number == pa.c.path_number) &
                     (an_r3tx.c.location_number == sg4.c.transmit_location) &
                     (an_r3tx.c.antenna_number == sg4.c.transmit_antenna),
                     isouter=True).

                join(lo_r4rx, (lo_r4rx.c.call_sign == pa.c.call_sign) &
                     (lo_r4rx.c.location_number == sg4.c.receiver_location),
                     isouter=True).
                join(an_r4rx, (an_r4rx.c.call_sign == pa.c.call_sign) &
                     (an_r4rx.c.path_number == pa.c.path_number) &
                     (an_r4rx.c.location_number == sg4.c.receiver_location) &
                     (an_r4rx.c.antenna_number == sg4.c.receiver_antenna),
                     isouter=True).
                join(sg5, (sg5.c.call_sign == pa.c.call_sign) &
                     (sg5.c.path_number == pa.c.path_number) &
                     (sg5.c.segment_number == 5), isouter=True).
                join(lo_r4tx, (lo_r4tx.c.call_sign == pa.c.call_sign) &
                     (lo_r4tx.c.location_number == sg5.c.transmit_location),
                     isouter=True).
                join(an_r4tx, (an_r4tx.c.call_sign == pa.c.call_sign) &
                     (an_r4tx.c.path_number == pa.c.path_number) &
                     (an_r4tx.c.location_number == sg5.c.transmit_location) &
                     (an_r4tx.c.antenna_number == sg5.c.transmit_antenna),
                     isouter=True).

                join(sg6, (sg6.c.call_sign == pa.c.call_sign) &
                     (sg6.c.path_number == pa.c.path_number) &
                     (sg6.c.segment_number == 6), isouter=True)).\
            group_by(pa.c.call_sign, pa.c.path_number,
                     fr.c.frequency_assigned, fr.c.frequency_upper_band,
                     sa.func.substr(em.c.emission_code, 1, 4)).\
            order_by(pa.c.call_sign, pa.c.path_number)
        path_freq_bucket = PathFreqBucket()
        row_num = 0
        for callsign, path_num, status, radio_service, name,  common_carrier, \
                mobile, grant_date, expired_date,  cancellation_date, \
                rx_callsign, rx_antenna_num,  freq_mhz, freq_upper_band_mhz, \
                tx_eirp, emission_code, \
                tx_lat_deg, tx_lat_min, tx_lat_sec, tx_lat_dir, tx_lon_deg, \
                tx_lon_min, tx_lon_sec, tx_lon_dir, tx_ground_elev_m, \
                tx_loc_type_code, tx_polarization, \
                tx_height_to_center_raat_m, tx_gain, \
                rx_lat_deg, rx_lat_min, rx_lat_sec, rx_lat_dir, \
                rx_lon_deg, rx_lon_min, rx_lon_sec, rx_lon_dir, \
                rx_ground_elev_m, rx_loc_type_code, rx_ant_model, \
                rx_height_to_center_raat_m, rx_gain, p_rx_indicator, \
                r1rx_lat_deg, r1rx_lat_min, r1rx_lat_sec, \
                r1rx_lat_dir, r1rx_lon_deg, r1rx_lon_min, \
                r1rx_lon_sec, r1rx_lon_dir, \
                r1rx_ground_elevation, r1rx_loc_type_code, \
                r1rx_height_to_center_raat_m, r1rx_reflector_height, \
                r1rx_reflector_width, r1rx_back_to_back_gain_tx, \
                r1rx_back_to_back_gain_rx, \
                r1tx_lat_deg, r1tx_lat_min, r1tx_lat_sec, \
                r1tx_lat_dir, r1tx_lon_deg, r1tx_lon_min, \
                r1tx_lon_sec, r1tx_lon_dir, \
                r1tx_ground_elevation, r1tx_loc_type_code, \
                r1tx_height_to_center_raat_m, r1tx_reflector_height, \
                r1tx_reflector_width, r1tx_back_to_back_gain_tx, \
                r1tx_back_to_back_gain_rx, \
                r2rx_lat_deg, r2rx_lat_min, r2rx_lat_sec, \
                r2rx_lat_dir, r2rx_lon_deg, r2rx_lon_min, \
                r2rx_lon_sec, r2rx_lon_dir, \
                r2rx_ground_elevation, r2rx_loc_type_code, \
                r2rx_height_to_center_raat_m, r2rx_reflector_height, \
                r2rx_reflector_width, r2rx_back_to_back_gain_tx, \
                r2rx_back_to_back_gain_rx, \
                r2tx_lat_deg, r2tx_lat_min, r2tx_lat_sec, \
                r2tx_lat_dir, r2tx_lon_deg, r2tx_lon_min, \
                r2tx_lon_sec, r2tx_lon_dir, \
                r2tx_ground_elevation, r2tx_loc_type_code, \
                r2tx_height_to_center_raat_m, r2tx_reflector_height, \
                r2tx_reflector_width, r2tx_back_to_back_gain_tx, \
                r2tx_back_to_back_gain_rx, \
                r3rx_lat_deg, r3rx_lat_min, r3rx_lat_sec, \
                r3rx_lat_dir, r3rx_lon_deg, r3rx_lon_min, \
                r3rx_lon_sec, r3rx_lon_dir, \
                r3rx_ground_elevation, r3rx_loc_type_code, \
                r3rx_height_to_center_raat_m, r3rx_reflector_height, \
                r3rx_reflector_width, r3rx_back_to_back_gain_tx, \
                r3rx_back_to_back_gain_rx, \
                r3tx_lat_deg, r3tx_lat_min, r3tx_lat_sec, \
                r3tx_lat_dir, r3tx_lon_deg, r3tx_lon_min, \
                r3tx_lon_sec, r3tx_lon_dir, \
                r3tx_ground_elevation, r3tx_loc_type_code, \
                r3tx_height_to_center_raat_m, r3tx_reflector_height, \
                r3tx_reflector_width, r3tx_back_to_back_gain_tx, \
                r3tx_back_to_back_gain_rx, \
                r4rx_lat_deg, r4rx_lat_min, r4rx_lat_sec, \
                r4rx_lat_dir, r4rx_lon_deg, r4rx_lon_min, \
                r4rx_lon_sec, r4rx_lon_dir, \
                r4rx_ground_elevation, r4rx_loc_type_code, \
                r4rx_height_to_center_raat_m, r4rx_reflector_height, \
                r4rx_reflector_width, r4rx_back_to_back_gain_tx, \
                r4rx_back_to_back_gain_rx, \
                r4tx_lat_deg, r4tx_lat_min, r4tx_lat_sec, \
                r4tx_lat_dir, r4tx_lon_deg, r4tx_lon_min, \
                r4tx_lon_sec, r4tx_lon_dir, \
                r4tx_ground_elevation, r4tx_loc_type_code, \
                r4tx_height_to_center_raat_m, r4tx_reflector_height, \
                r4tx_reflector_width, r4tx_back_to_back_gain_tx, \
                r4tx_back_to_back_gain_rx, \
                sg6_callsign in self._conn.execute(s):
            error_if(sg6_callsign is not None, "6-segment path encountered")
            repeaters: List[PathInfo.Repeater] = []
            for rrx_lat_deg, rrx_lat_min, rrx_lat_sec, \
                    rrx_lat_dir, rrx_lon_deg, rrx_lon_min, \
                    rrx_lon_sec, rrx_lon_dir, \
                    rrx_ground_elevation, rrx_loc_type_code, rrx_h2c_raat_m, \
                    rrx_reflector_height, rrx_reflector_width, \
                    rrx_back_to_back_gain_tx, rrx_back_to_back_gain_rx, \
                    rtx_lat_deg, rtx_lat_min, rtx_lat_sec, \
                    rtx_lat_dir, rtx_lon_deg, rtx_lon_min, \
                    rtx_lon_sec, rtx_lon_dir, \
                    rtx_ground_elevation, rtx_loc_type_code, rtx_h2c_raat_m, \
                    rtx_reflector_height, rtx_reflector_width, \
                    rtx_back_to_back_gain_tx, rtx_back_to_back_gain_rx \
                    in \
                    [(r1rx_lat_deg, r1rx_lat_min, r1rx_lat_sec,
                      r1rx_lat_dir, r1rx_lon_deg, r1rx_lon_min,
                      r1rx_lon_sec, r1rx_lon_dir,
                      r1rx_ground_elevation, r1rx_loc_type_code,
                      r1rx_height_to_center_raat_m,
                      r1rx_reflector_height, r1rx_reflector_width,
                      r1rx_back_to_back_gain_tx,  r1rx_back_to_back_gain_rx,
                      r1tx_lat_deg, r1tx_lat_min, r1tx_lat_sec,
                      r1tx_lat_dir, r1tx_lon_deg, r1tx_lon_min,
                      r1tx_lon_sec, r1tx_lon_dir,
                      r1tx_ground_elevation, r1tx_loc_type_code,
                      r1tx_height_to_center_raat_m,
                      r1tx_reflector_height, r1tx_reflector_width,
                      r1tx_back_to_back_gain_tx, r1tx_back_to_back_gain_rx),
                     (r2rx_lat_deg, r2rx_lat_min, r2rx_lat_sec,
                      r2rx_lat_dir, r2rx_lon_deg, r2rx_lon_min,
                      r2rx_lon_sec, r2rx_lon_dir,
                      r2rx_ground_elevation, r2rx_loc_type_code,
                      r2rx_height_to_center_raat_m,
                      r2rx_reflector_height, r2rx_reflector_width,
                      r2rx_back_to_back_gain_tx, r2rx_back_to_back_gain_rx,
                      r2tx_lat_deg, r2tx_lat_min, r2tx_lat_sec,
                      r2tx_lat_dir, r2tx_lon_deg, r2tx_lon_min,
                      r2tx_lon_sec, r2tx_lon_dir,
                      r2tx_ground_elevation, r2tx_loc_type_code,
                      r2tx_height_to_center_raat_m,
                      r2tx_reflector_height, r2tx_reflector_width,
                      r2tx_back_to_back_gain_tx, r2tx_back_to_back_gain_rx),
                     (r3rx_lat_deg, r3rx_lat_min, r3rx_lat_sec,
                      r3rx_lat_dir, r3rx_lon_deg, r3rx_lon_min,
                      r3rx_lon_sec, r3rx_lon_dir,
                      r3rx_ground_elevation, r3rx_loc_type_code,
                      r3rx_height_to_center_raat_m,
                      r3rx_reflector_height, r3rx_reflector_width,
                      r3rx_back_to_back_gain_tx, r3rx_back_to_back_gain_rx,
                      r3tx_lat_deg, r3tx_lat_min, r3tx_lat_sec,
                      r3tx_lat_dir, r3tx_lon_deg, r3tx_lon_min,
                      r3tx_lon_sec, r3tx_lon_dir,
                      r3tx_ground_elevation, r3tx_loc_type_code,
                      r3tx_height_to_center_raat_m,
                      r3tx_reflector_height, r3tx_reflector_width,
                      r3tx_back_to_back_gain_tx, r3tx_back_to_back_gain_rx),
                     (r4rx_lat_deg, r4rx_lat_min, r4rx_lat_sec,
                      r4rx_lat_dir, r4rx_lon_deg, r4rx_lon_min,
                      r4rx_lon_sec, r4rx_lon_dir,
                      r4rx_ground_elevation, r4rx_loc_type_code,
                      r4rx_height_to_center_raat_m,
                      r4rx_reflector_height, r4rx_reflector_width,
                      r4rx_back_to_back_gain_tx, r4rx_back_to_back_gain_rx,
                      r4tx_lat_deg, r4tx_lat_min, r4tx_lat_sec,
                      r4tx_lat_dir, r4tx_lon_deg, r4tx_lon_min,
                      r4tx_lon_sec, r4tx_lon_dir,
                      r4tx_ground_elevation, r4tx_loc_type_code,
                      r4tx_height_to_center_raat_m,
                      r4tx_reflector_height, r4tx_reflector_width,
                      r4tx_back_to_back_gain_tx, r4tx_back_to_back_gain_rx)]:
                if rtx_lat_deg is not None:
                    rx = \
                        PathInfo.Repeater(
                            ant_coord=PathInfo.AntCoord(
                                lat_deg=rrx_lat_deg, lat_min=rrx_lat_min,
                                lat_sec=rrx_lat_sec, lat_dir=rrx_lat_dir,
                                lon_deg=rrx_lon_deg, lon_min=rrx_lon_min,
                                lon_sec=rrx_lon_sec, lon_dir=rrx_lon_dir,
                                ground_elev_m=rrx_ground_elevation,
                                height_to_center_raat_m=rrx_h2c_raat_m,
                                loc_type_code=rrx_loc_type_code),
                            reflector_height=rrx_reflector_height,
                            reflector_width=rrx_reflector_width,
                            back_to_back_gain_tx=rrx_back_to_back_gain_tx,
                            back_to_back_gain_rx=rrx_back_to_back_gain_rx)
                    tx = \
                        PathInfo.Repeater(
                            ant_coord=PathInfo.AntCoord(
                                lat_deg=rtx_lat_deg, lat_min=rtx_lat_min,
                                lat_sec=rtx_lat_sec, lat_dir=rtx_lat_dir,
                                lon_deg=rtx_lon_deg, lon_min=rtx_lon_min,
                                lon_sec=rtx_lon_sec, lon_dir=rtx_lon_dir,
                                ground_elev_m=rtx_ground_elevation,
                                height_to_center_raat_m=rtx_h2c_raat_m,
                                loc_type_code=rtx_loc_type_code),
                            reflector_height=rtx_reflector_height,
                            reflector_width=rtx_reflector_width,
                            back_to_back_gain_tx=rtx_back_to_back_gain_tx,
                            back_to_back_gain_rx=rtx_back_to_back_gain_rx)
                    consistent, repeater = PathInfo.Repeater.combine(rx, tx)
                    if not consistent:
                        ComplaintInconsistentRepeater(
                            callsign=callsign, path_num=path_num,
                            repeater_idx=len(repeaters) + 1)
                    if not repeater:
                        break
                    repeaters.append(repeater)
            path = \
                PathInfo(
                    callsign=callsign, status=status,
                    radio_service=radio_service, name=name,
                    common_carrier=common_carrier, mobile=mobile,
                    rx_callsign=rx_callsign, rx_antenna_num=rx_antenna_num,
                    freq_range=FreqRange(emission_code=emission_code,
                                         center_mhz=freq_mhz)
                    if freq_upper_band_mhz is None
                    else FreqRange(emission_code=emission_code,
                                   start_mhz=freq_mhz,
                                   end_mhz=freq_upper_band_mhz),
                    tx_eirp=tx_eirp,
                    tx_ant=PathInfo.AntCoord(
                        lat_deg=tx_lat_deg, lat_min=tx_lat_min,
                        lat_sec=tx_lat_sec, lat_dir=tx_lat_dir,
                        lon_deg=tx_lon_deg, lon_min=tx_lon_min,
                        lon_sec=tx_lon_sec, lon_dir=tx_lon_dir,
                        ground_elev_m=tx_ground_elev_m,
                        height_to_center_raat_m=tx_height_to_center_raat_m,
                        loc_type_code=tx_loc_type_code),
                    tx_polarization=tx_polarization, tx_gain=tx_gain,
                    rx_ant=PathInfo.AntCoord(
                        lat_deg=rx_lat_deg, lat_min=rx_lat_min,
                        lat_sec=rx_lat_sec, lat_dir=rx_lat_dir,
                        lon_deg=rx_lon_deg, lon_min=rx_lon_min,
                        lon_sec=rx_lon_sec, lon_dir=rx_lon_dir,
                        ground_elev_m=rx_ground_elev_m,
                        height_to_center_raat_m=rx_height_to_center_raat_m,
                        loc_type_code=rx_loc_type_code),
                    rx_ant_model=rx_ant_model, rx_gain=rx_gain,
                    repeaters=repeaters, p_rx_indicator=p_rx_indicator,
                    path_num=path_num, grant_date=grant_date,
                    expired_date=expired_date,
                    cancellation_date=cancellation_date)
            row_num += 1
            if (row_num % 10000) == 0:
                Progressor.print_progress(f"{row_num}")

            if not path_freq_bucket.try_add(path):
                for p in path_freq_bucket:
                    yield p
                path_freq_bucket.clear()
                path_freq_bucket.try_add(path)
        for p in path_freq_bucket:
            yield p
    Progressor.print_progress()


class FsidGenerator:
    """ Generates 'fsid' field for AFC database

    Private attributes:
    _fsid_by_path -- Dictionary with already used fsids, indexed by (callsign,
                     path_num, start_mhz, end_mhz) tuples
    _next_fsid    -- Next fsid value to use
    """
    def __init__(self, prev_afc_db: Optional[str] = None) -> None:
        """ Constructor

        Arguments:
        prev_afc_db -- AFC database to continue FSIDs from. None to start from
                       1
        """
        self._fsid_by_path: Dict[Tuple[str, int, float, float], int] = {}
        self._next_fsid = 1
        if prev_afc_db:
            with Timing("Retrieving previous FSIDs"), \
                    DbPathIterator(db_filename=prev_afc_db) as dpi:
                for path in dpi:
                    error_if(path.fsid is None,
                             f"'fsid` field not found in {prev_afc_db}")
                    error_if(
                        path.path_num is None,
                        f"'path_number` field not found in {prev_afc_db}")
                    assert path.fsid is not None
                    assert path.path_num is not None
                    self._next_fsid = max(self._next_fsid, path.fsid + 1)
                    self._fsid_by_path[self._path_key(path)] = path.fsid

    def fsid(self, path: PathInfo) -> int:
        """ Returns FSID for given path """
        if path.path_num is None:
            ret1 = self._next_fsid
            self._next_fsid += 1
            return ret1
        ret2 = self._fsid_by_path.get(self._path_key(path))
        if ret2 is None:
            self._fsid_by_path[self._path_key(path)] = self._next_fsid
            ret2 = self._next_fsid
            self._next_fsid += 1
        return ret2

    def _path_key(self, path: PathInfo) -> Tuple[str, int, float, float]:
        """ Return tuple, consisting of components used to compute fsid """
        assert path.path_num is not None
        assert path.freq_range.start_mhz is not None
        assert path.freq_range.end_mhz is not None
        return (path.callsign, path.path_num,
                path.freq_range.start_mhz, path.freq_range.end_mhz)


class AfcWriter:
    """ Writes paths to AFC database
    Must be used as context manager (with 'with' clause)

    Private attributes:
    _metadata       -- Database metadata
    _no_bulk        -- True if no bulk write should be made (terminally slow,
                       but works on SqlAlchemy 1.2)
    _path_rows_data -- List of path table row data dictionaries, prepared for
                       bulk write
    _rep_rows_data  -- List of repeater table row data dictionaries, prepared
                       for bulk write
    _path_table     -- Table object for paths
    _rep_table      -- Table object for repeaters or None
    _validator      -- PathValidator to vet paths being written
    _fsid_gen       -- FsidGenerator object for 'fsid' field generation
    _db_format      -- AfcDb.Format to write database in. None to write in
                       universally compatible format
    _path_col_names -- Column names in path database
    _rep_col_names  -- Column names in repeater database
    """
    # Number of records to write to AFC database in a single transaction
    AFC_TRANSACTION_LENGTH = 500

    def __init__(self, db_filename: Optional[str], validator: PathValidator,
                 fsid_gen: FsidGenerator, compatibility: bool) -> None:
        """ Constructor

        Arguments:
        db_filename   -- Name of SQLite file to generate. None means creation
                         of fake writer that explodes on attempt to write
        validator     -- PathValidator to vet incoming paths
        fsid_gen      -- FsidGenerator object for 'fsid' field generation
        compatibility -- True to write database in universally-compatible
                         format, False to write in latest known format
        """
        self._validator = validator
        self._metadata = sa.MetaData()
        self._fsid_gen = fsid_gen
        self._no_bulk = False
        self._path_rows_data: List[ROW_DATA_TYPE] = []
        self._rep_rows_data: List[ROW_DATA_TYPE] = []
        self._db_format = None if compatibility else AfcDb.Format.latest()
        self._path_col_names: Set[str] = set()
        self._rep_col_names: Set[str] = set()

        sa_path_table_args: List[Any] = [AfcDb.PATH_TABLE_NAME, self._metadata]
        for col in AfcDb.columns(db_format=self._db_format,
                                 table_name=AfcDb.PATH_TABLE_NAME):
            sa_path_table_args.append(col.col_arg.apply(sa.Column))
            self._path_col_names.add(col.name)
        self._path_table = sa.Table(*sa_path_table_args)

        self._rep_table: Optional[sa.Table] = None
        if (self._db_format is None) or self._db_format.separate_repeaters:
            sa_rep_table_args: List[Any] = [AfcDb.REP_TABLE_NAME,
                                            self._metadata]
            for col in AfcDb.columns(db_format=self._db_format,
                                     table_name=AfcDb.REP_TABLE_NAME):
                sa_rep_table_args.append(col.col_arg.apply(sa.Column))
                self._rep_col_names.add(col.name)
            self._rep_table = sa.Table(*sa_rep_table_args)

        if db_filename is not None:
            engine = sa.create_engine("sqlite:///" + db_filename)
            self._metadata.create_all(engine)
            self._conn = engine.connect()
        else:
            self._conn = None

    def __enter__(self) -> "AfcWriter":
        """ Context manager entry. Returns self """
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
        """ Context manager exit. Write last transaction and close connection
        """
        if self._conn is not None:
            self._write()
            self._conn.close()

    def write_path(self, path: PathInfo) -> None:
        """ Writes PathInfo object to database """
        assert self._conn is not None
        if not self._validator.validate(path):
            return
        path_row_data: ROW_DATA_TYPE = \
            {"fsid": self._fsid_gen.fsid(path),
             "callsign": path.callsign,
             "status": path.status,
             "radio_service": path.radio_service,
             "name": path.name or "",
             "common_carrier": path.common_carrier,
             "mobile": path.mobile,
             "rx_callsign": path.rx_callsign or path.callsign,
             "rx_antenna_num": path.rx_antenna_num,
             "freq_assigned_start_mhz": path.freq_range.start_mhz,
             "freq_assigned_end_mhz": path.freq_range.end_mhz,
             "tx_eirp": path.tx_eirp,
             "emissions_des": path.freq_range.emission_code,
             "tx_lat_deg": path.tx_ant.lat,
             "tx_long_deg": path.tx_ant.lon,
             "tx_ground_elev_m": path.tx_ant.ground_elev_m,
             "tx_polarization": path.tx_polarization,
             "tx_height_to_center_raat_m": path.tx_ant.height_to_center_raat_m,
             "tx_gain": path.tx_gain,
             "rx_lat_deg": path.rx_ant.lat,
             "rx_long_deg": path.rx_ant.lon,
             "rx_ground_elev_m": path.rx_ant.ground_elev_m,
             "rx_ant_model": path.rx_ant_model,
             "rx_height_to_center_raat_m": path.rx_ant.height_to_center_raat_m,
             "rx_gain": path.rx_gain,
             "p_rx_indicator": path.p_rx_indicator,
             "path_number": path.path_num}
        empty_repeater = \
            PathInfo.Repeater(
                ant_coord=PathInfo.AntCoord(), reflector_height=None,
                reflector_width=None, back_to_back_gain_tx=None,
                back_to_back_gain_rx=None)

        # Modern multirepeater format
        path_row_data["p_rp_num"] = len(path.repeaters)
        for idx in range(1, AfcDb.MAX_REPEATERS + 1):
            r = path.repeaters[idx - 1] if idx <= len(path.repeaters) \
                else empty_repeater
            assert r is not None
            assert r.ant_coord is not None
            path_row_data[f"p_rp_{idx}_lat_degs"] = r.ant_coord.lat
            path_row_data[f"p_rp_{idx}_lon_degs"] = r.ant_coord.lon
            path_row_data[f"p_rp_{idx}_height_to_center_raat_m"] = \
                r.ant_coord.height_to_center_raat_m
            path_row_data[f"p_rp_{idx}_reflector_height_m"] = \
                r.reflector_height
            path_row_data[f"p_rp_{idx}_reflector_width_m"] = r.reflector_width
            path_row_data[f"p_rp_{idx}_back_to_back_gain_tx"] = \
                r.back_to_back_gain_tx
            path_row_data[f"p_rp_{idx}_back_to_back_gain_rx"] = \
                r.back_to_back_gain_rx

        # Legacy single-repeater format
        r = path.repeaters[-1] if path.repeaters else empty_repeater
        assert r is not None
        assert r.ant_coord is not None
        path_row_data["p_rp_lat_degs"] = r.ant_coord.lat
        path_row_data["p_rp_lon_degs"] = r.ant_coord.lon
        path_row_data["p_rp_height_to_center_raat_m"] = \
            r.ant_coord.height_to_center_raat_m

        # Separate repeaters table
        repeater_rows_data: List[ROW_DATA_TYPE] = []
        if self._rep_col_names:
            for idx, r in enumerate(path.repeaters):
                assert r.ant_coord is not None
                repeater_rows_data.append(
                    {"fsid": path_row_data["fsid"],
                     "prSeq": idx + 1,
                     "pr_lat_deg": r.ant_coord.lat,
                     "pr_lon_deg": r.ant_coord.lon,
                     "pr_height_to_center_raat_m":
                         r.ant_coord.height_to_center_raat_m,
                     "pr_reflector_height_m": r.reflector_height,
                     "pr_reflector_width_m": r.reflector_width,
                     "pr_back_to_back_gain_tx": r.back_to_back_gain_tx,
                     "pr_back_to_back_gain_rx": r.back_to_back_gain_rx})
                self._filter_columns(row_data=repeater_rows_data[-1],
                                     allowed_columns=self._rep_col_names)
        self._filter_columns(row_data=path_row_data,
                             allowed_columns=self._path_col_names)
        self._write(path_row_data, repeater_rows_data)

    def _write(self, path_row_data: Optional[ROW_DATA_TYPE] = None,
               rep_rows_data: Optional[List[ROW_DATA_TYPE]] = None) -> None:
        """ Writes given row dictionary to database

        Arguments:
        path_row_data -- Path row dictionary or None - in latter case write of
                         accumulated data to database is forced """
        if path_row_data is not None:
            self._path_rows_data.append(path_row_data)
        if rep_rows_data:
            self._rep_rows_data += rep_rows_data
        if not self._path_rows_data:
            return
        if (path_row_data is None) or self._no_bulk or \
                (len(self._path_rows_data) > self.AFC_TRANSACTION_LENGTH):
            transaction = self._conn.begin()
            ins = \
                sa.insert(self._path_table).execution_options(autocommit=False)
            if self._no_bulk:
                self._conn.execute(ins, self._path_rows_data)
            else:
                try:
                    self._conn.execute(ins.values(self._path_rows_data))
                except sa.exc.CompileError as ex:
                    logging.error(f"Database creation problem: {ex}")
                    if self._no_bulk:
                        raise
                    self._no_bulk = True
                    transaction.rollback()
                    transaction = self._conn.begin()
                    self._conn.execute(ins, self._path_rows_data)
            if self._rep_rows_data:
                ins = sa.insert(self._rep_table).\
                    execution_options(autocommit=False)
                if self._no_bulk:
                    for rep_row_data in self._rep_rows_data:
                        self._conn.execute(ins, rep_row_data)
                else:
                    self._conn.execute(ins.values(self._rep_rows_data))
            transaction.commit()
            self._path_rows_data = []
            self._rep_rows_data = []

    def _filter_columns(self, row_data: ROW_DATA_TYPE,
                        allowed_columns: Set[str]) -> None:
        """ Remove columns except allowed

        Arguments:
        row_data        -- Name-value dictionary
        allowed_columns -- Allowed column names """
        for column in (row_data.keys() - allowed_columns):
            del row_data[column]


class Downloader:
    """ Handles downloading FCC ULS file to local directory
    Must be used as context manager (with 'with' clause)

    Public attributes
    date     -- (upload_datetime, timezone_name) or (None, None)
    inc_date -- (upload_datetime, timezone_name) of latest increment data or
                (None, None_

    Private attributes:
    _zip_dir      -- Directory with zip files or None
    _unpacked_dir -- Directory with unpacked raw data
    _no_download  -- Don't do download, just use already existing files
    _retain       -- True to retain download directory, False to remove it
                     after operation
    _zip_file     -- None or .zip file to take all data from
    """
    # Weekly FCC ULS archive URL
    WEEKLY_URL = "https://data.fcc.gov/download/pub/uls/complete/l_micro.zip"

    # Daily updates FCC ULS archive URL. {dow} is placeholder for day of week
    # ("mon", "tue", ...)
    DAILY_URL = "https://data.fcc.gov/download/pub/uls/daily/l_mw_{dow}.zip"

    # Download chunk size
    DOWNLOAD_CHUNK_SIZE = 1024 * 1024

    # Special 'day of week' name, designating the full archive
    DOW_FULL = "full"

    # Days of week for daily update names
    DOWS = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]

    # Maximum download attempts
    MAX_DOWNLOAD_RETRIES = 10

    def __init__(self, from_dir: Optional[str], to_dir: Optional[str],
                 retry: bool, zip_file: Optional[str]) -> None:
        """ Constructor

        Arguments:
        from_dir    -- None or directory to read files from
        to_dir      -- None or directory to download files to
        retry       -- True to retry download several times
        zip_file    -- None or name of ZIP archive to take all data from
        """
        error_if(from_dir and (not os.path.isdir(from_dir)),
                 f"Directory '{from_dir}' not found")
        error_if(zip_file and (not os.path.isfile(zip_file)),
                 f"Archive '{zip_file}' not found")
        self._zip_dir = from_dir or to_dir
        self._unpacked_dir: Optional[str] = None
        self._no_download = from_dir is not None
        self._retry = retry
        self._retain = self._zip_dir is not None
        self._zip_file = zip_file
        self.date: Tuple[Optional[datetime.datetime], Optional[str]] = \
            (None, None)
        self.inc_date: Tuple[Optional[datetime.datetime], Optional[str]] = \
            (None, None)

    def __enter__(self) -> "Downloader":
        """ Context manager entry. Returns self """
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
        """ Context manager exit. Removes download directory if it is not to be
        retained """
        if self._unpacked_dir is not None:
            shutil.rmtree(self._unpacked_dir)
        if (not self._retain) and (self._zip_dir is not None):
            shutil.rmtree(self._zip_dir)

    @property
    def unpacked_directory(self):
        """ Root directory of unpacked files. Asserts if asked before
        download() """
        assert self.date[0] is not None
        return self._unpacked_dir

    def download(self) -> Dict[Optional[datetime.datetime], str]:
        """ Perform the download

        Returns dictionary. Keys are: None for last full database, datetimes
        for daily updates. Values are directories with unpacked raw ULS files
        for these days """
        dow_url: Callable[[str], str] = \
            lambda dow: self.WEEKLY_URL if (dow == self.DOW_FULL) \
            else self.DAILY_URL.format(dow=dow)
        # Creating zip directory (if needed) and unpacked files root directory
        if self._zip_dir is None:
            self._zip_dir = \
                tempfile.mkdtemp(
                    prefix=os.path.splitext(os.path.basename(__file__))[0])
        elif not self._zip_file:
            os.makedirs(self._zip_dir, exist_ok=True)

        self._unpacked_dir = \
            tempfile.mkdtemp(
                prefix=os.path.splitext(os.path.basename(__file__))[0])

        # Daily/full directory names
        dows = [self.DOW_FULL] + self.DOWS
        # Daily archives
        zip_files: Dict[str, str] = {}

        # Downloading/finding zip archives
        if self._zip_file:
            zip_files[self.DOW_FULL] = self._zip_file
        elif self._no_download:
            assert self._zip_dir is not None
            for dow, synonyms in ([(self.DOW_FULL, ["weekly", "full"])] +
                                  [(d, [d]) for d in self.DOWS]):
                for fn in \
                        [os.path.splitext(
                            os.path.basename(
                                urllib.parse.urlparse(
                                    dow_url(dow)).path))[0]] + \
                        synonyms:
                    full_file_name = os.path.join(self._zip_dir, fn) + ".zip"
                    if os.path.isfile(full_file_name):
                        zip_files[dow] = full_file_name
                        break
        else:
            old_cdhc = ssl._create_default_https_context
            ssl._create_default_https_context = ssl._create_unverified_context
            try:
                for dow in dows:
                    with Timing(f"Downloading '{dow_url(dow)}'"):
                        attempt = 0
                        num_attempts = \
                            self.MAX_DOWNLOAD_RETRIES if self._retry else 1
                        while attempt < num_attempts:
                            if attempt:
                                time.sleep(10)
                            attempt += 1
                            zip_file = self._download_file(dow_url(dow))
                            if zip_file is not None:
                                try:
                                    with zipfile.ZipFile(zip_file) as z:
                                        if z.testzip() is not None:
                                            zip_file = None
                                except zipfile.BadZipFile:
                                    assert zip_file is not None
                                    logging.warning(
                                        (f"Broken "
                                         f"'{os.path.basename(zip_file)}'"))
                                    zip_file = None
                            if zip_file is not None:
                                zip_files[dow] = zip_file
                                break
                            if attempt < num_attempts:
                                logging.warning("Retrying")
                                continue
                            logging.warning(
                                f"Failed to download {dow_url(dow)}")
                            break
            finally:
                ssl._create_default_https_context = old_cdhc
        error_if(self.DOW_FULL not in zip_files,
                 (f"Full (nonincremental) ULS data not "
                  f"{'found' if self._no_download else 'downloaded'}"))
        ret: Dict[Optional[datetime.datetime], str] = {}
        full_date: Optional[datetime.datetime] = None
        self.date = (None, None)
        self.inc_date = (None, None)
        # Unpacking archives and computing dates based on 'counts' files
        for dow in dows:
            if dow not in zip_files:
                continue
            unpacked_dow_dir = os.path.join(mst(self._unpacked_dir), dow)
            try:
                with zipfile.ZipFile(zip_files[dow]) as z:
                    z.extractall(unpacked_dow_dir)
            except zipfile.BadZipFile:
                error(f"Broken '{os.path.basename(zip_files[dow])}'")
            counts_filename = os.path.join(unpacked_dow_dir, "counts")
            error_if(not os.path.isfile(counts_filename),
                     f"'counts' file not found in "
                     f"'{os.path.basename(zip_files[dow])}'")
            with open(counts_filename, mode="r", encoding="ascii") as f:
                content_date, content_tz = \
                    self._parse_counts_date(counts_content=f.read(),
                                            url=dow_url(dow))
            if (self.date[0] is None) or (content_date > self.date[0]):
                self.date = (content_date, content_tz)
            if (dow == self.DOW_FULL):
                full_date = content_date
            elif (self.inc_date[0] is None) or \
                    (content_date > self.inc_date[0]):
                self.inc_date = (content_date, content_tz)
            assert full_date is not None
            if (dow == self.DOW_FULL) or (content_date > full_date):
                ret[None if (dow == self.DOW_FULL) else content_date] = \
                    unpacked_dow_dir
        return ret

    def _download_file(self, url: str) -> Optional[str]:
        """ Download file at given URL to download directory """
        assert self._zip_dir is not None
        filename = \
            os.path.join(self._zip_dir,
                         os.path.basename(urllib.parse.urlparse(url).path))
        download_started = False
        fail_reason = None
        try:
            with open(filename, mode="wb") as fo:
                size = 0
                with urllib.request.urlopen(url, timeout=30) as fi:
                    while True:
                        data = fi.read(self.DOWNLOAD_CHUNK_SIZE)
                        size += len(data)
                        Progressor.print_progress(f"{size/(1024*1024)}MB")
                        if data:
                            download_started = True
                            fo.write(data)
                        else:
                            break
        except (ConnectionError, urllib.error.URLError, socket.timeout) \
                as ex:
            fail_reason = str(ex)
        Progressor.print_progress()
        if download_started and (fail_reason is None):
            return filename
        logging.warning(
            (f"Download attempt for {url} failed"
             f"{(' (' + fail_reason + ')') if fail_reason else ''}"))
        return None

    def _parse_counts_date(self, counts_content: str, url: str) \
            -> Tuple[datetime.datetime, str]:
        """ Returns datetime and time zone name, extracted from given contents
        of 'counts' file """
        m = \
            re.match(
                r"^File Creation Date: " +
                r"\S{3}\s+(\S{3}\s+\d+\s+\d+:\d+:\d+)\s+(\S+)\s+(\d+)(\r|\n)",
                counts_content)
        error_if(not m, f"Unrecognized date format in 'counts' of '{url}'")
        assert m is not None

        old_lc_time = locale.getlocale(locale.LC_TIME)
        locale.setlocale(locale.LC_TIME, "C")
        try:
            return \
                (datetime.datetime.strptime(m.group(1) + " " + m.group(3),
                                            "%b %d %H:%M:%S %Y"),
                 m.group(2))
        except ValueError:
            error(f"Unrecognized date format in 'counts' of '{url}'")
            return (datetime.datetime.now(), "")  # Appeasing MyPy
        finally:
            locale.setlocale(locale.LC_TIME, old_lc_time)


class UlsFileReader:
    """ Iterates over records in given FCC ULS table file
    Must be used as context manager (with 'with' clause)

    Private attributes:
    _filename  -- File name
    _uls_table -- Correspondent UlsTable object
    _f         -- File object
    """
    def __init__(self, filename: str, uls_table: UlsTable) -> None:
        """ Constructor

        Arguments:
        filename  -- Name of ULS file to read
        uls_table -- Corresponded UlsTable object (must be non-None)
        """
        self._filename = filename
        self._uls_table = uls_table
        self._f = open(filename, mode="r", encoding="ascii",
                       errors="backslashreplace")

    def __enter__(self) -> "UlsFileReader":
        """ Context manager entry. Returns self """
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
        """ Context manager exit. Closes file """
        self._f.close()

    def __iter__(self) -> Iterator:
        """ Iterator

        Returns sequence of (row_data, line_number, record) tuples, where
        'row_data' is a dictionary of storable fields from the record, indexed
        by their names, 'line_number' is a line number in file, 'record' is raw
        record string from file """
        record = None
        line_count = 0
        for line in self._f:
            line_count += 1
            if record:
                record += line.rstrip("\r\n")
            else:
                record = line.rstrip("\r\n")
            field_count = record.count("|") + 1
            if field_count < len(self._uls_table.fields):
                continue
            error_if(
                field_count > len(self._uls_table.fields),
                (f"'{self._filename}' has invalid field count "
                 f"around line {line_count}"))
            row_data = self._get_row_data(record=record, line_num=line_count)
            yield (row_data, line_count, record)
            record = None
            if (line_count % 10000) == 0:
                Progressor.print_progress(f"{line_count}")

    def _get_row_data(self, record: str, line_num: int) -> ROW_DATA_TYPE:
        """ Extracts row data from given record

        Arguments:
        record   -- Record: string with the correct number of '|' in it
        line_num -- Line number

        Returns row data dictionary of storable fields from the record, indexed
        by their names
        """
        values = record.split("|")
        ret: ROW_DATA_TYPE = {}
        for idx, uls_field in enumerate(self._uls_table.fields):
            if uls_field.col_args is None:
                continue
            assert uls_field.col_name is not None
            value: ROW_VALUE_TYPE
            if (values[idx] == "") and uls_field.default:
                ComplaintReplacedMandatoryField(
                    table=self._uls_table.name, line_num=line_num,
                    column=uls_field.col_name, replacement=uls_field.default)
                values[idx] = uls_field.default
            if (values[idx] == "") and uls_field.nullable:
                value = None
            elif isinstance(uls_field.col_type, sa.Unicode):
                value = values[idx]
            elif isinstance(uls_field.col_type, sa.Integer):
                try:
                    value = int(values[idx])
                except ValueError:
                    error(
                        (f"'{self._filename}' has invalid integer value "
                         f"{values[idx]} of field {idx + 1} around line "
                         f"{line_num}"))
            elif isinstance(uls_field.col_type, sa.Float):
                try:
                    value = float(values[idx])
                except ValueError:
                    error(
                        (f"'{self._filename}' has invalid floating point "
                         f"value {values[idx]} of field {idx + 1} around line "
                         f"{line_num}"))
            elif isinstance(uls_field.col_type, sa.DateTime):
                # AFC does not use any fields, formatted as datetime, error
                # check is relaxed

                value = None
                # In second format milliseconds, not microseconds. Negligible
                for date_format in ("%m/%d/%Y", "%b %d %Y %I:%M:%S:%f%p"):
                    try:
                        value = datetime.datetime.strptime(values[idx],
                                                           date_format)
                        break
                    except ValueError:
                        pass
                if value is None:
                    logging.warning(
                        (f"'{self._filename}' has invalid date value "
                         f"'{values[idx]}' of field {idx + 1} around line "
                         f"{line_num}"))
            else:
                error(
                    (f"Field type {type(uls_field.col_type)} not yet "
                     f"supported in 'Uls.FileReader._get_row_data()'"))
            ret[uls_field.col_name] = value
        return ret


class UlsTableReader:
    """ Reader of data from ULS table.
    Handles daily patches on top of UlsFileReader

    Private attributes:
    _day_dirs  -- Directories where full file and updates located. Return value
                  of Downloader.download()
    _uls_table -- UlsTable object for ULS table being iterated
    """

    # Table filename pattern in FCC ULS archive. {table} is a placeholder for
    # table name ("AN", "EM", ...)
    TABLE_FILE_PATTERN = "{table}.dat"

    def __init__(self, uls_table: UlsTable,
                 day_dirs: Dict[Optional[datetime.datetime], str]) -> None:
        self._day_dirs = day_dirs
        self._uls_table = uls_table

    def __iter__(self) -> Iterator:
        """ Iterator

        Returns sequence of (filename, line_number, row_data, record) tuples,
        where 'row_data' is dictionary of storable row data, whereas 'filename'
        and  'line_number' specify location from which data was obtained.
        'record' is a raw record string
        """
        # Background: Unit of update is callsign. I.e. if data for callsign
        # present in so,me update, they completely override data for this
        # callsign in full table or in previous updates
        #
        # Keys are callsigns. Values are (day, rows) tuples, where 'day' is
        # datetime of file, from which content was retrieved, 'values' is a
        # list of row tuples. Row tuple is
        # (filename, linenum, row_as_data_dict, row_as_string)
        updates: Dict[str,
                      Tuple[datetime.datetime,
                            List[Tuple[str, int, ROW_DATA_TYPE, str]]]] = {}
        for date in \
                sorted([d for d in self._day_dirs.keys() if d is not None]):
            assert date is not None
            filename = \
                os.path.join(
                    self._day_dirs[date],
                    self.TABLE_FILE_PATTERN.format(table=self._uls_table.name))
            if not os.path.isfile(filename):
                continue
            with UlsFileReader(filename=filename, uls_table=self._uls_table) \
                    as fr:
                for row_data, line_num, record in fr:
                    callsign = row_data["call_sign"]
                    v = updates.get(callsign, (None, None))
                    if v[0] != date:
                        v = (date, [])
                        updates[callsign] = v
                    assert v[1] is not None
                    v[1].append((filename, line_num, row_data, record))
            for _, rows in updates.values():
                for filename, line_num, row_data, record in rows:
                    yield (filename, line_num, row_data, record)
        filename = \
            os.path.join(
                self._day_dirs[None],
                self.TABLE_FILE_PATTERN.format(table=self._uls_table.name))
        with UlsFileReader(filename=filename, uls_table=self._uls_table) as fr:
            for row_data, line_num, record in fr:
                if row_data["call_sign"] in updates:
                    continue
                yield (filename, line_num, row_data, record)


class UlsDatabaseFiller:
    """ Utilities to write ULS database
    Must be used as context manager (with 'with' clause)

    Private attributes:
    _day_dirs   -- Directories with downloaded FCC ULS files (return value of
                   Downloader.download())
    _skip_write -- Create database, but do not write into it (for performance
                   measurement purposes)
    _metadata   -- Database metadata
    _conn       -- Database connection
    """
    # Number of records to write to ULS database in a single transaction (1000
    # is beyond SqlAlchemy/SQlite capacity)
    ULS_TRANSACTION_LENGTH = 500

    class _TableWriter:
        """ Bulk (and otherwise) writer to ULS database table
        Must be used as context manager (with 'with' clause)

        Private attributes:
        _conn           -- Database connection
        _metadata       -- Database metadata
        _skip_write     -- True to not write to database
        _uls_table      -- UlsTable that corresponds to database table
        _no_bulk        -- Don't use bulk writes
        _sa_table       -- sa.Table object for database table
        _line_rows_dict -- Collection of row data dictionaries for future bulk
                           (or whatever) write: dictionary indexed by
                           (file_name, line_number) of where row data was taken
                           from
        """
        def __init__(self, conn: sa.engine.Connection, metadata: sa.MetaData,
                     name: str, skip_write: bool) -> None:
            """ Constructor

            Arguments:
            conn       -- Database connection
            metadata   -- Database metadata
            name       -- Table name
            skip_write -- True to not write to database
            """
            self._conn = conn
            self._metadata = metadata
            self._skip_write = skip_write
            self._uls_table = UlsTable.by_name(name)
            self._no_bulk = False
            self._sa_table: sa.Table = metadata.tables[name]
            self._line_rows_dict: Dict[Tuple[str, int], ROW_DATA_TYPE] = {}

        def __enter__(self) -> "UlsDatabaseFiller._TableWriter":
            """ Context manager entry. Returns self """
            return self

        def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
            """ Context manager exit. Writes what remains """
            self._bulk_write()

        def write(self, table_filename: str, line: int,
                  row_data: ROW_DATA_TYPE) -> None:
            """ Write row data to database table

            Arguments:
            table_filename -- File of row data origin
            line_number    -- Line number of row data origin
            row_data       -- Row data to write to database
            """
            if self._skip_write:
                return
            self._line_rows_dict[(table_filename, line)] = row_data
            if self._no_bulk or \
                    (len(self._line_rows_dict) >
                     UlsDatabaseFiller.ULS_TRANSACTION_LENGTH):
                self._bulk_write()
                self._line_rows_dict = {}

        def _bulk_write(self) -> None:
            """ Performs bulk write """
            if not self._line_rows_dict:
                return
            success = False
            if not self._no_bulk:
                try:
                    transaction = self._conn.begin()
                    rows: List[ROW_DATA_TYPE] = \
                        [self._line_rows_dict[fnln]
                         for fnln in sorted(self._line_rows_dict.keys())]
                    ins = sa.insert(self._sa_table).values(rows).\
                        execution_options(autocommit=False)
                    self._conn.execute(ins)
                    transaction.commit()
                    success = True
                except sa.exc.IntegrityError:
                    transaction.rollback()
                except sa.exc.CompileError:
                    self._no_bulk = True
                if success:
                    return
            for table_filename, line_num in \
                    sorted(self._line_rows_dict.keys()):
                row_data = self._line_rows_dict[(table_filename, line_num)]
                try:
                    ins = sa.insert(self._sa_table)
                    self._conn.execute(ins.execution_options(autocommit=True),
                                       row_data)
                except sa.exc.IntegrityError:
                    ComplaintDataDuplication(
                        table=os.path.basename(table_filename)[: 2],
                        line_num=line_num, row_data=row_data)

    def __init__(self, db_filename: str,
                 day_dirs: Dict[Optional[datetime.datetime], str],
                 skip_write: bool, force_pk: bool, relaxed: bool) -> None:
        """ Constructor

        Arguments:
        db_filename -- Target database filename
        day_dirs    -- Directories where full file and updates located. Return
                       value Downloader.download()
        skip_write  -- True to not perform database writing operations
        force_pk    -- True to create primary keys for all tables (e.g. for
                       reporting purposes)
        relaxed     -- True to do not create any primary keys and unique
                       indices (to be very close to source data)
        """
        self._day_dirs = day_dirs
        self._skip_write = skip_write
        assert not os.path.isfile(db_filename)
        engine = sa.create_engine("sqlite:///" + db_filename)
        self._metadata = sa.MetaData()
        for uls_table in UlsTable.all():
            args: List[Any] = \
                [f.col_args.apply(sa.Column) for f in uls_table.fields
                 if f.col_args] + \
                ([sa.PrimaryKeyConstraint(*uls_table.id_fields)]
                 if (uls_table.make_primary_key or force_pk) and (not relaxed)
                 else [])
            for idx in uls_table.indices:
                if (not force_pk) and relaxed:
                    idx.unique = False
                args.append(idx)
            sa_table = sa.Table(uls_table.name, self._metadata, *args)
            assert all((not getattr(sa_table.c, f).nullable)
                       for f in uls_table.id_fields)
            self._metadata.create_all(engine)
        self._conn = engine.connect()

    def __enter__(self) -> "UlsDatabaseFiller":
        """ Context manager entry. Returns self """
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
        """ Context manager exit. Closes database connection """
        self._conn.close()

    def load_table(
            self, name: str,
            row_filter: Optional[Callable[[int, ROW_DATA_TYPE],
                                          bool]] = None) -> None:
        """ Loads given database table from FCC ULS files

        Arguments:
        name        -- Table name
        row_filter  -- Optional filter function that determines what rows to
                       load. Arguments are line number and row data dictionary
        """
        row_filter1: Callable[[int, ROW_DATA_TYPE], bool] = \
            row_filter or (lambda ln, x: True)
        with Timing((f"Filling {name} table")), \
                self._TableWriter(conn=self._conn,
                                  metadata=self._metadata, name=name,
                                  skip_write=self._skip_write) as tw:
            for table_filename, line_num, row_data, _ in \
                    UlsTableReader(uls_table=UlsTable.by_name(name),
                                   day_dirs=self._day_dirs):
                if not row_filter1(line_num, row_data):
                    continue
                tw.write(table_filename=table_filename, line=line_num,
                         row_data=row_data)

    def get_paths_from_pa(self) \
            -> Tuple[Set[Tuple[str, int]], Set[Tuple[str, int, int]]]:
        """ Retrieves path information from PA table

        Returns (paths, antennas) tuple, where 'paths' is a set of
        (callsign, path_num) tuples and 'antennas' is set of
        (callsign, loc_num, ant_num) tuples """
        paths: Set[Tuple[str, int]] = set()
        tx_antennas: Set[Tuple[str, int, int]] = set()
        pa = self._metadata.tables["PA"]
        s = sa.select(
            [pa.c.call_sign, pa.c.path_number, pa.c.transmit_location_number,
             pa.c.transmit_antenna_number])
        callsign: str
        path_num: int
        tx_loc_num: int
        tx_ant_num: int
        for callsign, path_num, tx_loc_num, tx_ant_num in \
                self._conn.execute(s):
            paths.add((callsign, path_num))
            tx_antennas.add((callsign, tx_loc_num, tx_ant_num))
        return (paths, tx_antennas)

    def get_paths_from_pa_em(self) \
            -> Tuple[Set[Tuple[str, int]], Set[Tuple[str, int, int]]]:
        """ From PA retrieves information about paths and antennas that have
        counterparts in (already purged) EM

        Returns (paths, antennas) tuple, where 'paths' is a set of
        (callsign, path_num) tuples and 'antennas' is set of
        (callsign, loc_num, ant_num) tuples """
        paths: Set[Tuple[str, int]] = set()
        tx_antennas: Set[Tuple[str, int, int]] = set()
        pa = self._metadata.tables["PA"]
        em = self._metadata.tables["EM"]
        s = sa.select(
            [pa.c.call_sign, pa.c.path_number, pa.c.transmit_location_number,
             pa.c.transmit_antenna_number]).\
            select_from(
                pa.join(
                    em,
                    (pa.c.call_sign == em.c.call_sign) &
                    (pa.c.transmit_location_number == em.c.location_number) &
                    (pa.c.transmit_antenna_number == em.c.antenna_number)))
        callsign: str
        path_num: int
        tx_loc_num: int
        tx_ant_num: int
        for callsign, path_num, tx_loc_num, tx_ant_num in \
                self._conn.execute(s):
            paths.add((callsign, path_num))
            tx_antennas.add((callsign, tx_loc_num, tx_ant_num))
        return (paths, tx_antennas)

    def get_paths_from_pa_sg(self) \
            -> Tuple[Set[str], Set[str], Set[Tuple[str, int]],
                     Set[Tuple[str, int, int]]]:
        """ Retrieves various information from (already purged) PA and SG

        Return (tx_callsigns, rx_callsigns, locations, antennas) tuple where
        'tx_callsigns' is set of TX callsigns, 'rx_callsigns' is set or RX
        callsigns, specified in PA table (not all records have them),
        'locations' is a set of (callsign, loc_num) tuples for all
        involved parties (receivers, transmitters, repeaters), 'antennas is a
        set of (callsign, loc_num, ant_num) tuples for all involved parties
        (receivers, transmitters, repeaters) """
        tx_callsigns: Set[str] = set()
        rx_callsigns: Set[str] = set()
        locations: Set[Tuple[str, int]] = set()
        antennas: Set[Tuple[str, int, int]] = set()
        pa = self._metadata.tables["PA"]
        sg = self._metadata.tables["SG"]
        s = sa.select([sg.c.call_sign, sg.c.transmit_location,
                       sg.c.transmit_antenna, sg.c.receiver_location,
                       sg.c.receiver_antenna])
        for callsign, tx_loc_num, tx_ant_num, rx_loc_num, rx_ant_num in \
                self._conn.execute(s):
            tx_callsigns.add(callsign)
            locations.add((callsign, tx_loc_num))
            locations.add((callsign, rx_loc_num))
            antennas.add((callsign, tx_loc_num, tx_ant_num))
            antennas.add((callsign, rx_loc_num, rx_ant_num))
        s = sa.select([pa.c.receiver_callsign]).\
            where(pa.c.receiver_callsign != None)  # noqa
        for rx_callsign in self._conn.execute(s):
            rx_callsigns.add(rx_callsign[0])
        return (tx_callsigns, rx_callsigns, locations, antennas)

    def purge_paths_by_em(self) -> None:
        """ From PA table remove rows that do not correspondent rows in EM
        table """
        transaction = self._conn.begin()
        self._conn.execute(
            "DELETE FROM PA WHERE NOT EXISTS" +
            "(SELECT NULL FROM EM " +
            "WHERE (PA.call_sign = EM.call_sign) AND " +
            "(PA.transmit_location_number = EM.location_number) AND " +
            "(PA.transmit_antenna_number = EM.antenna_number))")
        transaction.commit()


class DownloadedPathIterator:
    """ Iterator over paths in downloaded FCC ULS files

    Private attributes:
    _day_dirs          -- Directories with downloaded FCC ULS files (return
                          value of Downloader.download())
    _paths_by_path     -- Index of _AfcPathInfo objects by paths
                          ((callsign, path_num) tuples)
    _paths_by_tx       -- Index of _AfcPathInfo objects by transmitters
                          ((callsign, loc_num, ant_num) tuples
    _paths_by_callsign -- Dictionary of lists of _AfcPathInfo objects, indexed
                          by callsign
    _locations         -- Dictionary of relevant excerpts from LO records,
                          indexed by (callsign, location) tuples
    _antennas          -- Dictionary of relevant excerpts from AN records,
                          indexed by (callsign, path_number, location, antenna)
                          tuples
    """
    class _AfcPathInfo:
        """ Bucket of data about path

        Attributes:
        callsign              -- Path callsign
        path_num              -- Path number
        tx_loc_num            -- TX location number
        tx_ant_num            -- TX antenna number
        rx_loc_num            -- RX location number
        rx_ant_num            -- RX antenna number
        p_rx_indicator        -- Passive repeater presence mark (None/'Y'/'N')
        rx_callsign           -- RX callsign (path callsign, if not specified)
        tx_eirp               -- Maximum TX EIRP (initially None)
        license_status        -- License status (initially None)
        radio_service         -- Radio service string (initially None)
        entity_name           -- Entity name (initially None)
        common_carrier        -- Common carrier string (initially None)
        mobile                -- Mobile flag string (initially None)
        grant_date            -- License grant date string (initially None)
        expired_date          -- License expired date string (initially None)
        cancellation_date     -- License cancellation date string (initially
                                 None)
        em_des_by_num_by_freq -- Lists of emission descriptors, ordered by
                                 frequency numbers, then by frequency
        freq_by_hash          -- FreqRange object, indexed by hash (initially
                                 empty)
        segments              -- List of Segment objects. Initially empty
        """
        class Segment:
            """ Information about transmission segment

            Attributes:
            tx_loc_num -- TX location number
            tx_ant_num -- TX antenna number
            rx_loc_num -- RX location number
            rx_ant_num -- RX antenna number
            """
            def __init__(self, tx_loc_num: int, tx_ant_num: int,
                         rx_loc_num: int, rx_ant_num: int) -> None:
                self.tx_loc_num = tx_loc_num
                self.tx_ant_num = tx_ant_num
                self.rx_loc_num = rx_loc_num
                self.rx_ant_num = rx_ant_num

        def __init__(self, callsign: str, path_num: int, tx_loc_num: int,
                     tx_ant_num: int, rx_loc_num: int, rx_ant_num: int,
                     rx_callsign: Optional[str],
                     p_rx_indicator: Optional[str]) -> None:
            """ Constructor

            Arguments:
            callsign       -- Callsign
            path_num       -- Path number
            tx_loc_num     -- TX location number
            tx_ant_num     -- TX antenna number
            rx_loc_num     -- RX location number
            rx_ant_num     -- RX antenna number
            rx_callsign    -- RX callsign from PA row or None
            p_rx_indicator -- Passive repeater presence mark (None/'Y'/'N')
            """
            self.callsign = callsign
            self.path_num = path_num
            self.tx_loc_num = tx_loc_num
            self.tx_ant_num = tx_ant_num
            self.rx_loc_num = rx_loc_num
            self.rx_ant_num = rx_ant_num
            self.rx_callsign = rx_callsign or callsign
            self.p_rx_indicator = p_rx_indicator
            self.tx_eirp: Optional[float] = None
            self.license_status: Optional[str] = None
            self.radio_service: Optional[str] = None
            self.entity_name: Optional[str] = None
            self.common_carrier: Optional[str] = None
            self.mobile: Optional[str] = None
            self.grant_date: Optional[str] = None
            self.expired_date: Optional[str] = None
            self.cancellation_date: Optional[str] = None
            self.em_des_by_num_by_freq: Dict[int, Dict[float, List[str]]] = {}
            self.freq_by_hash: Dict[int, FreqRange] = {}
            self.segments: \
                Dict[int, "DownloadedPathIterator._AfcPathInfo.Segment"] = {}

        def callsign_key(self) -> str:
            """ Key for DownloadedPathIterator._paths_by_callsign """
            return self.callsign

        def path_key(self) -> Tuple[str, int]:
            """ Key for DownloadedPathIterator._paths_by_path """
            return (self.callsign, self.path_num)

        def tx_key(self) -> Tuple[str, int, int]:
            """ Key for DownloadedPathIterator._paths_by_tx """
            return (self.callsign, self.tx_loc_num, self.tx_ant_num)

        def __str__(self):
            """ String representation (for debugging) """
            return (f"{self.callsign}/{self.path_num or 0}. "
                    f"TX: {self.tx_loc_num}/{self.tx_ant_num}. "
                    f"RX: {self.rx_loc_num}/{self.rx_ant_num}.")

        def __repr__(self):
            """ Debugger representation (for debugging) """
            return f"<{self}>"

    class _AntInfo:
        """ Antenna information

        Attributes:
        ant_cord             -- Optional PathInfo.AntCoord object
        model                -- Optional antenna model string
        polarization         -- Optional antenna polarization string
        gain                 -- Optional antenna gain value
        reflector_height     -- Optional reflector height
        reflector_width      -- Optional reflector width
        back_to_back_gain_tx -- Optional back-to-back TX gain
        back_to_back_gain_rx -- Optional back-to-back RX gain
        """
        def __init__(
                self, ant_coord: Optional[PathInfo.AntCoord] = None,
                model: Optional[str] = None,
                polarization: Optional[str] = None,
                gain: Optional[float] = None,
                reflector_height: Optional[float] = None,
                reflector_width: Optional[float] = None,
                back_to_back_gain_tx: Optional[float] = None,
                back_to_back_gain_rx: Optional[float] = None) -> None:
            """ Constructor

            Arguments:
            ant_cord             -- Optional PathInfo.AntCoord object
            model                -- Optional antenna model string
            polarization         -- Optional antenna polarization string
            gain                 -- Optional antenna gain value
            reflector_height     -- Optional reflector height
            reflector_width      -- Optional reflector width
            back_to_back_gain_tx -- Optional back-to-back TX gain
            back_to_back_gain_rx -- Optional back-to-back RX gain
            """
            self.ant_coord = ant_coord
            self.model = model
            self.polarization = polarization
            self.gain = gain
            self.reflector_height = reflector_height
            self.reflector_width = reflector_width
            self.back_to_back_gain_tx = back_to_back_gain_tx
            self.back_to_back_gain_rx = back_to_back_gain_rx

    def __init__(self, day_dirs: Dict[Optional[datetime.datetime], str]) \
            -> None:
        """ Constructor (does all read-in)

        Arguments:
        day_dirs     -- Directories where full file and updates located. Return
                        value Downloader.download()
        """
        self._day_dirs = day_dirs
        self._paths_by_path: Dict[Tuple[str, int],
                                  "DownloadedPathIterator._AfcPathInfo"] = {}
        self._paths_by_tx: \
            Dict[Tuple[str, int, int],
                 List["DownloadedPathIterator._AfcPathInfo"]] = {}
        self._paths_by_callsign: \
            Dict[str, List["DownloadedPathIterator._AfcPathInfo"]] = {}
        self._locations: Dict[Tuple[str, int], ROW_DATA_TYPE] = {}
        self._antennas: Dict[Tuple[str, int, int, int], ROW_DATA_TYPE] = {}
        self._process_pa()
        self._process_hd()
        self._process_em()
        self._process_fr()
        self._purge_nonafc()
        self._process_en()
        self._process_sg()
        self._prepare_loc_ant()
        self._process_lo()
        self._process_an()

    def _process_pa(self) -> None:
        """ Reads paths from PA table """
        with Timing("Reading 'PA' data"):
            for _, line_num, row_data, _ in \
                    UlsTableReader(day_dirs=self._day_dirs,
                                   uls_table=UlsTable.by_name("PA")):
                key = (row_data["call_sign"], row_data["path_number"])
                if key in self._paths_by_path:
                    ComplaintDataDuplication(
                        table="PA", line_num=line_num, row_data=row_data)
                    continue
                afc_path = \
                    self._AfcPathInfo(
                        callsign=row_data["call_sign"],
                        path_num=row_data["path_number"],
                        tx_loc_num=row_data["transmit_location_number"],
                        tx_ant_num=row_data["transmit_antenna_number"],
                        rx_loc_num=row_data["receiver_location_number"],
                        rx_ant_num=row_data["receiver_antenna_number"],
                        rx_callsign=row_data["receiver_callsign"],
                        p_rx_indicator=row_data["passive_receiver_indicator"])
                self._paths_by_path[afc_path.path_key()] = afc_path
                self._paths_by_tx.setdefault(afc_path.tx_key(),
                                             []).append(afc_path)
                self._paths_by_callsign.setdefault(afc_path.callsign_key(),
                                                   []).append(afc_path)

    def _process_hd(self) -> None:
        """ Reads license information from HD table, removes unlicensed paths
        """
        with Timing("Reading 'HD' data"):
            for _, line_num, row_data, _ in \
                    UlsTableReader(day_dirs=self._day_dirs,
                                   uls_table=UlsTable.by_name("HD")):
                afc_paths = self._paths_by_callsign.get(row_data["call_sign"])
                if not afc_paths:
                    continue
                if afc_paths[0].license_status is not None:
                    ComplaintDataDuplication(
                        table="HD", line_num=line_num, row_data=row_data)
                    continue
                if row_data["license_status"] not in "AL":
                    self._remove_afc_paths_by_callsign(row_data["call_sign"])
                else:
                    for afc_path in afc_paths:
                        afc_path.license_status = row_data["license_status"]
                        afc_path.radio_service = row_data["radio_service_code"]
                        afc_path.common_carrier = row_data["common_carrier"]
                        afc_path.mobile = row_data["mobile"]
                        afc_path.grant_date = row_data["grant_date"]
                        afc_path.expired_date = row_data["expired_date"]
                        afc_path.cancellation_date = \
                            row_data["cancellation_date"]

    def _process_em(self) -> None:
        """ Reads emission information from EM table """
        with Timing("Reading 'EM' data"):
            for _, line_num, row_data, _ in \
                    UlsTableReader(day_dirs=self._day_dirs,
                                   uls_table=UlsTable.by_name("EM")):
                key = (row_data["call_sign"], row_data["location_number"],
                       row_data["antenna_number"])
                if key not in self._paths_by_tx:
                    continue
                if not row_data["emission_code"]:
                    continue
                em_bw_mhz = \
                    FreqRange.emission_bw_mhz(mst(row_data["emission_code"]))
                if em_bw_mhz is None:
                    ComplaintEmissionCode(
                        callsign=row_data["call_sign"],
                        emission_code=row_data["emission_code"],
                        line_num=line_num)
                    continue
                for afc_path in self._paths_by_tx[key]:
                    afc_path.em_des_by_num_by_freq.setdefault(
                        row_data["frequency_number"], {}).\
                        setdefault(row_data["frequency_assigned"], []).\
                        append(row_data["emission_code"])

    def _process_fr(self):
        """ Obtains power (EIRP) information from FR table """
        with Timing("Reading 'FR' data"):
            for _, line_num, row_data, _ in \
                    UlsTableReader(day_dirs=self._day_dirs,
                                   uls_table=UlsTable.by_name("FR")):
                key = (row_data["call_sign"], row_data["location_number"],
                       row_data["antenna_number"])
                if key not in self._paths_by_tx:
                    continue
                for afc_path in self._paths_by_tx[key]:
                    by_freq_dict = \
                        afc_path.em_des_by_num_by_freq.get(
                            row_data["freq_seq_id"], {})
                    emission_codes: List[str]
                    if row_data["frequency_upper_band"] is not None:
                        # Frequency range specified. Emission codes(s)
                        # retrieved by frequency number
                        emission_codes = \
                            sum(by_freq_dict.values(), [])
                    else:
                        # Center frequency specuified. Emission code(s)
                        # retrieved by frequency number and center frequency
                        emission_codes = \
                            by_freq_dict.get(row_data["frequency_assigned"],
                                             [])
                    if not emission_codes:
                        ComplaintMissingEmission(
                            callsign=row_data["call_sign"],
                            loc_num=row_data["location_number"],
                            ant_num=row_data["antenna_number"],
                            freq_num=row_data["freq_seq_id"],
                            center_freq=None
                            if row_data["frequency_upper_band"] is not None
                            else row_data["frequency_assigned"],
                            line_num=line_num)
                    for emission_code in emission_codes:
                        if row_data["frequency_upper_band"] is not None:
                            freq_range = \
                                FreqRange(
                                    emission_code=emission_code,
                                    start_mhz=mf(
                                        row_data["frequency_assigned"]),
                                    end_mhz=mf(
                                        row_data["frequency_upper_band"]))
                        elif FreqRange.emission_bw_mhz(emission_code) \
                                is not None:
                            freq_range = \
                                FreqRange(
                                    emission_code=emission_code,
                                    center_mhz=row_data["frequency_assigned"])
                        if freq_range and freq_range.is_6ghz():
                            afc_path.freq_by_hash[hash(freq_range)] = \
                                freq_range
                    tx_eirp = row_data["eirp"]
                    if (tx_eirp is not None) and \
                            ((afc_path.tx_eirp is None) or
                             (tx_eirp > afc_path.tx_eirp)):
                        afc_path.tx_eirp = tx_eirp

    def _purge_nonafc(self):
        """ Purges paths not fit for AFC database (non-6GHz, not having license
        information, not having 6GHz emissions """
        with Timing("Pruning non-6GHz/enabled paths"):
            callsigns_to_remove: List[str] = []
            transmitters_to_remove: Dict[str, Set[Tuple[int, int]]] = {}
            for callsign, afc_paths in self._paths_by_callsign.items():
                if afc_paths[0].license_status is None:
                    callsigns_to_remove.append(callsign)
                    continue
                ttr: Set[Tuple[int, int]] = set()
                for afc_path in afc_paths:
                    if not afc_path.freq_by_hash:
                        ttr.add((afc_path.tx_loc_num, afc_path.tx_ant_num))
                if ttr:
                    transmitters_to_remove[callsign] = ttr
            for callsign in callsigns_to_remove:
                self._remove_afc_paths_by_callsign(callsign)
            for callsign, ttr in transmitters_to_remove.items():
                self._remove_afc_paths_by_tx(callsign=callsign, ttr=ttr)

    def _remove_afc_paths_by_tx(self, callsign: str,
                                ttr: Set[Tuple[int, int]]) -> None:
        """ Removes paths by transmitter

        Arguments:
        callsign -- Callsign of paths to remove
        ttr      -- Set of (tx_loc_num, tx_ant_num) tuples of transmitters to
                    remove
        """
        paths = self._paths_by_callsign[callsign]
        idx = 0
        while idx < len(paths):
            afc_path = paths[idx]
            if (afc_path.tx_ant_num, afc_path.tx_loc_num) in ttr:
                del self._paths_by_path[afc_path.path_key()]
                paths.pop(idx)
            else:
                idx += 1
        for tx_loc_num, tx_ant_num in ttr:
            del self._paths_by_tx[(callsign, tx_loc_num, tx_ant_num)]
        if not paths:
            del self._paths_by_callsign[callsign]

    def _remove_afc_paths_by_callsign(self, callsign: str) -> None:
        """ Removes paths of given callsign """
        removed_tx_keys: Set[Tuple[str, int, int]] = set()
        for afc_path in self._paths_by_callsign[callsign]:
            del self._paths_by_path[afc_path.path_key()]
            if afc_path.tx_key() not in removed_tx_keys:
                del self._paths_by_tx[afc_path.tx_key()]
                removed_tx_keys.add(afc_path.tx_key())
        del self._paths_by_callsign[callsign]

    def _process_en(self):
        """ Retrieves entity names from EN table """
        with Timing("Reading 'EN' data"):
            for _, _, row_data, _ in \
                    UlsTableReader(day_dirs=self._day_dirs,
                                   uls_table=UlsTable.by_name("EN")):
                afc_paths = self._paths_by_callsign.get(row_data["call_sign"])
                if not afc_paths:
                    continue
                for afc_path in afc_paths:
                    afc_path.entity_name = row_data["entity_name"]

    def _process_sg(self):
        """ Retrieves segment (repeater) info from SG table """
        with Timing("Reading 'SG' data"):
            for _, line_num, row_data, _ in \
                    UlsTableReader(day_dirs=self._day_dirs,
                                   uls_table=UlsTable.by_name("SG")):
                afc_path = self._paths_by_path.get((row_data["call_sign"],
                                                    row_data["path_number"]))
                if afc_path is None:
                    continue
                if row_data["segment_number"] in afc_path.segments:
                    ComplaintDataDuplication(
                        table="SG", line_num=line_num, row_data=row_data)
                    continue
                afc_path.segments[row_data["segment_number"]] = \
                    self._AfcPathInfo.Segment(
                        tx_loc_num=row_data["transmit_location"],
                        tx_ant_num=row_data["transmit_antenna"],
                        rx_loc_num=row_data["receiver_location"],
                        rx_ant_num=row_data["receiver_antenna"])

    def _prepare_loc_ant(self):
        """ In _locations and _antennas builds collections of required
        locations and antennas """
        with Timing("Pruning unrelated locations/antennas"):
            for afc_path in self._paths_by_path.values():
                for tx_loc_num, tx_ant_num, rx_loc_num, rx_ant_num in \
                        [(afc_path.tx_loc_num, afc_path.tx_ant_num,
                          afc_path.rx_loc_num, afc_path.rx_ant_num)] + \
                        [(seg.tx_loc_num, seg.tx_ant_num, seg.rx_ant_num,
                          seg.tx_ant_num)
                         for seg in afc_path.segments.values()]:
                    for loc_num, ant_num in [(tx_loc_num, tx_ant_num),
                                             (rx_loc_num, rx_ant_num)]:
                        self._locations[(afc_path.callsign, loc_num)] = None
                        self._antennas[(afc_path.callsign,
                                        afc_path.path_num, loc_num,
                                        ant_num)] = None

    def _process_lo(self) -> None:
        """ Reads locations from LO table """
        lo_fields = ("lat_degrees", "lat_minutes", "lat_seconds",
                     "lat_direction", "long_degrees", "long_minutes",
                     "long_seconds", "long_direction", "ground_elevation",
                     "location_type_code")
        with Timing("Reading 'LO' data"):
            for _, line_num, row_data, _ in \
                    UlsTableReader(day_dirs=self._day_dirs,
                                   uls_table=UlsTable.by_name("LO")):
                key = (row_data["call_sign"], row_data["location_number"])
                if key not in self._locations:
                    continue
                if self._locations[key]:
                    ComplaintDataDuplication(
                        table="LO", line_num=line_num, row_data=row_data)
                    self._locations[key]["duplicate"] = True
                    continue
                self._locations[key] = \
                    {field: row_data[field] for field in lo_fields}

    def _process_an(self) -> None:
        """ Reads antenna data from AN table """
        an_fields = ("polarization_code", "height_to_center_raat", "gain",
                     "antenna_model", "reflector_height", "reflector_width",
                     "back_to_back_tx_dish_gain", "back_to_back_rx_dish_gain")
        with Timing("Reading 'AN' data"):
            for _, line_num, row_data, _ in \
                    UlsTableReader(day_dirs=self._day_dirs,
                                   uls_table=UlsTable.by_name("AN")):
                key = (row_data["call_sign"], row_data["path_number"],
                       row_data["location_number"], row_data["antenna_number"])
                if key not in self._antennas:
                    continue
                if self._antennas[key]:
                    ComplaintDataDuplication(
                        table="AN", line_num=line_num, row_data=row_data)
                    self._antennas[key]["duplicate"] = True
                    continue
                self._antennas[key] = \
                    {field: row_data[field] for field in an_fields}

    def __iter__(self) -> Iterator:
        """ Iterator. Yields PathInfo objects """
        row_count = 0
        for callsign in sorted(self._paths_by_callsign.keys()):
            afc_paths = self._paths_by_callsign[callsign]
            for afc_path in sorted(afc_paths, key=lambda path: path.path_num):
                repeaters: List[PathInfo.Repeater] = []
                repeater_rx_ai: Optional["DownloadedPathIterator._AntInfo"] = \
                    None
                drop = False
                expected_idx = 1
                # Populating repeater list
                for seg_idx, segment in sorted(afc_path.segments.items()):
                    if seg_idx != expected_idx:
                        # Segment index absent in SG
                        ComplaintMissingSegment(
                            callsign=callsign, path_num=afc_path.path_num,
                            seg_num=expected_idx)
                        drop = True
                        break
                    expected_idx += 1
                    if seg_idx > 1:
                        assert repeater_rx_ai is not None
                        assert repeater_rx_ai.ant_coord is not None
                        repeater_tx_ai = \
                            self._ant_info(
                                callsign=callsign, path_num=afc_path.path_num,
                                loc_num=segment.tx_loc_num,
                                ant_num=segment.tx_ant_num)
                        if not repeater_tx_ai.ant_coord:
                            drop = True
                            break
                        # Repeater information from ingress and egress segments
                        repeater_rx_tx: List[PathInfo.Repeater] = []
                        for ai in (repeater_rx_ai, repeater_tx_ai):
                            assert ai is not None
                            repeater_rx_tx.append(
                                PathInfo.Repeater(
                                    ant_coord=ai.ant_coord,
                                    reflector_height=ai.reflector_height,
                                    reflector_width=ai.reflector_width,
                                    back_to_back_gain_tx=ai.
                                    back_to_back_gain_tx,
                                    back_to_back_gain_rx=ai.
                                    back_to_back_gain_rx))
                        consistent, repeater = \
                            PathInfo.Repeater.combine(repeater_rx_tx[0],
                                                      repeater_rx_tx[1])
                        if not consistent:
                            # Ingress and egress repeater data mismatch
                            ComplaintInconsistentRepeater(
                                callsign=callsign, path_num=afc_path.path_num,
                                repeater_idx=seg_idx - 1)
                        assert repeater is not None
                        repeaters.append(repeater)
                    if seg_idx < len(afc_path.segments):
                        repeater_rx_ai = \
                            self._ant_info(
                                callsign=callsign, path_num=afc_path.path_num,
                                loc_num=segment.rx_loc_num,
                                ant_num=segment.rx_ant_num)
                        if not repeater_rx_ai.ant_coord:
                            drop = True
                            break
                if drop:
                    continue
                assert afc_path.license_status is not None
                assert afc_path.radio_service is not None
                tx_ant_info = \
                    self._ant_info(
                        callsign=callsign, path_num=afc_path.path_num,
                        loc_num=afc_path.tx_loc_num,
                        ant_num=afc_path.tx_ant_num)
                if not tx_ant_info.ant_coord:
                    continue
                rx_ant_info = \
                    self._ant_info(
                        callsign=callsign, path_num=afc_path.path_num,
                        loc_num=afc_path.rx_loc_num,
                        ant_num=afc_path.rx_ant_num)
                if not rx_ant_info.ant_coord:
                    continue
                path_freq_bucket = PathFreqBucket()
                for freq_range in afc_path.freq_by_hash.values():
                    path = \
                        PathInfo(
                            callsign=callsign, status=afc_path.license_status,
                            radio_service=afc_path.radio_service,
                            name=afc_path.entity_name,
                            common_carrier=afc_path.common_carrier,
                            mobile=afc_path.mobile,
                            rx_callsign=afc_path.rx_callsign,
                            rx_antenna_num=afc_path.rx_ant_num,
                            freq_range=freq_range, tx_eirp=afc_path.tx_eirp,
                            tx_ant=tx_ant_info.ant_coord,
                            tx_polarization=tx_ant_info.polarization,
                            tx_gain=tx_ant_info.gain,
                            rx_ant=rx_ant_info.ant_coord,
                            rx_ant_model=rx_ant_info.model,
                            rx_gain=rx_ant_info.gain,
                            repeaters=repeaters,
                            p_rx_indicator=afc_path.p_rx_indicator,
                            path_num=afc_path.path_num,
                            grant_date=afc_path.grant_date,
                            expired_date=afc_path.expired_date,
                            cancellation_date=afc_path.cancellation_date)
                    path_freq_bucket.try_add(path)
                for p in path_freq_bucket:
                    row_count += 1
                    if (row_count % 10000) == 0:
                        Progressor.print_progress(f"{row_count}")
                    yield p

    def _ant_info(self, callsign: str, path_num: int, loc_num: int,
                  ant_num: int) -> "DownloadedPathIterator._AntInfo":
        """ Builds _AntInfo for given antenna

        Arguments:
        callsign -- Path callsign
        loc_num  -- Location number
        ant_num  -- Antenna number
        """
        location: Optional[ROW_DATA_TYPE] = \
            self._locations.get((callsign, loc_num))
        antenna: Optional[ROW_DATA_TYPE] = \
            self._antennas.get((callsign, path_num, loc_num, ant_num))
        ret = \
            self._AntInfo(
                ant_coord=PathInfo.AntCoord(
                    lat_deg=oi(location["lat_degrees"]) if location else None,
                    lat_min=oi(location["lat_minutes"]) if location else None,
                    lat_sec=of(location["lat_seconds"]) if location else None,
                    lat_dir=ost(location["lat_direction"]) if location
                    else None,
                    lon_deg=oi(location["long_degrees"]) if location else None,
                    lon_min=oi(location["long_minutes"]) if location else None,
                    lon_sec=of(location["long_seconds"]) if location else None,
                    lon_dir=ost(location["long_direction"]) if location
                    else None,
                    ground_elev_m=of(location["ground_elevation"]) if location
                    else None,
                    height_to_center_raat_m=of(
                        antenna["height_to_center_raat"]) if antenna else None,
                    loc_type_code=ost(location["location_type_code"])
                    if location else None),
                model=ost(antenna["antenna_model"]) if antenna else None,
                gain=of(antenna["gain"]) if antenna else None,
                polarization=ost(antenna["polarization_code"])
                if antenna else None,
                reflector_height=of(antenna["reflector_height"])
                if antenna else None,
                reflector_width=of(antenna["reflector_width"])
                if antenna else None,
                back_to_back_gain_tx=of(antenna["back_to_back_tx_dish_gain"])
                if antenna else None,
                back_to_back_gain_rx=of(antenna["back_to_back_rx_dish_gain"])
                if antenna else None)
        return ret


class KmlWriter:
    """ KML file writer
    Intended to be used as context manager (with 'with' clause)

    Private attributes:
    _filename    -- Name of .kml file to write to. None to do nothing
    _comment     -- Map comment or None
    _map_points  -- Dictionary of MapPoint objects, indexed by
                    (latitude, longitude) tuples
    _map_segs    -- Dictionary of MapSeg objects, indexed by
                    (src_lat, src_lon, dst_lat, dst_lon) tuples
    _center_lat  -- Optional latitude of center of circle to draw
    _center_lon  -- Optional longitude of center of circle to draw
    _range_miles -- Optional radius in miles of circle to draw
    _arrows      -- True to draw arrows
    """
    # Maximum feature count in KML file
    KML_MAX_FEATURES = 2000

    # KML colors
    C_RED = "0000ff"
    C_BLUE = "ff0000"
    C_BLACK = "000000"
    C_YELLOW = "00ffff"
    C_GREEN = "00ff00"

    class AntInfo:
        """ Information about single antenna

        Attributes:
        path_name   -- Path name
        role        -- Role string
        raat        -- Height above ground in meters
        rx_callsign -- RX callsign (if pertinent, specified and different from
                       path callsign) or None
        """
        def __init__(self, path: PathInfo, role: str, raat: float,
                     rx_callsign: Optional[str] = None) -> None:
            """ Constructor

            Arguments:
            path        -- Path this antenna belongs to
            role        -- Role string
            raat        -- Height above ground in meters
            rx_callsign -- RX callsign (if pertinent, specified and different
                           from path callsign) or None
            """
            self.path_name = path.path_name()
            self.role = role
            self.raat = raat
            self.rx_callsign = \
                rx_callsign if rx_callsign != path.callsign else None

        def description(self) -> str:
            """ More detailed description """
            rcs = self.rx_callsign
            return \
                (f"{self.path_name}{'' if rcs is None else ('-' + rcs)} "
                 f"{self.role}: {self.raat}m")

        def sort_key(self) -> Any:
            """ Sort key for point description """
            return (self.path_name, self.raat, self.role)

        def __hash__(self) -> int:
            """ Hash implementation """
            return \
                hash((self.path_name, self.role, self.raat, self.rx_callsign))

        def __eq__(self, other: Any) -> bool:
            """ Equality comparison """
            return isinstance(other, self.__class__) and \
                (self.__dict__ == other.__dict__)

    class MapPoint:
        """ Information about map point

        Private attributes:
        _antennas -- Set of AntInfo objects
        """
        def __init__(self) -> None:
            """ Constructor """
            self._antennas: Set["KmlWriter.AntInfo"] = set()

        def add(self, ant_info: "KmlWriter.AntInfo") -> None:
            """ Adds antenna information """
            self._antennas.add(ant_info)

        def name(self) -> str:
            """ Map name for location """
            assert self._antennas
            top_ant = None
            for ant in self._antennas:
                if (top_ant is None) or (ant.raat > top_ant.raat):
                    top_ant = ant
            assert top_ant is not None
            return (f"{top_ant.path_name} {top_ant.role}"
                    f"{'*' if len(self._antennas) > 1 else ''}")

        def raat(self) -> float:
            """ Height above ground - of lowest antenna """
            assert self._antennas
            bottom_ant = None
            for ant in self._antennas:
                if (bottom_ant is None) or (ant.raat < bottom_ant.raat):
                    bottom_ant = ant
            assert bottom_ant is not None
            return bottom_ant.raat

        def description(self) -> str:
            """ Map description of point """
            return \
                ", ".join(ant.description()
                          for ant in
                          sorted(self._antennas,
                                 key=operator.methodcaller("sort_key")))

    class MapSeg:
        """ Map segment

        Public attributes:
        src_raat  -- Minimum antenna height at source point
        dst_raat  -- Minimum antenna height at destination point
        src_arrow -- Arrow should be drawn at source point
        dst_arrow -- Arrow should be drawn at destination point
        color     -- 6-character hexadecimal KML color string ("bbggrr")

        Private attributes:
        _paths -- Dictionary of per-feature path names
        """

        Feature = enum.Enum("Feature", "Body Src Dst")

        def __init__(self) -> None:
            """ Constructor """
            self.src_raat: Optional[float] = None
            self.dst_raat: Optional[float] = None
            self.src_arrow = False
            self.dst_arrow = False
            self.color: Optional[str] = None
            self._paths: Dict["KmlWriter.MapSeg.Feature", Set[str]] = {}

        def add(self, path: PathInfo, src: PathInfo.AntCoord,
                dst: PathInfo.AntCoord, src_arrow: bool,
                dst_arrow: bool, color: str,
                color_priority: Optional[List[str]] = None) -> None:
            """ Add segment

            Arguments:
            path           -- Path segment belongs to
            src            -- Source point
            dst            -- Destination point
            src_arrow      -- Arrow should be drawn at source point
            dst_arrow      -- Arrow should be drawn at destination point
            color          -- 6-character hexadecimal KML color string
                              ("bbggrr")
            color_priority -- None or list of color strings that specify
                              resulting color (color with higher index beats
                              one with lower). None means most recent wins
            """
            if (self.src_raat is None) or \
                    (self.src_raat > mf(src.height_to_center_raat_m)):
                self.src_raat = mf(src.height_to_center_raat_m)
            if (self.dst_raat is None) or \
                    (self.dst_raat > mf(dst.height_to_center_raat_m)):
                self.dst_raat = mf(dst.height_to_center_raat_m)
            self._paths.setdefault(self.Feature.Body,
                                   set()).add(path.path_name())
            if src_arrow:
                self._paths.setdefault(self.Feature.Src,
                                       set()).add(path.path_name())
            if dst_arrow:
                self._paths.setdefault(self.Feature.Dst,
                                       set()).add(path.path_name())

            self.src_arrow |= src_arrow
            self.dst_arrow |= dst_arrow
            if (self.color is None) or (color_priority is None) or \
                    (color_priority.index(color) >
                     color_priority.index(self.color)):
                self.color = color

        def name(self, feature: "KmlWriter.MapSeg.Feature") -> str:
            """ Name for map of given feature """
            feature_paths = self._paths.get(feature)
            if not feature_paths:
                return ""
            return \
                f"{min(feature_paths)}{'*' if len(feature_paths) > 1 else ''}"

        def description(self, feature: "KmlWriter.MapSeg.Feature") -> str:
            """ Description for map of given feature """
            feature_paths = self._paths.get(feature)
            if not feature_paths:
                return ""
            return ", ".join(sorted(feature_paths))

        def num_features(self) -> int:
            """ Returns number of KML features in segment """
            return \
                1 + (1 if self.src_arrow else 0) + (1 if self.dst_arrow else 0)

    def __init__(self, filename: Optional[str], comment: Optional[str] = None,
                 center_lat: Optional[float] = None,
                 center_lon: Optional[float] = None,
                 range_miles: Optional[float] = None, arrows: bool = False) \
            -> None:
        """ Constructor

        Arguments:
        filename    -- Name of KML file to write to. None to do nothing
        comment     -- Map comment or None
        center_lat  -- Optional latitude of center of circle to draw
        center_lon  -- Optional longitude of center of circle to draw
        range_miles -- Optional radius in miles of circle to draw
        arrows      -- True to draw arrows
        """
        self._filename = filename
        self._comment = comment
        self._map_points: Dict[Tuple[float, float], "KmlWriter.MapPoint"] = {}
        self._map_segs: Dict[Tuple[float, float, float, float],
                             "KmlWriter.MapSeg"] = {}
        self._center_lat = center_lat
        self._center_lon = center_lon
        self._range_miles = range_miles
        self._arrows = arrows

    def add_path(self, path: PathInfo, color: str,
                 color_priority: Optional[List[str]] = None) -> None:
        """ Add given path to map

        Arguments:
        path           -- Path to add
        color          -- 6-character hexadecimal KML color string ("bbggrr")
        color_priority -- None or list of color strings that specify
                          resulting color (color with higher index beats
                          one with lower). None means most recent wins
       """
        self._add_ant_coord(path=path, ant_coord=path.tx_ant, role="TX")
        loc = path.tx_ant
        for repeater in path.repeaters:
            assert repeater.ant_coord is not None
            self._add_segment(path=path, src=loc, dst=repeater.ant_coord,
                              color=color, color_priority=color_priority)
            self._add_ant_coord(path=path, ant_coord=repeater.ant_coord,
                                role="Repeater")
            loc = repeater.ant_coord
        self._add_ant_coord(path=path, ant_coord=path.rx_ant, role="RX",
                            rx_callsign=path.rx_callsign)
        self._add_segment(path=path, src=loc, dst=path.rx_ant,
                          color=color, color_priority=color_priority)

    def num_features(self) -> int:
        """ Returns number of KML features already in collection """
        return (1 if self._has_circle() else 0) + \
            len(self._map_points) + \
            sum(s.num_features() for s in self._map_segs.values())

    def num_path_features(self, path: PathInfo) -> int:
        """ Returns number of KML features in given path """
        return (2 + 2 * len(path.repeaters)) + \
            ((1 + len(path.repeaters)) * (2 if self._arrows else 1))

    def __enter__(self) -> "KmlWriter":
        """ Context manager entry. Returns self  """
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, exc_tb: Any) -> None:
        """ Context manager exit """
        if self._filename is None:
            return
        root = \
            etree.Element("kml",
                          attrib={"xmlns": "http://www.opengis.net/kml/2.2"})
        document = etree.SubElement(root, "Document")
        if self._comment:
            self._make_tag(document, "name", text=self._comment)
        if self._has_circle():
            circle_style = self._make_tag(document, "Style",
                                          attrs={"id": "CircleStyle"})
            line_style = self._make_tag(circle_style, "LineStyle")
            self._make_tag(line_style, "color", text="20" + self.C_RED)
            self._make_tag(line_style, "width", text="2")
        if self._arrows:
            arrow_style = self._make_tag(document, "Style",
                                         attrs={"id": "ArrowStyle"})
            line_style = self._make_tag(arrow_style, "LineStyle")
            self._make_tag(line_style, "color", text="ff" + self.C_BLUE)
            self._make_tag(line_style, "width", text="3")
        for line_color in sorted({s.color for s in self._map_segs.values()}):
            seg_style = \
                self._make_tag(document, "Style",
                               attrs={"id": "SegmentStyle_" + line_color})
            line_style = self._make_tag(seg_style, "LineStyle")
            self._make_tag(line_style, "color", text="ff" + line_color)
            self._make_tag(line_style, "width", text="1")
        if self._has_circle():
            placemark = self._make_tag(document, "Placemark")
            self._make_tag(placemark, "name", text="Range Boundary")
            self._make_tag(placemark, "styleUrl", text="#CircleStyle")
            ls = self._make_tag(placemark, "LineString")
            coordinates = \
                " ".join(f"{lon},{lat}" for lat, lon in
                         GeoCalc.circle(center_lat=mf(self._center_lat),
                                        center_lon=mf(self._center_lon),
                                        radius_miles=mf(self._range_miles),
                                        num_vertex=60,
                                        close=True))
            self._make_tag(ls, "coordinates", text=coordinates)
        for (lat, lon), map_point in self._map_points.items():
            placemark = self._make_tag(document, "Placemark")
            self._make_tag(placemark, "name", text=map_point.name())
            self._make_tag(placemark, "description",
                           text=map_point.description())
            point = self._make_tag(placemark, "Point")
            self._make_tag(point, "extrude", text="1")
            self._make_tag(point, "altitude_mode", text="relativeToGround")
            self._make_tag(point, "coordinates",
                           text=f"{lon},{lat},{map_point.raat()}")
        for (src_lat, src_lon, dst_lat, dst_lon), map_seg \
                in self._map_segs.items():
            placemark = self._make_tag(document, "Placemark")
            self._make_tag(
                placemark, "name", text=map_seg.name(self.MapSeg.Feature.Body))
            self._make_tag(placemark, "description",
                           text=map_seg.description(self.MapSeg.Feature.Body))
            self._make_tag(placemark, "styleUrl",
                           text="#SegmentStyle_" + map_seg.color)
            line = self._make_tag(placemark, "LineString")
            self._make_tag(line, "extrude", text="1")
            self._make_tag(line, "altitude_mode", text="relativeToGround")
            self._make_tag(
                line, "coordinates",
                text=(f"{src_lon},{src_lat},{map_seg.src_raat} "
                      f"{dst_lon},{dst_lat},{map_seg.dst_raat}"))
            for begin_lat, begin_lon, end_lat, end_lon, arrow, feature in \
                    [(src_lat, src_lon, dst_lat, dst_lon, map_seg.dst_arrow,
                      self.MapSeg.Feature.Dst),
                     (dst_lat, dst_lon, src_lat, src_lon, map_seg.src_arrow,
                      self.MapSeg.Feature.Src)]:
                if not arrow:
                    continue
                lat1, lon1, lat2, lon2 = \
                    GeoCalc.arrow_head(
                        begin_lat=begin_lat, begin_lon=begin_lon,
                        end_lat=end_lat, end_lon=end_lon, angle=1,
                        size_miles=0.5)
                placemark = self._make_tag(document, "Placemark")
                self._make_tag(placemark, "name", text=map_seg.name(feature))
                self._make_tag(placemark, "description",
                               text=map_seg.description(feature))
                self._make_tag(placemark, "styleUrl", text="#ArrowStyle")
                line = self._make_tag(placemark, "LineString")
                self._make_tag(
                    line, "coordinates",
                    text=(f"{lon1},{lat1} {end_lon},{end_lat} {lon2},{lat2}"))
        with open(self._filename, "w", encoding="ascii") as f:
            f.write(
                minidom.parseString(
                    etree.tostring(
                        root, encoding="utf-8")).toprettyxml(indent="    "))

    def _add_ant_coord(self, path: PathInfo, ant_coord: PathInfo.AntCoord,
                       role: str, rx_callsign: Optional[str] = None) -> None:
        """ Adds antenna information to _map_points

        Arguments:
        path        -- Path antenna belongs to
        ant_coord   -- Antenna coordinates
        role        -- Antenna role string
        rx_callsign -- RX callsign (if appropriate) or None
        """
        self._map_points.setdefault(
            (mf(ant_coord.lat), mf(ant_coord.lon)),
            self.MapPoint()).add(
                self.AntInfo(path=path, role=role,
                             raat=mf(ant_coord.height_to_center_raat_m),
                             rx_callsign=rx_callsign))

    def _has_circle(self) -> bool:
        """ True if circle drawing was requested """
        return (self._center_lat is not None) and \
            (self._center_lon is not None) and (self._range_miles is not None)

    def _add_segment(self, path: PathInfo, src: PathInfo.AntCoord,
                     dst: PathInfo.AntCoord, color: str,
                     color_priority: Optional[List[str]]) -> None:
        """ Adds map segment

        Arguments
        path           -- PathInfo segment belongs to
        src            -- AntCoord of one end
        dst            -- AntCoord of other end
        color          -- 6-character hexadecimal KML color string ("bbggrr")
        color_priority -- None or list of color strings that specify
                          resulting color (color with higher index beats
                          one with lower). None means most recent wins
        """
        src_arrow = False
        dst_arrow = self._arrows
        if (mf(src.lat), mf(src.lon)) > (mf(dst.lat), mf(dst.lon)):
            src, dst, src_arrow, dst_arrow = dst, src, dst_arrow, src_arrow
        self._map_segs.setdefault(
            (mf(src.lat), mf(src.lon), mf(dst.lat), mf(dst.lon)),
            self.MapSeg()).add(path=path, src=src, dst=dst,
                               src_arrow=src_arrow, dst_arrow=dst_arrow,
                               color=color, color_priority=color_priority)

    def _make_tag(self, parent: etree.Element, name: str,
                  text: Optional[str] = None,
                  attrs: Optional[Dict[str, str]] = None) -> etree.Element:
        """ Creates subtag

        Arguments:
        parent -- Parent tag
        name   -- Tag name
        text   -- Tag text or None
        attrs  -- Optional dictionary with attributes
        """
        ret = etree.SubElement(parent, name, attrib=attrs or {})
        if text is not None:
            ret.text = text
        return ret


def check_download_switches(args: Any):
    """ Checks consistency off download-related switches """
    error_if(
        (int(args.from_dir is not None) + int(args.to_dir is not None) +
         int(args.zip is not None)) > 1,
        "--from_dir, --to_dir and --zip are mutually exclusive. No more " +
        "than one of them can be specified")


def do_latest(args: Any) -> None:
    """ Processes 'latest' subcommand

    Arguments:
    args -- Argparsed command line arguments """
    check_download_switches(args)
    error_if(args.zip and args.inc,
             "--zip and --inc switches can't be specified simultaneously")
    with Downloader(from_dir=args.from_dir, to_dir=args.to_dir,
                    retry=args.retry, zip_file=args.zip) as dl:
        dl.download()
        date, tz = dl.inc_date if args.inc else dl.date
        assert date is not None
        print(f"{date.strftime('%Y-%m-%d %H:%M:%S')} {tz}")


def do_download_uls(args: Any) -> None:
    """ Processes 'download_uls' subcommand

    Arguments:
    args -- Argparsed command line arguments """
    check_download_switches(args)
    db_filename: str = args.DST_DB
    if os.path.isfile(db_filename):
        error_if(not args.overwrite, f"File {db_filename} already exists")
        try:
            os.unlink(db_filename)
        except OSError:
            error(
                (f"Can't delete '{db_filename}'. Maybe it is open in another "
                 f"program"))
    with Timing(epilogue="Entire download"), \
            Complainer(filename=args.report), \
            Downloader(from_dir=args.from_dir, to_dir=args.to_dir,
                       retry=args.retry, zip_file=args.zip) as dl:
        day_dirs = dl.download()
        if args.skip_write:
            args.data = "all"
        with UlsDatabaseFiller(
                db_filename=args.DST_DB, day_dirs=day_dirs,
                skip_write=args.skip_write, force_pk=args.strict,
                relaxed=(not args.strict) and (args.data == "all")) as uf:
            uf.load_table(name="PA")
            if args.data != "all":
                paths, tx_antennas = uf.get_paths_from_pa()

            def em_row_checker(line_num: int, row_data: ROW_DATA_TYPE) -> bool:
                """ Checks validity and appropriateness of EM table row """
                if args.data == "all":
                    return True
                if row_data["emission_code"] and \
                        (FreqRange.emission_bw_mhz(
                            mst(row_data['emission_code'])) is None):
                    ComplaintEmissionCode(
                        callsign=mst(row_data["call_sign"]),
                        emission_code=mst(row_data['emission_code']),
                        line_num=line_num)
                    return False
                if args.data != "6g":
                    return True
                return ((row_data["call_sign"], row_data["location_number"],
                         row_data["antenna_number"]) in tx_antennas) and \
                    FreqRange(
                        center_mhz=of(row_data["frequency_assigned"]),
                        emission_code=ost(row_data["emission_code"])).is_6ghz()

            uf.load_table(name="EM", row_filter=em_row_checker)
            if args.data != "all":
                paths, tx_antennas = uf.get_paths_from_pa_em()
                if args.data == "6g":
                    with Timing("Purging non-6GHz rows from EM"):
                        uf.purge_paths_by_em()
            uf.load_table(
                name="SG",
                row_filter=lambda ln, r: True if args.data == "all" else
                (r["call_sign"], r["path_number"]) in paths)
            if args.data != "all":
                tx_callsigns, rx_callsigns, locations, antennas = \
                    uf.get_paths_from_pa_sg()
            uf.load_table(
                name="LO",
                row_filter=lambda ln, r: True if args.data == "all" else
                (((r["call_sign"], r["location_number"]) in locations) or
                 (r["call_sign"] in rx_callsigns)))
            uf.load_table(
                name="AN",
                row_filter=lambda ln, r: True if args.data == "all" else
                (((r["call_sign"], r["location_number"], r["antenna_number"])
                  in antennas) or
                 (r["call_sign"] in rx_callsigns)))
            uf.load_table(
                name="CP",
                row_filter=lambda ln, r: True if args.data == "all" else
                ((r["call_sign"] in tx_callsigns) or
                 (r["call_sign"] in rx_callsigns)))
            uf.load_table(
                name="EN",
                row_filter=lambda ln, r: True if args.data == "all" else
                ((r["call_sign"] in tx_callsigns) or
                 (r["call_sign"] in rx_callsigns)))
            uf.load_table(
                name="FR",
                row_filter=lambda ln, r: True if args.data == "all" else
                ((r["call_sign"], r["location_number"], r["antenna_number"])
                 in tx_antennas))
            uf.load_table(
                name="HD",
                row_filter=lambda ln, r: True if args.data == "all" else
                ((r["call_sign"] in tx_callsigns) or
                 (r["call_sign"] in rx_callsigns)))


def do_uls_to_afc(args: Any) -> None:
    """ Processes 'uls_to_afc' subcommand

    Arguments:
    args -- Argparsed command line arguments """
    error_if(
        not os.path.isfile(args.SRC_DB),
        f"Can't find source database file '{args.SRC_DB}'")
    error_if(args.prev_afc and (not os.path.isfile(args.prev_afc)),
             f"Can't find previous AFC database '{args.prev_afc}'")
    if os.path.isfile(args.DST_DB):
        try:
            os.unlink(args.DST_DB)
        except OSError:
            error((f"Can't delete {args.DST_DB}. May be it opened in "
                   f"another program"))
    fsid_gen = FsidGenerator(args.prev_afc)
    with Timing("Generating AFC database"), \
            DbPathIterator(args.SRC_DB) as pi, \
            Complainer(filename=args.report), \
            AfcWriter(args.DST_DB, validator=PathValidator(),
                      fsid_gen=fsid_gen,
                      compatibility=args.compatibility) as aw:
        for path in pi:
            aw.write_path(path)


def do_compare(args: Any) -> None:
    """ Processes 'compare' subcommand

    Arguments:
    args -- Argparsed command line arguments """
    spatial_resolution: float = 0
    spatial_top = 0
    if args.spatial:
        m = re.match(r"^([0-9.]+)(;(\d+))?", args.spatial)
        error_if(not m, "Invalid format of --spatial' switch argument")
        assert m is not None
        spatial_resolution = float(m.group(1))
        if m.group(3):
            spatial_top = int(m.group(3))

    for f in (args.DB1, args.DB2):
        error_if(not os.path.isfile(f), f"Source database {f} not found")
    for f in (args.added, args.removed, args.common):
        error_if(f in (args.DB1, args.DB2),
                 f"'{f}' specified as both input and output file")
        if (f is not None) and os.path.isfile(f):
            try:
                os.unlink(f)
            except OSError:
                error((f"Can't delete '{f}'. May be it opened in another "
                       f"program"))

    path_name: Callable[[str, Optional[int]], str] = \
        lambda cs, pn: cs + (("/" + str(pn)) if pn is not None else "")
    # Dictionaries of path names for first and second database, computed with
    # and without path numbers
    path_hashes: Dict[bool, List[Dict[int, str]]] = {}
    # Numbers of paths in first and second database
    path_counts: List[int] = []
    # True if both database has path numbers
    has_path_num = True
    # Paths in first and second database
    paths: List[Set[Tuple[str, Optional[int]]]] = []
    for f in (args.DB1, args.DB2):
        for wpn in (True, False):
            path_hashes.setdefault(wpn, []).append({})
        path_counts.append(0)
        paths.append(set())
        with Timing(f"Scanning {f}"), DbPathIterator(f) as pi:
            for path in pi:
                if path.is_afc() and path.tx_ant and path.rx_ant:
                    path_counts[-1] += 1
                    paths[-1].add((path.callsign, path.path_num))
                    has_path_num &= path.path_num is not None
                    for wpn in (True, False):
                        ph = path.path_hash(with_path_num=wpn)
                        if (ph in path_hashes[wpn][-1]) and \
                                ((path.path_num is not None) == wpn):
                            logging.warning(
                                (f"Script bug: '{f}' has hash collision "
                                 f"between {path_hashes[wpn][-1][ph]} and "
                                 f"{path.path_name()}"))
                        path_hashes[wpn][-1][ph] = path.path_name()
    added_hashes = \
        set(path_hashes[has_path_num][1].keys()) - \
        set(path_hashes[has_path_num][0].keys())
    removed_hashes = \
        set(path_hashes[has_path_num][0].keys()) - \
        set(path_hashes[has_path_num][1].keys())
    spatial_bins: Dict[Tuple[float, float], int] = {}
    with KmlWriter(
            filename=args.kml,
            comment=(f"Paths unique for "
                     f"{os.path.splitext(os.path.basename(args.DB1))[0]} "
                     f"(red) and for "
                     f"{os.path.splitext(os.path.basename(args.DB2))[0]} "
                     f"(green)"), arrows=False) as kw:
        for hashes, src_db, dst_db, color in \
                [(removed_hashes, args.DB1, args.removed, KmlWriter.C_RED),
                 (added_hashes, args.DB2, args.added, KmlWriter.C_GREEN)]:
            if (not dst_db) and (not args.spatial) and (not args.kml):
                continue
            with Timing(f"Scanning {src_db}"), \
                    DbPathIterator(src_db) as pi, \
                    AfcWriter(dst_db, validator=PathValidator(pass_all=True),
                              fsid_gen=FsidGenerator(""),
                              compatibility=True) as aw:
                for path in pi:
                    if path.path_hash(with_path_num=has_path_num) in hashes:
                        if dst_db:
                            aw.write_path(path)
                        if args.spatial:
                            key = (round(path.tx_ant.lat /
                                         spatial_resolution) *
                                   spatial_resolution,
                                   round(path.tx_ant.lon /
                                         spatial_resolution) *
                                   spatial_resolution)
                            spatial_bins.setdefault(key, 0)
                            spatial_bins[key] += 1
                        if args.kml:
                            kw.add_path(path, color=color)
        if args.kml:
            if kw.num_features() > KmlWriter.KML_MAX_FEATURES:
                logging.warning(
                    (f"Number of KML features is {kw.num_features()} which "
                     f"exceeds allowed maximum "
                     f"({KmlWriter.KML_MAX_FEATURES})"))
            else:
                print(f"Number of features in KML file: {kw.num_features()}")
    if args.common:
        with Timing(f"Generating {args.common} from {args.DB1}"), \
                DbPathIterator(args.DB1) as pi, \
                AfcWriter(args.common, validator=PathValidator(pass_all=True),
                          fsid_gen=FsidGenerator(""),
                          compatibility=True) as aw:
            for path in pi:
                if path.path_hash(with_path_num=has_path_num) not in \
                        removed_hashes:
                    aw.write_path(path)
    if args.unique:
        d = sorted(paths[0] - paths[1])
        print(f"{len(d)} unique paths in first database: "
              f"{', '.join(path_name(p[0], p[1]) for p in d)}")
        d = sorted(paths[1] - paths[0])
        print(f"{len(d)} unique paths in second database: "
              f"{', '.join(path_name(p[0], p[1]) for p in d)}")
        # For each database - maps path names to sets of path hashes
        hp: List[Dict[str, Set[int]]] = []
        for ph in path_hashes[has_path_num]:
            hp.append({})
            for path_hash, path_name in ph.items():
                hp[-1].setdefault(path_name, set()).add(path_hash)
        miscalculated_paths: List[str] = \
            [pn for pn, phs in hp[0].items() if phs != hp[1].get(pn, phs)]
        print((f"{len(miscalculated_paths)} paths, computed differently: "
               f"{', '.join(sorted(miscalculated_paths))}"))
    print(f"Paths added"
          f"{' ('+os.path.basename(args.added)+')' if args.added else ''}"
          f": {len(added_hashes)}")
    print(f"Paths removed"
          f"{' ('+os.path.basename(args.removed)+')' if args.removed else ''}"
          f": {len(removed_hashes)}")
    print(f"Paths common"
          f"{' ('+os.path.basename(args.common)+')' if args.common else ''}"
          f": {path_counts[0] - len(removed_hashes)}")
    if args.spatial:
        num_printed = 0
        for (lat, lon), num in \
                sorted(spatial_bins.items(), key=lambda kvp: kvp[1],
                       reverse=True):
            print(f"Changes around {abs(lat)}{'N' if lat >= 0 else 'S'}, "
                  f"{abs(lon)}{'E' if lon >= 0 else 'W'}: {num}")
            num_printed += 1
            if spatial_top and (num_printed >= spatial_top):
                break


def do_gmap(args: Any) -> None:
    """ Processes 'gmap' subcommand

    Arguments:
    args -- Argparsed command line arguments """
    error_if(
        not os.path.isfile(args.SRC_DB),
        f"Can't find source database file '{args.SRC_DB}'")
    error_if(
        ((args.lat is None) != (args.distance is None)) or
        ((args.lon is None) != (args.distance is None)),
        "--lat, --lon, --distance should be specified together or not at all")
    error_if((not args.path) and (args.lat is None),
             "Neither paths nor vicinity specified")
    requested_path_names: Set[str] = set(args.path)
    found_path_names: Set[str] = set()
    paths: List[PathInfo] = []
    with DbPathIterator(args.SRC_DB) as pi, Timing("Making path list"):
        error_if(any(("/" in p) for p in requested_path_names) and
                 (not pi.has_path_num()),
                 "Source database does not have path numbers, so no path " +
                 "numbers may be specified in --path")
        for path in pi:
            if not path and path.is_afc():
                continue
            if (path.callsign in requested_path_names) or \
                    (path.path_name() in requested_path_names) or \
                    ((args.lat is not None) and
                     path.within_range(center_lat=args.lat,
                                       center_lon=args.lon,
                                       range_miles=args.distance)):
                paths.append(path)
                found_path_names.add(path.callsign)
                found_path_names.add(path.path_name())
    paths_not_found = requested_path_names - found_path_names
    error_if(paths_not_found,
             (f"Following paths were not found: "
              f"{', '.join(sorted(paths_not_found))}"))
    with KmlWriter(filename=args.KML_FILE,
                   comment=(f"Paths in {args.distance} miles around "
                            f"{abs(args.lat)}"
                            f"{'N' if args.lat >= 0 else 'S'}, "
                            f"{abs(args.lon)}"
                            f"{'E' if args.lon >= 0 else 'W'}"),
                   center_lat=args.lat, center_lon=args.lon,
                   range_miles=args.distance,
                   arrows=args.direction) as kw:

        for idx, path in enumerate(paths):
            if (kw.num_features() + kw.num_path_features(path)) > \
                    KmlWriter.KML_MAX_FEATURES:
                logging.warning(
                    (f"Number of KML features exceeds "
                     f"{KmlWriter.KML_MAX_FEATURES}. {idx} of {len(paths)} "
                     f"paths written"))
                break
            kw.add_path(path, color=KmlWriter.C_BLACK)


def do_download_afc(args: Any) -> None:
    """ Processes 'download_afc' subcommand

    Arguments:
    args -- Argparsed command line arguments """
    check_download_switches(args)
    error_if(args.prev_afc and (not os.path.isfile(args.prev_afc)),
             f"Can't find previous AFC database '{args.prev_afc}'")
    db_filename: str = args.DST_DB
    if os.path.isfile(db_filename):
        try:
            os.unlink(db_filename)
        except OSError:
            error(
                (f"Can't delete '{db_filename}'. Maybe it is open in another "
                 f"program"))
    fsid_gen = FsidGenerator(args.prev_afc)
    with Timing(epilogue="Entire download"), \
            Complainer(filename=args.report), \
            Downloader(from_dir=args.from_dir, to_dir=args.to_dir,
                       retry=args.retry, zip_file=args.zip) as dl:
        dpi = DownloadedPathIterator(day_dirs=dl.download())
        with Timing("Writing AFC database"), \
                AfcWriter(args.DST_DB, validator=PathValidator(),
                          fsid_gen=fsid_gen,
                          compatibility=args.compatibility) as aw:
            for path in dpi:
                aw.write_path(path)


def do_csv(args: Any) -> None:
    """ Processes 'csv' subcommand

    Arguments:
    args -- Argparsed command line arguments """
    check_download_switches(args)
    callsigns: Set[str] = set(args.callsign)
    uls_table = UlsTable.try_by_name(args.table)
    error_if(uls_table is None,
             "Incorrect/unsupported ULS table name specified")
    assert uls_table is not None

    def process_row(csv_writer: Any, callsign: str, record: str):
        if (not callsigns) or (callsign in callsigns):
            csv_writer.writerow(record.split("|"))

    with Downloader(from_dir=args.from_dir, to_dir=args.to_dir,
                    retry=args.retry, zip_file=args.zip) as dl, \
            open(args.CSV_FILE, mode="w", encoding="ascii", newline="",
                 errors="backslashreplace") as f:
        csv_writer = csv.writer(f, dialect="excel")
        csv_writer.writerow([field.dsc for field in uls_table.fields])
        day_dirs = dl.download()
        if args.day:
            directory = os.path.join(dl.unpacked_directory, args.day)
            error_if(not os.path.isdir(directory),
                     f"Directory {directory} not found. Wrong --day parameter")
            with UlsFileReader(
                    filename=os.path.join(
                        directory,
                        UlsTableReader.TABLE_FILE_PATTERN.format(
                            table=args.table)),
                    uls_table=uls_table) as ufr:
                for row_data, _, record in ufr:
                    process_row(csv_writer=csv_writer,
                                callsign=row_data["calls_ign"], record=record)
        else:
            for _, _, row_data, record in \
                    UlsTableReader(uls_table=uls_table, day_dirs=day_dirs):
                process_row(csv_writer=csv_writer,
                            callsign=row_data["call_sign"], record=record)


def do_xlsx(args: Any) -> None:
    """ Processes 'xlsx' subcommand

    Arguments:
    args -- Argparsed command line arguments """
    check_download_switches(args)
    error_if("openpyxl" not in sys.modules,
             "Python 'openpyxl' module not installed. Please install it " +
             "with 'pip install openpyxl' command")

    callsigns: Set[str] = set(args.callsign)
    filename = args.XLSX_FILE or ("_".join(args.callsign) + ".xlsx")
    if os.path.isfile(filename):
        try:
            os.unlink(filename)
        except OSError:
            error(
                (f"Can't delete '{filename}'. Maybe it is open in "
                 f"another program"))
    with Downloader(from_dir=args.from_dir, to_dir=args.to_dir,
                    retry=args.retry, zip_file=args.zip) as dl, \
            Timing(epilogue="Entire download"):
        day_dirs = dl.download()
        wb = openpyxl.Workbook()
        for tab_idx, uls_table in enumerate(UlsTable.all()):
            with Timing(f"Scanning {uls_table.name}"):
                if tab_idx == 0:
                    ws = wb.active
                    ws.title = uls_table.name
                else:
                    ws = wb.create_sheet(uls_table.name)
                for field_idx, field in enumerate(uls_table.fields):
                    ws.cell(row=1, column=field_idx + 1, value=field.dsc)
                    min_width = 0
                    if isinstance(field.col_type, sa.Unicode):
                        min_width = field.col_type.length
                    elif isinstance(field.col_type, sa.DateTime):
                        min_width = 10
                    cn = spreadsheet_col_name(field_idx + 1)
                    ws.column_dimensions[cn].width = \
                        max(ws.column_dimensions[cn].width, min_width)
                row_count = 1
                for _, _, row_data, record in \
                        UlsTableReader(uls_table=uls_table, day_dirs=day_dirs):
                    if row_data["call_sign"] not in callsigns:
                        continue
                    for value_idx, value in enumerate(record.split("|")):
                        field = uls_table.fields[value_idx]
                        v: Optional[Union[int, float, str]] = value
                        if not value:
                            pass
                        elif isinstance(field.col_type, sa.Integer):
                            v = int(value)
                        elif isinstance(field.col_type, sa.Float):
                            v = float(value)
                        ws.cell(row=row_count + 1, column=value_idx + 1,
                                value=v)
                    row_count += 1
                if row_count == 1:
                    logging.warning(
                        (f"No data on callsign(s) {', '.join(args.callsign)} "
                         f"found in table {uls_table.name}"))
        wb.save(filename)


def do_help(args: Any) -> None:
    """Execute "help" command.

    Arguments:
    args -- Parsed command line arguments (also contains 'argument_parser' and
            'subparsers' fields)
    """
    if args.subcommand is None:
        args.argument_parser.print_help()
    else:
        args.subparsers.choices[args.subcommand].print_help()


def main(argv: List[str]) -> None:
    """Do the job.

    Arguments:
    argv -- Program arguments
    """
    # Container of common switches
    switches_common = \
        argparse.ArgumentParser(add_help=False)
    switches_common.add_argument(
        "--progress", action="store_true",
        help="Print download progress information")
    switches_common.add_argument(
        "--quiet", action="store_true",
        help="Do not print messages on nonfatal data inconsistencies")

    # Container of report-specific switches
    switches_report = \
        argparse.ArgumentParser(add_help=False)
    switches_report.add_argument(
        "--report", metavar="FILE_NAME",
        help="Write ULS data inconsistency report to given file")

    # Container of download-specific switches
    switches_download = \
        argparse.ArgumentParser(add_help=False)
    switches_download.add_argument(
        "--to_dir", metavar="DOWNLOAD_DIRECTORY",
        help="Download raw FCC ULS files to given directory and retain " +
        "them there. Default is to download to temporary directory and " +
        "remove it afterwards")
    switches_download.add_argument(
        "--from_dir", metavar="DOWNLOAD_DIRECTORY",
        help="Use raw FCC ULS files from given firectory (created from " +
        "--to_dir in or some other way). Default is to download files from " +
        "FCC server")
    switches_download.add_argument(
        "--retry", action="store_true",
        help="Retry download several times if it aborts. For use e.g. on " +
        "terminally overheated machines")
    switches_download.add_argument(
        "--zip", metavar="ZIP_FILE_NAME",
        help="Take ULS tables from given single (presumably - full) .zip " +
        "archive instead of downloading from FCC ULS site. For development " +
        "purposes")

    # Container of AFC database generation switches
    switches_afc = \
        argparse.ArgumentParser(add_help=False)
    switches_afc.add_argument(
        "--compatibility", action="store_true",
        help="Generate AFC database compatible with all AFC engine " +
        "releases. Default is to generate database for latest AFC engine " +
        "release")
    switches_afc.add_argument(
        "--prev_afc", metavar="PREV_AFC_DATABASE",
        help="Open AFC database 'fsid' field computed as sequential number " +
        "counted from previous AFC database - specified by this switch. " +
        "Default is to assign sequential numbers, starting from 1")

    # Top level parser
    argument_parser = argparse.ArgumentParser(
        description=f"ULS download and manipulation tool. V{VERSION}")
    subparsers = argument_parser.add_subparsers(dest="subcommand",
                                                metavar="SUBCOMMAND")
    # Subparser for 'latest' command
    parser_latest = subparsers.add_parser(
        "latest", parents=[switches_common, switches_download],
        help="Prints date of the latest ULS database on FCC server")
    parser_latest.add_argument(
        "--inc", action="store_true",
        help="Print date of latest incremental data (ignoring date of full " +
        "data)")
    parser_latest.set_defaults(func=do_latest)

    # Subparser for 'download_uls' command
    parser_download = subparsers.add_parser(
        "download_uls", parents=[switches_common, switches_download,
                                 switches_report],
        help="Download ULS data and make ULS database from them")
    parser_download.add_argument(
        "--data", choices=["all", "used", "6g"], default="6g",
        help=(f"What part of data to put to database: 'all' - data from all "
              f"rows of processed tables (may take several hours), 'used' - "
              f"data belonging to defined paths on all bands, '6g' - data "
              f"belonging to defined paths on 6GHz band "
              f"([{FreqRange.LOW_6G_START_MHZ}MHz to "
              f"{FreqRange.LOW_6G_END_MHZ}MHz] and "
              f"[{FreqRange.HIGH_6G_START_MHZ}MHz to "
              f"{FreqRange.HIGH_6G_END_MHZ}]MHz). '6g' is the default"))
    parser_download.add_argument(
        "--overwrite", action="store_true",
        help="Overwrite database if it is already exists")
    parser_download.add_argument(
        "--skip_write", action="store_true",
        help="Create target database, but don't write into it. For " +
        "development purposes")
    parser_download.add_argument(
        "--strict", action="store_true",
        help="Upheld primary keys in all tables (e.g. to have richer error " +
        "report). By default it is being upheld not in all tables")
    parser_download.add_argument(
        "DST_DB",
        help="Name of SQLite file to create")
    parser_download.set_defaults(func=do_download_uls)

    # Subparser for 'uls_to_afc' command
    parser_uls_to_afc = subparsers.add_parser(
        "uls_to_afc", parents=[switches_afc, switches_common, switches_report],
        help="Make AFC database from ULS database")
    parser_uls_to_afc.add_argument(
        "SRC_DB",
        help="Name of SQLite file with previously downloaded ULS database")
    parser_uls_to_afc.add_argument(
        "DST_DB",
        help="Name of file for AFC SQLite database. Overrides if file " +
        "already exists")
    parser_uls_to_afc.set_defaults(func=do_uls_to_afc)

    # Subparser for 'download_afc' command
    parser_download_afc = subparsers.add_parser(
        "download_afc",
        parents=[switches_afc, switches_common, switches_download,
                 switches_report],
        help="Download ULS data and make AFC database from them")
    parser_download_afc.add_argument(
        "DST_DB",
        help="Name of file for AFC SQLite database. Overrides if file " +
        "already exists")
    parser_download_afc.set_defaults(func=do_download_afc)

    # Subparser for 'gmap' command
    parser_gmap = subparsers.add_parser(
        "gmap", parents=[switches_common],
        help="Export KML to import to Google maps")
    parser_gmap.add_argument(
        "--lat", metavar="CENTER_LATITUDE_DEGREES", type=float,
        help="Latitude of center point in degrees")
    parser_gmap.add_argument(
        "--lon", metavar="CENTER_LONGITUDE_DEGREES", type=float,
        help="Latitude of center point in degrees")
    parser_gmap.add_argument(
        "--distance", metavar="DISTANCE_IN_MILES", type=float,
        help="Maximum distance in miles from center point")
    parser_gmap.add_argument(
        "--path", metavar="CALLSIGN[/NUM]", action="append", default=[],
        help="Path to put on map. This switch may be specified several times")
    parser_gmap.add_argument(
        "--direction", action="store_true",
        help="Show direction arrows on paths. Note that this adds clutter " +
        "and reduces the number of paths that can be drawn (arrow is a " +
        "separate feature, KML has a limit on number of features)")
    parser_gmap.add_argument(
        "SRC_DB",
        help="ULS or AFC database file")
    parser_gmap.add_argument(
        "KML_FILE",
        help="Name of KML file to generate")
    parser_gmap.set_defaults(func=do_gmap)

    # Subparser for 'compare' command
    parser_compare = subparsers.add_parser(
        "compare",
        parents=[switches_common],
        help="Compare paths in two ULS or AFC databases")
    parser_compare.add_argument(
        "--added", metavar="DB_FOR_ADDED",
        help="AFC-style database with added and changed paths (paths " +
        "present in second, but not in first database)")
    parser_compare.add_argument(
        "--removed", metavar="DB_FOR_REMOVED",
        help="AFC-style database with removed and changed paths (paths " +
        "present in first, but not in second database)")
    parser_compare.add_argument(
        "--common", metavar="DB_FOR_COMMON",
        help="Generate AFC-style database with common paths (paths present " +
        "in both databases). Computed from paths in DB1")
    parser_compare.add_argument(
        "--unique", action="store_true",
        help="Print lists of unique paths from first and second databases")
    parser_compare.add_argument(
        "--spatial", metavar="resolution[;top]",
        help="Print spatial statistics. 'resolution' is resolution in " +
        "latitude/longitude degrees, 'top' is the number of top spots to " +
        "print (default - print all)")
    parser_compare.add_argument(
        "--kml", metavar="KML_FILE",
        help="Create KML file (for displaying in Google maps) with added " +
        "and removed paths")
    parser_compare.add_argument(
        "DB1",
        help="First ULS or AFC database file to compare")
    parser_compare.add_argument(
        "DB2",
        help="Second ULS or AFC database file to compare")
    parser_compare.set_defaults(func=do_compare)

    # Subparser for 'csv' command
    parser_csv = subparsers.add_parser(
        "csv", parents=[switches_download, switches_common],
        help="Download and convert single or consolidated (with daily " +
        "updates) .dat file to Excel-style .csv (fully or for single " +
        "callsign)")
    parser_csv.add_argument(
        "--callsign", metavar="CALLSIGN", action="append", default=[],
        help="In output file leave only rows for given callsign. Default - " +
        "keep all callsigns, this switch may be specified several times (to " +
        "have several callsigns in output file")
    parser_csv.add_argument(
        "--day", metavar="DOW_OR_FULL",
        choices=[Downloader.DOW_FULL] + Downloader.DOWS,
        help="Select single .dat file: 'full' for full database, weekday " +
        "('mon', 'tue'...) for single day update. Default is to use " +
        "consolidated data (full database with all updates)")
    parser_csv.add_argument(
        "--table", metavar="TABLE_NAME", required=True,
        choices=sorted([t.name for t in UlsTable.all()]),
        help="Table name ('AN', 'EM'...)")
    parser_csv.add_argument(
        "CSV_FILE",
        help="Name of CSV file to generate")
    parser_csv.set_defaults(func=do_csv)

    # Subparser for 'xlsx' command
    parser_csv = subparsers.add_parser(
        "xlsx", parents=[switches_download, switches_common],
        help="Download and put to spreadsheet data on given callsign(s)")
    parser_csv.add_argument(
        "--callsign", metavar="CALLSIGN", action="append", default=[],
        required=True,
        help="Callsign to put to spreadsheet. This parameter is mandatory " +
        "and may be specified several times")
    parser_csv.add_argument(
        "XLSX_FILE", nargs="?",
        help="Name of XLSX file to generate. If omitted then callsign (or " +
        "concatenation thereof) with .xlsx extension is used")
    parser_csv.set_defaults(func=do_xlsx)

    # Subparser for 'help' command
    parser_help = subparsers.add_parser(
        "help", add_help=False, usage="%(prog)s subcommand",
        help="Prints help on given subcommand")
    parser_help.add_argument(
        "subcommand", metavar="SUBCOMMAND", nargs="?",
        choices=subparsers.choices,
        help="Name of subcommand to print help about (use " +
        "\"%(prog)s --help\" to get list of all subcommands)")
    parser_help.set_defaults(func=do_help, subparsers=subparsers,
                             argument_parser=argument_parser)
    if not argv:
        argument_parser.print_help()
        sys.exit(1)
    args = argument_parser.parse_args(argv)

    # Set up logging
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(
        logging.Formatter(
            f"{os.path.basename(__file__)}. %(levelname)s: %(message)s"))
    logging.getLogger().addHandler(console_handler)
    logging.getLogger().setLevel(
        logging.ERROR if getattr(args, "quiet", False) else logging.WARNING)
    Progressor.enabled = getattr(args, "progress", False)

    # Do the needful
    args.func(args)


if __name__ == "__main__":
    main(sys.argv[1:])
