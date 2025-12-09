#!/usr/bin/env python3
#
# Copyright (C) 2021 Broadcom. All rights reserved. The term "Broadcom"
# refers solely to the Broadcom Inc. corporate affiliate that owns
# the software below. This work is licensed under the OpenAFC Project License,
# a copy of which is included with this software program
#

"""
Description

The execution always done in specific order: configuration, database, testing.
In case there are not available valid responses to compare with
need to make acquisition and create new database as following.
    ./afc_tests.py --addr <address> --log debug --cmd run
"""

import argparse
import certifi
import csv
import datetime
import hashlib
import inspect
import io
import json
import logging
import openpyxl as oxl
import os
import re
import requests
import shutil
import sqlite3
import subprocess
import sys
import time

import smtplib
import ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from bs4 import BeautifulSoup
from deepdiff import DeepDiff
from multiprocessing.pool import Pool
from string import Template
from _afc_types import *
from _afc_errors import *
from _version import __version__
from _wfa_types import *

AFC_URL_SUFFIX = '/fbrat/ap-afc/'
AFC_REQ_NAME = 'availableSpectrumInquiry'
AFC_WEBUI_URL_SUFFIX = '/fbrat/ratapi/v1/'
AFC_WEBUI_REQ_NAME = 'availableSpectrumInquirySec'
headers = {'content-type': 'application/json'}

AFC_TEST_DB_FILENAME = 'afc_input.sqlite3'
TBL_REQS_NAME = 'test_vectors'
TBL_RESPS_NAME = 'test_data'
TBL_USERS_NAME = 'user_config'
TBL_AFC_CFG_NAME = 'afc_config'
TBL_AP_CFG_NAME = 'ap_config'

AFC_PROT_NAME = 'https'

# metadata variables
TESTCASE_ID = "testCaseId"
# mandatory keys that need to be read from input text file
MANDATORY_METADATA_KEYS = {TESTCASE_ID}

app_log = logging.getLogger(__name__)


class TestCfg(dict):
    """Keep test configuration"""

    def __init__(self):
        dict.__init__(self)
        self.update({
            'cmd': '',
            'port': 443,
            'url_path': AFC_PROT_NAME + '://',
            'log_level': logging.INFO,
            'db_filename': AFC_TEST_DB_FILENAME,
            'tests': None,
            'is_test_by_index': True,
            'resp': '',
            'stress': 0,
            'precision': None})

    def _send_recv(self, params):
        """Run AFC test and wait for respons"""
        data = params.split("'")
        get_req = ''
        for item in data:
            try:
                json.loads(item)
            except ValueError as e:
                continue
            get_req = item
            break

        new_req_json = json.loads(get_req.encode('utf-8'))
        new_req = json.dumps(new_req_json, sort_keys=True)
        if (self['webui'] is False):
            params_data = {
                'conn_type': self['conn_type'],
                'debug': self['debug'],
                'edebug': cfg['elaborated_debug'],
                'gui': self['gui']
            }
        else:
            # emulating request calll from webui
            params_data = {
                'debug': 'True',
                'edebug': cfg['elaborated_debug'],
                'gui': 'True'
            }
        if (self['cache'] == False):
            params_data['nocache'] = 'True'

        ser_cert = not self['verif']
        cli_certs = ()
        if (self['prot'] == AFC_PROT_NAME and
                self['verif'] == False):
            # add mtls certificates if explicitly provided
            if not isinstance(self['cli_cert'], type(None)):
                cli_certs = ("".join(self['cli_cert']),
                             "".join(self['cli_key']))
            # add tls certificates if explicitly provided
            if not isinstance(self['ca_cert'], type(None)):
                ser_cert = "".join(self['ca_cert'])
        app_log.debug(f"Client {cli_certs}, Server {ser_cert}")

        before_ts = time.monotonic()
        rawresp = requests.post(
            self['url_path'],
            params=params_data,
            data=new_req,
            headers=headers,
            timeout=600,  # 10 min
            verify=self['verif'])
        resp = rawresp.json()

        tId = resp.get('taskId')
        if ((self['conn_type'] == 'async') and
                (not isinstance(tId, type(None)))):
            tState = resp.get('taskState')
            params_data['task_id'] = tId
            while (tState == 'PENDING') or (tState == 'PROGRESS'):
                app_log.debug('_run_test() state %s, tid %s, status %d',
                              tState, tId, rawresp.status_code)
                time.sleep(2)
                rawresp = requests.get(self['url_path'],
                                       params=params_data)
                if rawresp.status_code == 200:
                    resp = rawresp.json()
                    break

        tm_secs = time.monotonic() - before_ts
        app_log.info('Test done at %.1f secs', tm_secs)
        return new_req, resp


class TestResultComparator:
    """ AFC Response comparator

    Private instance attributes:
    _precision -- Precision for results' comparison in dB. 0 means exact match
    """

    def __init__(self, precision):
        """ Constructor

        Arguments:
        precision -- Precision for results' comparison in dB. 0 means exact
                     match
        """
        assert precision >= 0
        self._precision = precision

    def compare_results(self, ref_str, result_str):
        """ Compares reference and actual AFC responses

        Arguments:
        ref_str    -- Reference response JSON in string representation
        result_str -- Actual response json in string representation
        Returns list of difference description strings. Empty list means match
        """
        # List of difference description strings
        diffs = []
        # Reference and actual JSON dictionaries
        jsons = []
        for s, kind in [(ref_str, "reference"), (result_str, "result")]:
            try:
                jsons.append(json.loads(s))
            except json.JSONDecodeError as ex:
                diffs.append(f"Failed to decode {kind} JSON data: {ex}")
                return diffs
        self._recursive_compare(jsons[0], jsons[1], [], diffs)
        return diffs

    def _recursive_compare(self, ref_json, result_json, path, diffs):
        """ Recursive comparator of JSON nodes

        Arguments:
        ref_json    -- Reference response JSON dictionary
        result_json -- Actual response JSON dictionary
        path        -- Path (sequence of indices) to node in question
        diffs       -- List of difference description strings to update
        """
        # Items in questions in JSON dictionaries
        ref_item = self._get_item(ref_json, path)
        result_item = self._get_item(result_json, path)
        # Human readable path representation for difference messages
        path_repr = f"[{']['.join(str(idx) for idx in path)}]"
        if ref_item == result_item:
            return  # Items are equal - nothing to do
        # So, items are different. What's the difference?
        if isinstance(ref_item, dict):
            # One item is dictionary. Other should also should be dictionary...
            if not isinstance(result_item, dict):
                diffs.append(f"Different item types at {path_repr}")
                return
            # ... with same set of keys
            ref_keys = set(ref_item.keys())
            result_keys = set(result_item.keys())
            for unique_key in (ref_keys ^ result_keys):
                if self._compare_channel_lists(ref_json, result_json,
                                               path + [unique_key], diffs):
                    ref_keys -= {unique_key}
                    result_keys -= {unique_key}
            if ref_keys != result_keys:
                msg = f"Different set of keys at {path_repr}"
                for kind, elems in [("reference", ref_keys - result_keys),
                                    ("result", result_keys - ref_keys)]:
                    if elems:
                        msg += \
                            f" Unique {kind} keys: {', '.join(sorted(elems))}."
                diffs.append(msg)
                return
            # Comparing values for individual keys
            for key in sorted(ref_keys):
                self._recursive_compare(ref_json, result_json, path + [key],
                                        diffs)
        elif isinstance(ref_item, list):
            # One item is list. Other should also be list...
            if not isinstance(result_item, list):
                diffs.append(f"Different item types at {path_repr}")
                return
            # If this is a channel list (or part thereof - handle it)
            if self._compare_channel_lists(ref_json, result_json, path, diffs):
                return
            # Proceeding with comparison of other list kinds
            if len(ref_item) != len(result_item):
                diffs.append(
                    (f"Different list lengths at at {path_repr}: "
                     f"{len(ref_item)} elements in reference vs "
                     f"{len(result_item)} elements in result"))
                return
            # Comparing individual elements
            for i in range(len(ref_item)):
                self._recursive_compare(ref_json, result_json, path + [i],
                                        diffs)
        else:
            # Items should be scalars
            for item, kind in [(ref_item, "Reference"),
                               (result_item, "Result")]:
                if not isinstance(item, (int, float, str)):
                    diffs.append((f"{kind} data contains unrecognized item "
                                  f"type at {path_repr}"))
                    return
            diffs.append((f"Difference at {path}: reference content is "
                          f"{ref_item}, result content is {result_item}"))

    def _compare_channel_lists(self, ref_json, result_json, path, diffs):
        """ Trying to compare channel lists

        Arguments:
        ref_json    -- Reference response JSON dictionary
        result_json -- Actual response JSON dictionary
        path        -- Path (sequence of indices) to node in question
        diffs       -- List of difference description strings to update
        Returns true if channel list comparison was done, no further action
        required, False if node is not a channel list, should be compared as
        usual
        """
        if path[-1] == "channelCfi":
            # Comparison will be made at "maxEirp"
            return True
        # Human readable path representation for difference messages
        path_repr = f"[{']['.join(str(idx) for idx in path)}]"
        # EIRP dictionaries, indexed by channel identification (number or
        # frequency range)
        ref_channels = {}
        result_channels = {}
        if path[-1] == "maxEirp":
            # Channel numbers
            for kind, src, chan_dict in \
                    [("reference", ref_json, ref_channels),
                     ("result", result_json, result_channels)]:
                try:
                    numbers = self._get_item(src, path[:-1] + ["channelCfi"],
                                             default_last=[])
                    chan_dict.update(
                        dict(zip([str(n) for n in numbers],
                                 self._get_item(src, path, default_last=[]))))
                except (TypeError, ValueError, KeyError):
                    diffs.append((f"Unrecognized  channel list structure at "
                                  f"{path_repr} in {kind}"))
                    return True
        elif path[-1] == "availableFrequencyInfo":
            # Channel frequencies
            for kind, src, chan_dict in \
                    [("reference", ref_json, ref_channels),
                     ("result", result_json, result_channels)]:
                try:
                    for freq_info in self._get_item(src, path,
                                                    default_last=[]):
                        fr = freq_info["frequencyRange"]
                        low = fr['lowFrequency']
                        high = fr['highFrequency']
                        for freq in range(low, high):
                            chan_dict[f"[{freq} - {freq+1}"] = \
                                float(freq_info.get("maxPSD") or
                                      freq_info.get("maxPsd"))
                except (TypeError, ValueError, KeyError):
                    diffs.append((f"Unrecognized frequency list structure at "
                                  f"{path_repr} in {kind}"))
                    return True
        else:
            return False
        # Now will compare two channel dictionaries
        # First looking for unique channels
        for this_kind, this_channels, other_kind, other_channels in \
                [("reference", ref_channels, "result", result_channels),
                 ("result", result_channels, "reference", ref_channels)]:
            for channel in sorted(set(this_channels.keys()) -
                                  set(other_channels.keys())):
                diffs.append(
                    (f"Channel {channel} present in {path_repr} of "
                     f"{this_kind} with EIRP limit of "
                     f"{this_channels[channel]}dBm, but absent in "
                     f"{other_kind}"))
        # Then looking for different EIRPs on common channels
        for channel in sorted(set(ref_channels.keys()) &
                              set(result_channels.keys())):
            diff = abs(ref_channels[channel] - result_channels[channel])
            if diff <= self._precision:
                continue
            diffs.append(
                (f"Different values in {path_repr} for channel {channel}: "
                 f"reference has EIRP of {ref_channels[channel]}dBm, "
                 f"result has EIRP of {result_channels[channel]}dBm, "
                 f"difference is: {diff:g}dB"))
        return True

    def _get_item(self, j, path, default_last=None):
        """ Retrieves item by sequence of indices

        Arguments:
        j            -- JSON dictionary
        path         -- Sequence of indices
        default_last -- What to return if item at last index is absent. None
                        means throw exception (if nonlast item is absent -
                        exception is also thrown)
        Returns retrieved item
        """
        for path_idx, elem_idx in enumerate(path):
            try:
                j = j[elem_idx]
            except (KeyError, IndexError):
                if (default_last is not None) and \
                        (path_idx == (len(path) - 1)):
                    return default_last
                raise
        return j


def json_lookup(key, json_obj, val):
    """Lookup for key in json and change it value if required"""
    keepit = []

    def lookup(key, json_obj, val, keepit):
        if isinstance(json_obj, dict):
            for k, v in json_obj.items():
                if isinstance(v, (dict, list)):
                    lookup(key, v, val, keepit)
                elif k == key:
                    keepit.append(v)
                    if val:
                        json_obj[k] = val
        elif isinstance(json_obj, list):
            for node in json_obj:
                lookup(key, node, val, keepit)
        return keepit

    found = lookup(key, json_obj, val, keepit)
    return found


def get_md5(fname):
    hash_md5 = hashlib.md5()
    with open(fname, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


def create_email_attachment(filename):
    part = None
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        # Add header as key/value pair to attachment part
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f"attachment; filename= {filename}",)
        return part


def send_email(cfg):
    """Send an email to predefined adress using gmail smtp server"""
    sender = cfg.get('email_from')
    recipient = cfg.get('email_to')

    if sender is None or recipient is None:
        app_log.warning("Sender or recipient email is not defined.")
        return

    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()"
                  f" from: {sender}, to: {recipient}")
    if isinstance(cfg['email_to'], type(None)):
        app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()"
                      f" Not sending email because no receiver specified")
        return
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender, cfg['email_pwd'])

        body = f"Please find attached responses."

        message = MIMEMultipart("alternative")
        message['Subject'] = f"AFC test results"
        message['From'] = sender
        message['To'] = recipient
        if not isinstance(cfg['email_cc'], type(None)):
            message['Cc'] = cfg['email_cc']

        # Turn these into plain/html MIMEText objects
        message.attach(MIMEText(body, "plain"))
        message.attach(create_email_attachment(cfg['outfile'][0]))

        server.sendmail(sender, recipient, message.as_string())


def _resubmit_cookies(ssn):
    """ Part of cookie drop fix

    There are two problems with requests.Session in this script:
    1. External cookie loss. 'session' cookie never get back to rat_server.
       Reason is not yet known.
       To mitigate this problem this function reads all cookies and resubmits
       them with path of /, all domains.
       Maybe more nuanced appropach may do it better, but for purpose of this
       script this is, probably, good enough
    2. Internal cookie loss. For some reason session.post() on sign in causes
       two spurious GET. As a result all-important session cookie with login
       sussess gets lost - it even is not passed to first GET.
       To mitigate this problem 'allow_redirects=False' passed to all session
       get()/post() calls
    """
    saved_cookies = dict(ssn.cookies.get_dict())
    ssn.cookies.clear()
    for name, value in saved_cookies.items():
        ssn.cookies.set(name, value)


def _send_recv(cfg, req_data, ssn=None):
    """Send AFC request and receiver it's response"""
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()")

    new_req_json = json.loads(req_data.encode('utf-8'))
    new_req = json.dumps(new_req_json, sort_keys=True)
    if (cfg['webui'] is False):
        params_data = {
            'conn_type': cfg['conn_type'],
            'debug': cfg['debug'],
            'edebug': cfg['elaborated_debug'],
            'gui': cfg['gui']
        }
        if (cfg['cache'] == False):
            params_data['nocache'] = 'True'
        post_func = requests.post
    else:
        # emulating request call from webui
        params_data = {
            'debug': 'True',
            'gui': 'True'
        }
        headers['Accept-Encoding'] = 'gzip, defalte'
        headers['Referer'] = cfg['base_url'] + 'fbrat/www/index.html'

        csrf_token = ssn.cookies.get('csrf_token', default='')
        if csrf_token:
            headers['X-Csrf-Token'] = csrf_token

        app_log.debug(
            f"({os.getpid()}) {inspect.stack()[0][3]}()\n"
            f"Cookies: {requests.utils.dict_from_cookiejar(ssn.cookies)}")
        post_func = ssn.post

    ser_cert = ()
    cli_certs = None
    if ((cfg['prot'] == AFC_PROT_NAME and
            cfg['verif']) or (cfg['ca_cert'])):

        # add mtls certificates if explicitly provided
        if not isinstance(cfg['cli_cert'], type(None)):
            cli_certs = ("".join(cfg['cli_cert']), "".join(cfg['cli_key']))

        # add tls certificates if explicitly provided
        if not isinstance(cfg['ca_cert'], type(None)):
            ser_cert = "".join(cfg['ca_cert'])
            cfg['verif'] = True
        else:
            os.environ['REQUESTS_CA_BUNDLE'] = certifi.where()
            app_log.debug(f"REQUESTS_CA_BUNDLE "
                          f"{os.environ.get('REQUESTS_CA_BUNDLE')}")
            if "REQUESTS_CA_BUNDLE" in os.environ:
                ser_cert = "".join(os.environ.get('REQUESTS_CA_BUNDLE'))
                cfg['verif'] = True
            else:
                app_log.error(f"Missing CA certificate while forced.")
                return

    app_log.debug(f"Client {cli_certs}, Server {ser_cert}")

    try:
        rawresp = post_func(
            cfg['url_path'],
            params=params_data,
            data=new_req,
            headers=headers,
            timeout=600,  # 10 min
            cert=cli_certs,
            allow_redirects=False,
            verify=ser_cert if cfg['verif'] else False)
        if cfg['webui']:
            _resubmit_cookies(ssn)
        rawresp.raise_for_status()
    except (requests.exceptions.HTTPError,
            requests.exceptions.ConnectionError) as err:
        app_log.error(f"{err}")
        return

    resp = rawresp.json()

    tId = resp.get('taskId')
    if ((cfg['conn_type'] == 'async') and
            (not isinstance(tId, type(None)))):
        tState = resp.get('taskState')
        params_data['task_id'] = tId
        while (tState == 'PENDING') or (tState == 'PROGRESS'):
            app_log.debug('_run_test() state %s, tid %s, status %d',
                          tState, tId, rawresp.status_code)
            time.sleep(2)
            rawresp = requests.get(cfg['url_path'],
                                   params=params_data)
            if rawresp.status_code == 200:
                resp = rawresp.json()
                break

    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()"
                  f" Resp status: {rawresp.status_code}")
    return resp


def _login(cfg, ssn):
    """ Making login, opening session """
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()")

    ser_cert = ()
    cli_certs = None
    app_log.debug(f"=== ser {type(ser_cert)}")
    if ((cfg['prot'] == AFC_PROT_NAME and
            cfg['verif']) or (cfg['ca_cert'])):

        # add mtls certificates if explicitly provided
        if not isinstance(cfg['cli_cert'], type(None)):
            cli_certs = ("".join(cfg['cli_cert']), "".join(cfg['cli_key']))

        # add tls certificates if explicitly provided
        if not isinstance(cfg['ca_cert'], type(None)):
            ser_cert = "".join(cfg['ca_cert'])
            cfg['verif'] = True
        else:
            os.environ['REQUESTS_CA_BUNDLE'] = certifi.where()
            app_log.debug(f"REQUESTS_CA_BUNDLE "
                          f"{os.environ.get('REQUESTS_CA_BUNDLE')}")
            if "REQUESTS_CA_BUNDLE" in os.environ:
                ser_cert = "".join(os.environ.get('REQUESTS_CA_BUNDLE'))
                cfg['verif'] = True
            else:
                app_log.error(f"Missing CA certificate while forced.")
                return

    app_log.debug(f"Client {cli_certs}, Server {ser_cert}")

    # get login
    ssn.headers.update({
        'Accept-Encoding': 'gzip, defalte'
    })
    url_login = cfg['base_url'] + 'fbrat/user/sign-in'
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()\n"
                  f"===> URL {url_login}\n"
                  f"===> Status {ssn.headers}\n"
                  f"===> Cookies: {ssn.cookies}\n")

    try:
        rawresp = ssn.get(url_login,
                          stream=False,
                          cert=cli_certs,
                          allow_redirects=False,
                          verify=ser_cert if cfg['verif'] else False)
        _resubmit_cookies(ssn)
    except (requests.exceptions.HTTPError,
            requests.exceptions.ConnectionError) as err:
        app_log.error(f"{err}")
        return
    csrf_token = ssn.cookies.get('csrf_token', default='')
    if not csrf_token:
        app_log.error("CSRF token not found in login response")
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()\n"
                  f"<--- Status {rawresp.status_code}\n"
                  f"<--- Headers {rawresp.headers}\n"
                  f"<--- Cookies: {ssn.cookies}\n"
                  f"<--- CSRF token: {csrf_token}\n")

    # fetch username and password from test db
    con = sqlite3.connect(cfg['db_filename'])
    cur = con.cursor()
    cur.execute('SELECT * FROM %s\n' % TBL_USERS_NAME)
    found_user = cur.fetchall()
    con.close()
    found_json = json.loads(found_user[0][1])
    app_log.debug(f"Found Users: {found_json['username']}")

    form_data = {
        'next': '/',
        'reg_next': '/',
        'csrf_token': csrf_token,
        'username': found_json['username'],
        'password': found_json['password']
    }
    ssn.headers.update({
        'Content-Type': 'application/x-www-form-urlencoded',
        'Referer': url_login,
        'X-Csrf-Token': csrf_token
    })
    try:
        rawresp = ssn.post(url_login,
                           data=form_data,
                           stream=False,
                           cert=cli_certs,
                           allow_redirects=False,
                           verify=ser_cert if cfg['verif'] else False)
        _resubmit_cookies(ssn)
    except (requests.exceptions.HTTPError,
            requests.exceptions.ConnectionError) as err:
        app_log.error(f"{err}")
        return
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()\n"
                  f"<--- Status {rawresp.status_code}\n"
                  f"<--- Headers {rawresp.headers}\n"
                  f"<--- Cookies: {ssn.cookies}\n")

    return


def make_db(filename):
    """Create DB file only with schema"""
    app_log.debug('%s()', inspect.stack()[0][3])
    if os.path.isfile(filename):
        app_log.debug('%s() The db file is exists, no need to create new one.',
                      inspect.stack()[0][3])
        return True

    app_log.info('Create DB tables (%s, %s) from source files',
                 TBL_REQS_NAME, TBL_RESPS_NAME)
    con = sqlite3.connect(filename)
    cur = con.cursor()
    cur.execute('CREATE TABLE IF NOT EXISTS ' + TBL_REQS_NAME +
                ' (test_id varchar(50), data json)')
    cur.execute('CREATE TABLE IF NOT EXISTS ' + TBL_RESPS_NAME +
                ' (test_id varchar(50), data json, hash varchar(255))')
    con.close()
    return True


def compare_afc_config(cfg):
    """
    Compare AFC configuration from the DB with provided one.
    """
    app_log.debug('%s()', inspect.stack()[0][3])

    if not os.path.isfile(cfg['db_filename']):
        app_log.error('Missing DB file %s', cfg['db_filename'])
        return AFC_ERR

    con = sqlite3.connect(cfg['db_filename'])
    cur = con.cursor()
    cur.execute('SELECT * FROM %s' % TBL_AFC_CFG_NAME)
    found_cfgs = cur.fetchall()
    con.close()

    # get record from the input file
    if isinstance(cfg['infile'], type(None)):
        app_log.debug('Missing input file to compare with.')
        return AFC_OK

    filename = cfg['infile'][0]
    with open(filename, 'r') as fp_test:
        while True:
            rec = fp_test.read()
            if not rec:
                break
            try:
                get_rec = json.loads(rec)
            except (ValueError, TypeError) as e:
                continue
            break
    app_log.debug(json.dumps(get_rec, sort_keys=True, indent=4))

    get_cfg = ''
    app_log.debug('Found %d config records', len(found_cfgs))
    idx = 0
    max_idx = len(found_cfgs)
    if not isinstance(cfg['idx'], type(None)):
        idx = cfg['idx']
        if idx >= max_idx:
            app_log.error("The index (%d) is out of range (0 - %d).",
                          idx, max_idx - 1)
            return AFC_ERR
        max_idx = idx + 1
    while idx < max_idx:
        for item in list(found_cfgs[idx]):
            try:
                get_cfg = json.loads(item)
            except (ValueError, TypeError) as e:
                continue
            break
        app_log.debug("Record %d:\n%s",
                      idx,
                      json.dumps(get_cfg['afcConfig'],
                                 sort_keys=True,
                                 indent=4))

        get_diff = DeepDiff(get_cfg['afcConfig'],
                            get_rec,
                            report_repetition=True)
        app_log.info("rec %d:\n%s", idx, get_diff)
        idx += 1

    return AFC_OK


def start_acquisition(cfg):
    """
    Fetch test vectors from the DB, drop previous response table,
    run tests and fill responses in the DB with hash values
    """
    app_log.debug(f"{inspect.stack()[0][3]}()")

    found_reqs, found_resp, ids, test_ids = _convert_reqs_n_resps_to_dict(cfg)

    # check if to make acquisition of all tests
    con = sqlite3.connect(cfg['db_filename'])
    cur = con.cursor()
    # drop response table and create new one if all testcases required
    # to reacquisition
    all_resps = False
    if (len(test_ids) == len(found_reqs)):
        all_resps = True
    if all_resps:
        try:
            app_log.debug(f"{inspect.stack()[0][3]}() "
                          f"drop table {TBL_RESPS_NAME}")
            cur.execute('DROP TABLE ' + TBL_RESPS_NAME)
        except Exception as OperationalError:
            # awkward but bearable
            app_log.debug('Missing table %s', TBL_RESPS_NAME)
        cur.execute('CREATE TABLE ' + TBL_RESPS_NAME +
                    ' (test_id varchar(50), data json, hash varchar(255))')

    app_log.info(f"Number of tests to make acquisition - {len(test_ids)}")
    for test_id in test_ids:
        req_id = ids[test_id][0]
        app_log.debug(f"Request: {req_id}")
        resp = _send_recv(cfg, json.dumps(found_reqs[test_id][0]))
        if isinstance(resp, type(None)):
            app_log.error(f"Test {test_ids} ({req_id}) is Failed.")
            continue

        json_lookup('availabilityExpireTime', resp, '0')
        upd_data = json.dumps(resp, sort_keys=True)
        hash_obj = hashlib.sha256(upd_data.encode('utf-8'))
        app_log.debug(f"{inspect.stack()[0][3]}() new: "
                      f"{hash_obj.hexdigest()}")
        if all_resps:
            cur.execute('INSERT INTO ' + TBL_RESPS_NAME + ' values ( ?, ?, ?)',
                        [req_id,
                         upd_data,
                         hash_obj.hexdigest()])
            con.commit()
        elif (test_id in found_resp.keys() and
              found_resp[test_id][1] == hash_obj.hexdigest()):
            app_log.debug(f"Skip to update hash for {req_id}. "
                          f"Found the same value.")
            continue
        else:
            hash = hash_obj.hexdigest()
            cur.execute('UPDATE ' + TBL_RESPS_NAME + ' SET ' +
                        'data = ?, hash = ? WHERE test_id =?',
                        (upd_data, hash, ids[test_id][0]))
            con.commit()
    con.close()
    return AFC_OK


def process_jsonline(line):
    """
    Function to process the input line from .txt file containing comma
    separated json strings
    """

    # convert the input line to list of dictioanry/dictionaries
    line_list = json.loads("[" + line + "]")
    request_dict = line_list[0]
    metadata_dict = line_list[1] if len(line_list) > 1 else {}

    return request_dict, metadata_dict


def get_db_req_resp(cfg):
    """
    Function to retrieve request and response records from the database
    """
    con = sqlite3.connect(cfg['db_filename'])
    cur = con.cursor()
    cur.execute('SELECT * FROM %s' % TBL_REQS_NAME)
    found_reqs = cur.fetchall()
    db_reqs_list = [row[0] for row in found_reqs]

    cur.execute('SELECT * FROM %s' % TBL_RESPS_NAME)
    found_resps = cur.fetchall()
    db_resp_list = [row[0] for row in found_resps]
    con.close()

    return db_reqs_list, db_resp_list


def insert_reqs_int(filename, con, cur):
    """
    Insert requests from input file to a table in test db.

    """
    with open(filename, 'r') as fp_test:
        while True:
            dataline = fp_test.readline()
            if not dataline:
                break

            # process dataline arguments
            app_log.debug(f"= {dataline}")
            request_json, metadata_json = process_jsonline(dataline)

            # reject the request if mandatory metadata arguments are not
            # present
            if not MANDATORY_METADATA_KEYS.issubset(
                    set(metadata_json.keys())):
                # missing mandatory keys in test case input
                app_log.error("Test case input does not contain required"
                              " mandatory arguments: %s",
                              ", ".join(list(
                                  MANDATORY_METADATA_KEYS - set(metadata_json.keys()))))
                return AFC_ERR

            app_log.info(f"Insert new request in DB "
                         f"({metadata_json[TESTCASE_ID]})")
            app_log.debug(f"+ {metadata_json[TESTCASE_ID]}")
            cur.execute('INSERT INTO ' + TBL_REQS_NAME + ' VALUES ( ?, ?)',
                        (metadata_json[TESTCASE_ID],
                         json.dumps(request_json),))
            con.commit()
    con.close()
    return AFC_OK


def insert_reqs(cfg):
    """
    Insert requests from input file to a table in test db.
    Drop previous table of requests.

    """
    app_log.debug(f"{inspect.stack()[0][3]}()")

    if isinstance(cfg['infile'], type(None)):
        app_log.error(f"Missing input file")
        return AFC_ERR

    filename = cfg['infile'][0]
    app_log.debug(f"{inspect.stack()[0][3]}() {filename}")

    if not os.path.isfile(filename):
        app_log.error(f"Missing raw test data file {filename}")
        return AFC_ERR

    if not os.path.isfile(cfg['db_filename']):
        app_log.error(f"Unable to find test db file.")
        return AFC_ERR

    con = sqlite3.connect(cfg['db_filename'])
    # drop existing table of requests and create new one
    app_log.info(f"Drop table of requests ({TBL_REQS_NAME})")
    cur = con.cursor()
    try:
        cur.execute('DROP TABLE ' + TBL_REQS_NAME)
    except Exception as OperationalError:
        app_log.debug(f"Fail to drop, missing table {TBL_REQS_NAME}")
    cur.execute('CREATE TABLE ' + TBL_REQS_NAME +
                ' (test_id varchar(50), data json)')
    con.commit()
    return insert_reqs_int(filename, con, cur)


def extend_reqs(cfg):
    """
    Insert requests from input file to a table in test db.
    Drop previous table of requests.

    """
    app_log.debug(f"{inspect.stack()[0][3]}()")

    if isinstance(cfg['infile'], type(None)):
        app_log.error(f"Missing input file")
        return AFC_ERR

    filename = cfg['infile'][0]
    app_log.debug(f"{inspect.stack()[0][3]}() {filename}")

    if not os.path.isfile(filename):
        app_log.error(f"Missing raw test data file {filename}")
        return AFC_ERR

    if not os.path.isfile(cfg['db_filename']):
        app_log.error(f"Unable to find test db file.")
        return AFC_ERR

    con = sqlite3.connect(cfg['db_filename'])
    # add more rows to existing table of requests
    app_log.info(f"Extending table of requests ({TBL_REQS_NAME})")
    cur = con.cursor()
    return insert_reqs_int(filename, con, cur)


def insert_devs(cfg):
    """
    Insert device descriptors from input file to a table in test db.
    Drop previous table of devices.
    """
    app_log.debug(f"{inspect.stack()[0][3]}()")

    if isinstance(cfg['infile'], type(None)):
        app_log.error(f"Missing input file")
        return AFC_ERR

    filename = cfg['infile'][0]
    app_log.debug(f"{inspect.stack()[0][3]}() {filename}")

    if not os.path.isfile(filename):
        app_log.error(f"Missing raw test data file {filename}")
        return AFC_ERR

    if not os.path.isfile(cfg['db_filename']):
        app_log.error(f"Unable to find test db file.")
        return AFC_ERR

    con = sqlite3.connect(cfg['db_filename'])
    # drop existing table of requests and create new one
    app_log.info(f"Drop table of devices ({TBL_AP_CFG_NAME})")
    cur = con.cursor()
    try:
        cur.execute('DROP TABLE ' + TBL_AP_CFG_NAME)
    except Exception as OperationalError:
        app_log.debug(f"Fail to drop, missing table {TBL_AP_CFG_NAME}")
    cur.execute('CREATE TABLE ' + TBL_AP_CFG_NAME +
                ' (ap_config_id, data json, user_id)')
    cnt = 1
    con.commit()
    with open(filename, 'r') as fp_test:
        while True:
            dataline = fp_test.readline()
            if not dataline or (len(dataline) < 72):
                break

            # process dataline arguments
            app_log.debug(f"= {dataline}")

            cur.execute('INSERT INTO ' + TBL_AP_CFG_NAME +
                        ' VALUES ( ?, ?, ?)', (cnt, dataline[:-1], 1))
            con.commit()
            cnt += 1
    con.close()
    return AFC_OK


def add_reqs(cfg):
    """Prepare DB source files"""
    app_log.debug(f"{inspect.stack()[0][3]}()")

    if isinstance(cfg['infile'], type(None)):
        app_log.error('Missing input file')
        return AFC_ERR

    filename = cfg['infile'][0]
    app_log.debug('%s() %s', inspect.stack()[0][3], filename)

    if not os.path.isfile(filename):
        app_log.error('Missing raw test data file %s', filename)
        return AFC_ERR

    if not make_db(cfg['db_filename']):
        return AFC_ERR

    # fetch available requests and responses
    db_reqs_list, db_resp_list = get_db_req_resp(cfg)

    con = sqlite3.connect(cfg['db_filename'])
    with open(filename, 'r') as fp_test:
        while True:
            dataline = fp_test.readline()
            if not dataline:
                break

            # process dataline arguments
            request_json, metadata_json = process_jsonline(dataline)

            # reject the request if mandatory metadata arguments are not
            # present
            if not MANDATORY_METADATA_KEYS.issubset(
                    set(metadata_json.keys())):
                # missing mandatory keys in test case input
                app_log.error("Test case input does not contain required"
                              " mandatory arguments: %s",
                              ", ".join(list(
                                  MANDATORY_METADATA_KEYS - set(metadata_json.keys()))))
                return AFC_ERR

            # check if the test case already exists in the database test
            # vectors
            if metadata_json[TESTCASE_ID] in db_reqs_list:
                app_log.error("Test case: %s already exists in database",
                              metadata_json[TESTCASE_ID])
                break

            app_log.info("Executing test case: %s", metadata_json[TESTCASE_ID])
            new_req, resp = cfg._send_recv(json.dumps(request_json))

            # get request id from a request, response not always has it
            # the request contains test category
            new_req_json = json.loads(new_req.encode('utf-8'))
            req_id = json_lookup('requestId', new_req_json, None)

            resp_res = json_lookup('shortDescription', resp, None)
            if (resp_res[0] != 'Success') \
                and (req_id[0].lower().find('urs') == -1) \
                    and (req_id[0].lower().find('ibp') == -1):
                app_log.error('Failed in test response - %s', resp_res)
                break

            app_log.info('Got response for the request')
            json_lookup('availabilityExpireTime', resp, '0')

            app_log.info('Insert new request in DB')
            cur = con.cursor()
            cur.execute('INSERT INTO ' + TBL_REQS_NAME + ' VALUES ( ?, ?)',
                        (metadata_json[TESTCASE_ID], new_req,))
            con.commit()

            app_log.info('Insert new resp in DB')
            upd_data = json.dumps(resp, sort_keys=True)
            hash_obj = hashlib.sha256(upd_data.encode('utf-8'))
            cur = con.cursor()
            cur.execute('INSERT INTO ' + TBL_RESPS_NAME + ' values ( ?, ?, ?)',
                        [metadata_json[TESTCASE_ID],
                         upd_data, hash_obj.hexdigest()])
            con.commit()
    con.close()
    return AFC_OK


def dump_table(conn, tbl_name, out_file, pref):
    app_log.debug(f"{inspect.stack()[0][3]}() {tbl_name}")
    fp_new = ''

    if 'single' in out_file:
        fp_new = open(out_file['single'], 'w')

    conn.execute(f"SELECT * FROM {tbl_name}")
    found_data = conn.fetchall()
    for val in enumerate(found_data):
        if isinstance(fp_new, io.IOBase):
            fp_new.write(f"{str(val)}\n")
        elif 'split' in out_file:
            tbl_fname = {
                TBL_REQS_NAME: '_Request.txt',
                TBL_RESPS_NAME: '_Response.txt'
            }
            new_json = json.loads(val[1][1].encode('utf-8'))
            prefix, name, nbr = val[1][0].split('.')
            app_log.debug(f"{inspect.stack()[0][3]}() {name} {nbr}")
            # omit URS testcases
            if (name.lower().find('urs') != -1) or (pref and not pref == prefix):
                continue

            fp_test = open(f"{out_file['split']}/{prefix}_{name}_{nbr}" +
                           f"{tbl_fname[tbl_name]}", 'a')
            fp_test.write(f"{val[1][1]}\n")
            fp_test.close()
        else:
            # Just dump to the console
            app_log.info(f"{val[1]}")

    if isinstance(fp_new, io.IOBase):
        fp_new.close()


def dump_database(cfg):
    """Dump data from test DB tables"""
    app_log.debug(f"{inspect.stack()[0][3]}()")
    find_key = ''
    found_tables = []
    # keep configuration for output path and files
    # 'single' - only single file for whole output
    # 'split' - separate file for each response
    out_file = {}

    if not os.path.isfile(cfg['db_filename']):
        app_log.error(f"Missing DB file {cfg['db_filename']}")
        return AFC_ERR

    set_dump_db_opts = {
        'wfa': [(TBL_REQS_NAME,), (TBL_RESPS_NAME,)],
        'all': [(TBL_REQS_NAME,), (TBL_RESPS_NAME,)],
        'req': [(TBL_REQS_NAME,)],
        'resp': [(TBL_RESPS_NAME,)],
        'ap': [('ap_config',)],
        'cfg': [('afc_config',)],
        'user': [('user_config',)]
    }

    prefix = {
        'wfa': "AFCS",
        'all': None
    }

    tbl = 'True'
    if isinstance(cfg['table'], list):
        tbl = cfg['table'][0]
    con = sqlite3.connect(cfg['db_filename'])
    cur = con.cursor()

    if tbl in set_dump_db_opts:
        # Dump only tables with requests and responses
        found_tables.extend(set_dump_db_opts[tbl])
    elif tbl == 'True':
        # Dump all tables if no options provided
        cur.execute(f"SELECT name FROM sqlite_master WHERE type='table';")
        found_tables = cur.fetchall()

    pref = None
    if tbl == 'wfa' or tbl == 'all':
        if tbl in prefix:
            pref = prefix[tbl]

        out_file['split'] = './'
        if not isinstance(cfg['outpath'], type(None)):
            out_file['split'] = cfg['outpath'][0] + '/'

        out_file['split'] += WFA_TEST_DIR
        if os.path.exists(out_file['split']):
            shutil.rmtree(out_file['split'])
        os.mkdir(out_file['split'])
    elif isinstance(cfg['outfile'], type(None)):
        app_log.error(f"Missing output filename.\n")
        return AFC_ERR
    else:
        out_file['single'] = cfg['outfile'][0]

    for tbl in enumerate(found_tables):
        app_log.debug(f"Dump {tbl} to {out_file}")
        dump_table(cur, tbl[1][0], out_file, pref)
    con.close()
    return AFC_OK


def export_admin_config(cfg):
    """Export admin server configuration"""
    app_log.debug('%s()', inspect.stack()[0][3])

    con = sqlite3.connect(cfg['db_filename'])
    cur = con.cursor()
    cur.execute('SELECT COUNT(*) FROM ' + TBL_AP_CFG_NAME)
    found_rcds = cur.fetchall()
    with open(cfg['outfile'][0], 'w') as fp_exp:
        cur.execute('SELECT * FROM %s' % TBL_AFC_CFG_NAME)
        found_cfg = cur.fetchall()
        app_log.debug('Found AfcCfg: %s', found_cfg)

        cur.execute('SELECT * FROM %s\n' % TBL_USERS_NAME)
        found_user = cur.fetchall()
        app_log.debug('Found Users: %s\n', found_user)

        cur.execute('SELECT * FROM %s\n' % TBL_AP_CFG_NAME)
        found_aps = cur.fetchall()
        con.close()

        afc = ''
        for count, val in enumerate(found_cfg):
            afc += str(val[1]) + ','
        app_log.debug('Found AFCs: %s\n', afc[:-1])

        aps = ''
        idx = 0
        for count, val in enumerate(found_aps):
            aps += str(val[1]) + ','
        app_log.debug('Found APs: %s\n', aps[:-1])

        out_str = '{"afcAdminConfig":{ "afcConfigs": [' + afc[:-1] + ']}' + ', '\
                  '"userConfig":' + found_user[0][1] + ', '\
                  '"apConfig":[' + aps[:-1] + ']}'
        fp_exp.write(out_str)
    app_log.info('Server admin config exported to %s', cfg['outfile'][0])
    return AFC_OK


def dry_run_test(cfg):
    """Run one or more requests from provided file"""
    if isinstance(cfg['infile'], type(None)):
        app_log.error('Missing input file')
        return AFC_ERR

    filename = cfg['infile'][0]
    app_log.debug('%s() %s', inspect.stack()[0][3], filename)

    if not os.path.isfile(filename):
        app_log.error('Missing raw test data file %s', filename)
        return AFC_ERR

    with open(filename, 'r') as fp_test:
        while True:
            dataline = fp_test.readline()
            if not dataline:
                break
            app_log.info('Request:')
            app_log.info(dataline)

            # process dataline arguments
            request_json, _ = process_jsonline(dataline)

            resp = _send_recv(cfg, json.dumps(request_json))

            # get request id from a request, response not always has it
            # the request contains test category
            new_req_json = json.loads(json.dumps(request_json).encode('utf-8'))
            req_id = json_lookup('requestId', new_req_json, None)

            resp_res = json_lookup('shortDescription', resp, None)
            if (resp_res[0] != 'Success') \
                    and (req_id[0].lower().find('urs') == -1):
                app_log.error('Failed in test response - %s', resp_res)
                app_log.debug(resp)
                break

            app_log.info('Got response for the request')
            app_log.info('Resp:')
            app_log.info(resp)
            app_log.info('\n\n')

            json_lookup('availabilityExpireTime', resp, '0')
            upd_data = json.dumps(resp, sort_keys=True)
            hash_obj = hashlib.sha256(upd_data.encode('utf-8'))
    return AFC_OK


def get_nbr_testcases(cfg):
    """ Find APs count on DB table """
    if not os.path.isfile(cfg['db_filename']):
        print('INFO: Missing DB file %s', cfg['db_filename'])
        return False
    con = sqlite3.connect(cfg['db_filename'])
    cur = con.cursor()
    cur.execute('SELECT count("requestId") from ' + TBL_REQS_NAME)
    found_data = cur.fetchall()
    db_inquiry_count = found_data[0][0]
    con.close()
    app_log.debug("found %s ap lists from db table", db_inquiry_count)
    return db_inquiry_count


def collect_tests2combine(sh, rows, t_ident, t2cmb, cmb_t):
    """
       Lookup for combined test vectors,
       build of required test vectors to combine
    """
    app_log.debug('%s()\n', inspect.stack()[0][3])
    for i in range(1, rows + 1):
        cell = sh.cell(row=i, column=PURPOSE_CLM)
        if ((cell.value is None) or
            (AFC_TEST_IDENT.get(cell.value.lower()) is None) or
                (cell.value == 'SRI')):
            continue

        if (t_ident != 'all') and (cell.value.lower() != t_ident):
            continue

        cell = sh.cell(row=i, column=COMBINED_CLM)
        if cell.value is not None and \
                cell.value.upper() != 'NO':
            raw_list = str(cell.value)

            test_case_id = sh.cell(row=i, column=UNIT_NAME_CLM).value
            test_case_id += "."
            test_case_id += sh.cell(row=i, column=PURPOSE_CLM).value
            test_case_id += "."
            test_case_id += str(sh.cell(row=i, column=TEST_VEC_CLM).value)

            cmb_t[test_case_id] = []
            for t in raw_list.split(','):
                if '-' in t:
                    # found range of test vectors
                    left, right = t.split('-')
                    t2cmb_ident = ''
                    for r in AFC_TEST_IDENT:
                        if r in left.lower():
                            min = int(left.replace(r.upper(), ''))
                            max = int(right.replace(r.upper(), '')) + 1
                            t2cmb_ident = r.upper()
                            for cnt in range(min, max):
                                tcase = t2cmb_ident + str(cnt)
                                t2cmb[tcase] = ''
                                cmb_t[test_case_id] += [tcase]
                else:
                    # found single test vector
                    t2cmb[t] = ''
                    cmb_t[test_case_id] += [t]


def _parse_tests_dev_desc(sheet, fp_new, rows):
    app_log.debug('%s()\n', inspect.stack()[0][3])
    for i in range(1, rows + 1):
        res_str = ""
        cell = sheet.cell(row=i, column=PURPOSE_CLM)
        if ((cell.value is None) or
            (AFC_TEST_IDENT.get(cell.value.lower()) is None) or
                (cell.value == 'SRI')):
            continue

        # skip combined test vectors because device descriptor is missing
        cell = sheet.cell(row=i, column=COMBINED_CLM)
        if cell.value is not None and \
                cell.value.upper() != 'NO':
            continue

        res_str += build_device_desc(
            sheet.cell(row=i, column=INDOOR_DEPL_CLM).value,
            sheet.cell(row=i, column=SER_NBR_CLM).value,
            sheet.cell(row=i, column=RULESET_CLM).value,
            sheet.cell(row=i, column=ID_CLM).value,
            True)

        fp_new.write(res_str + '\n')
    return res_str


def _parse_tests_all(sheet, fp_new, rows, test_ident):
    app_log.debug('%s()\n', inspect.stack()[0][3])
    # collect tests to combine in next loop
    tests2combine = dict()
    # gather combined tests
    combined_tests = dict()
    collect_tests2combine(sheet, rows, test_ident,
                          tests2combine,
                          combined_tests)
    if len(combined_tests):
        app_log.info('Found combined test vectors: %s',
                     ' '.join(combined_tests))
        app_log.info('Found test vectors to combine: %s',
                     ' '.join(tests2combine))
    for i in range(1, rows + 1):
        cell = sheet.cell(row=i, column=PURPOSE_CLM)
        if ((cell.value is None) or
            (AFC_TEST_IDENT.get(cell.value.lower()) is None) or
                (cell.value == 'SRI')):
            continue

        if (test_ident != 'all') and (cell.value.lower() != test_ident):
            continue

        uut = sheet.cell(row=i, column=UNIT_NAME_CLM).value
        purpose = sheet.cell(row=i, column=PURPOSE_CLM).value
        test_vec = sheet.cell(row=i, column=TEST_VEC_CLM).value

        test_case_id = uut + "." + purpose + "." + str(test_vec)

        # Prepare request header '{"availableSpectrumInquiryRequests": [{'
        res_str = REQ_INQUIRY_HEADER
        # check if the test case is combined
        cell = sheet.cell(row=i, column=COMBINED_CLM)
        if cell.value is not None and \
                cell.value.upper() != 'NO':
            for item in combined_tests[test_case_id]:
                res_str += tests2combine[item] + ','
            res_str = res_str[:-1]
        else:
            #
            # Inquired Channels
            #
            res_str += '{' + REQ_INQ_CHA_HEADER
            cell = sheet.cell(row=i, column=GLOBALOPERATINGCLASS_131)
            res_str += '{' + REQ_INQ_CHA_GL_OPER_CLS + str(cell.value)
            cell = sheet.cell(row=i, column=CHANNEL_CFI_131)
            if cell.value is not None:
                res_str += ', ' + REQ_INQ_CHA_CHANCFI + str(cell.value)
            res_str += '}, '
            cell = sheet.cell(row=i, column=GLOBALOPERATINGCLASS_132)
            res_str += '{' + REQ_INQ_CHA_GL_OPER_CLS + str(cell.value)
            cell = sheet.cell(row=i, column=CHANNEL_CFI_132)
            if cell.value is not None:
                res_str += ', ' + REQ_INQ_CHA_CHANCFI + str(cell.value)
            res_str += '}, '
            cell = sheet.cell(row=i, column=GLOBALOPERATINGCLASS_133)
            res_str += '{' + REQ_INQ_CHA_GL_OPER_CLS + str(cell.value)
            cell = sheet.cell(row=i, column=CHANNEL_CFI_133)
            if cell.value is not None:
                res_str += ', ' + REQ_INQ_CHA_CHANCFI + str(cell.value)
            res_str += '}, '
            cell = sheet.cell(row=i, column=GLOBALOPERATINGCLASS_134)
            res_str += '{' + REQ_INQ_CHA_GL_OPER_CLS + str(cell.value)
            cell = sheet.cell(row=i, column=CHANNEL_CFI_134)
            if cell.value is not None:
                res_str += ', ' + REQ_INQ_CHA_CHANCFI + str(cell.value)
            res_str += '}, '
            cell = sheet.cell(row=i, column=GLOBALOPERATINGCLASS_136)
            res_str += '{' + REQ_INQ_CHA_GL_OPER_CLS + str(cell.value)
            cell = sheet.cell(row=i, column=CHANNEL_CFI_136)
            if cell.value is not None:
                res_str += ', ' + REQ_INQ_CHA_CHANCFI + str(cell.value)
            res_str += '}, '
            cell = sheet.cell(row=i, column=GLOBALOPERATINGCLASS_137)
            res_str += '{' + REQ_INQ_CHA_GL_OPER_CLS + str(cell.value)
            cell = sheet.cell(row=i, column=CHANNEL_CFI_137)
            if cell.value is not None:
                res_str += ', ' + REQ_INQ_CHA_CHANCFI + str(cell.value)
            res_str += '}' + REQ_INQ_CHA_FOOTER + ' '
            #
            # Device descriptor
            #
            res_str += REQ_DEV_DESC_HEADER

            res_str += build_device_desc(
                sheet.cell(row=i, column=INDOOR_DEPL_CLM).value,
                sheet.cell(row=i, column=SER_NBR_CLM).value,
                sheet.cell(row=i, column=RULESET_CLM).value,
                sheet.cell(row=i, column=ID_CLM).value,
                False)
            res_str += ','
            #
            # Inquired Frequency Range
            #
            res_str += REQ_INQ_FREQ_RANG_HEADER
            freq_range = AfcFreqRange()
            freq_range.set_range_limit(
                sheet.cell(
                    row=i,
                    column=INQ_FREQ_RNG_LOWFREQ_A),
                'low')
            freq_range.set_range_limit(
                sheet.cell(
                    row=i,
                    column=INQ_FREQ_RNG_HIGHFREQ_A),
                'high')
            try:
                res_str += freq_range.append_range()
            except IncompleteFreqRange as e:
                app_log.debug(f"{e} -  row {i}")
            freq_range = AfcFreqRange()
            freq_range.set_range_limit(
                sheet.cell(
                    row=i,
                    column=INQ_FREQ_RNG_LOWFREQ_B),
                'low')
            freq_range.set_range_limit(
                sheet.cell(
                    row=i,
                    column=INQ_FREQ_RNG_HIGHFREQ_B),
                'high')
            try:
                tmp_str = freq_range.append_range()
                res_str += ', ' + tmp_str
            except IncompleteFreqRange as e:
                app_log.debug(f"{e} -  row {i}")
            res_str += REQ_INQ_FREQ_RANG_FOOTER

            cell = sheet.cell(row=i, column=MINDESIREDPOWER)
            if (cell.value):
                res_str += REQ_MIN_DESIRD_PWR + str(cell.value) + ', '
            #
            # Location
            #
            res_str += REQ_LOC_HEADER
            cell = sheet.cell(row=i, column=INDOORDEPLOYMENT)
            res_str += REQ_LOC_INDOORDEPL + str(cell.value) + ', '

            # Location - elevation
            res_str += REQ_LOC_ELEV_HEADER
            cell = sheet.cell(row=i, column=ELE_VERTICALUNCERTAINTY)
            if isinstance(cell.value, int):
                res_str += REQ_LOC_VERT_UNCERT + str(cell.value) + ', '
            cell = sheet.cell(row=i, column=ELE_HEIGHTTYPE)
            res_str += REQ_LOC_HEIGHT_TYPE + '"' + str(cell.value) + '"'
            cell = sheet.cell(row=i, column=ELE_HEIGHT)
            if isinstance(cell.value, int) or isinstance(cell.value, float):
                res_str += ', ' + REQ_LOC_HEIGHT + str(cell.value)
            res_str += '}, '

            # Location - uncertainty reqion
            geo_coor = AfcGeoCoordinates(sheet, i)
            try:
                res_str += geo_coor.collect_coordinates()
            except IncompleteGeoCoordinates as e:
                app_log.debug(e)

            res_str += REQ_LOC_FOOTER

            cell = sheet.cell(row=i, column=REQ_ID_CLM)
            if isinstance(cell.value, str):
                req_id = cell.value
            else:
                req_id = ""
            res_str += REQ_REQUEST_ID + '"' + req_id + '"'
            res_str += '}'

            # collect test vectors required for combining
            # build test case id in format <purpose><test vector>
            short_tcid = ''.join(test_case_id.split('.', 1)[1].split('.'))
            if short_tcid in tests2combine.keys():
                tests2combine[short_tcid] = res_str.split(
                    REQ_INQUIRY_HEADER)[1]

        res_str += REQ_INQUIRY_FOOTER
        cell = sheet.cell(row=i, column=VERSION_CLM)
        res_str += REQ_VERSION + '"' + str(cell.value) + '"'
        res_str += REQ_FOOTER

        # adding metadata parameters
        res_str += ', '

        # adding test case id
        res_str += META_HEADER
        res_str += META_TESTCASE_ID + '"' + test_case_id + '"'
        res_str += META_FOOTER
        fp_new.write(res_str + '\n')
    return AFC_OK


def parse_tests(cfg):
    app_log.debug('%s()\n', inspect.stack()[0][3])
    filename = ''
    out_fname = ''

    if isinstance(cfg['infile'], type(None)):
        app_log.error('Missing input file')
        return AFC_ERR

    filename = cfg['infile'][0]

    if not os.path.isfile(filename):
        app_log.error('Missing raw test data file %s', filename)
        return AFC_ERR

    test_ident = cfg['test_id']

    if isinstance(cfg['outfile'], type(None)):
        out_fname = test_ident + NEW_AFC_TEST_SUFX
    else:
        out_fname = cfg['outfile'][0]

    wb = oxl.load_workbook(filename, data_only=True)

    for sh in wb:
        app_log.debug('Sheet title: %s', sh.title)
        app_log.debug('rows %d, cols %d\n', sh.max_row, sh.max_column)

    # Look for request sheet
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == "Availability Requests":
            break

    wb.active = s
    sheet = wb.active
    nbr_rows = sheet.max_row
    app_log.debug('Rows range 1 - %d', nbr_rows + 1)
    app_log.info('Export tests into file: %s', out_fname)
    fp_new = open(out_fname, 'w')

    # partial parse if only certain parameters required.
    # only for device descriptor for now.
    if cfg['dev_desc']:
        res_str = _parse_tests_dev_desc(sheet, fp_new, nbr_rows)
        fp_new.write(res_str + '\n')
    else:
        _parse_tests_all(sheet, fp_new, nbr_rows, test_ident)

    fp_new.close()
    return AFC_OK


def test_report(fname, runtimedata, testnumdata, testvectordata,
                test_result, upd_data):
    """Procedure to generate AFC test results report.
    Args:
        runtimedata(str) : Tested case running time.
        testnumdata(str): Tested case id.
        testvectordata(int): Tested vector id.
        test_result(str: Test result Pass is 0 or Fail is 1)
        upd_data(list): List for test response data
    Return:
        Create test results file.
    """
    ts_time = datetime.datetime.fromtimestamp(time.time()).\
        strftime('%Y_%m_%d_%H%M%S')
    # Test results output args
    data_names = ['Date', 'Test Number', 'Test Vector', 'Running Time',
                  'Test Result', 'Response data']
    data = {'Date': ts_time, 'Test Number': testnumdata,
            'Test Vector': testvectordata, 'Running Time': runtimedata,
            'Test Result': test_result, 'Response data': upd_data}
    with open(fname, "a") as f:
        file_writer = csv.DictWriter(f, fieldnames=data_names)
        if os.stat(fname).st_size == 0:
            file_writer.writeheader()
        file_writer.writerow(data)


def _run_tests(cfg, reqs, resps, comparator, ids, test_cases):
    """
    Run tests
    reqs: {testcaseid: [request_json_str]}
    resps: {testcaseid: [response_json_str, response_hash]}
    comparator: reference to object
    test_cases: [testcaseids]
    """
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}() {test_cases} "
                  f"{cfg['url_path']}")

    if not len(resps):
        app_log.info(f"Unable to compare response data."
                     f"Suggest to make acquisition of responses.")
        return AFC_ERR

    all_test_res = AFC_OK
    accum_secs = 0

    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}() "
                  f"{type(cfg['webui'])} ")
    ssn = None
    if cfg['webui'] is True:
        ssn = requests.Session()
        _login(cfg, ssn)

    for test_case in test_cases:
        # Default reset test_res value
        test_res = AFC_OK

        req_id = ids[test_case][0]
        app_log.info(f"Prepare to run test - {req_id}")
        if test_case not in reqs:
            app_log.warning(f"The requested test case {test_case} is "
                            f"invalid/not available in database")
            continue

        request_data = reqs[test_case][0]
        app_log.debug(f"{inspect.stack()[0][3]}() {request_data}")

        before_ts = time.monotonic()
        resp = _send_recv(cfg, json.dumps(request_data), ssn)
        tm_secs = time.monotonic() - before_ts
        res = f"id {test_case} name {req_id} status $status time {tm_secs:.1f}"
        res_template = Template(res)

        if isinstance(resp, type(None)):
            test_res = AFC_ERR
            all_test_res = AFC_ERR
        elif 'error' in resp:
            app_log.error(f"Test case {req_id} returned error: {resp['error']}")
            test_res = AFC_ERR
        else:
            if cfg['webui'] is True:
                # remove the mapping info from the response
                # to make sure the base data matches - not checking map results
                parent = resp['availableSpectrumInquiryResponses'][0]
                if 'vendorExtensions' in parent:
                    parent.pop('vendorExtensions')

            json_lookup('availabilityExpireTime', resp, '0')
            upd_data = json.dumps(resp, sort_keys=True)

            diffs = []
            hash_obj = hashlib.sha256(upd_data.encode('utf-8'))
            diffs = comparator.compare_results(ref_str=resps[test_case][0],
                                               result_str=upd_data)
            if (resps[test_case][1] == hash_obj.hexdigest()) \
                    if cfg['precision'] is None else (not diffs):
                res = res_template.substitute(status="Ok")
            else:
                for line in diffs:
                    app_log.error(f"  Difference: {line}")
                app_log.error(hash_obj.hexdigest())
                test_res = AFC_ERR

        if test_res == AFC_ERR:
            res = res_template.substitute(status="Fail")
            app_log.error(res)
            all_test_res = AFC_ERR

        accum_secs += tm_secs
        app_log.info(res)

        # For saving test results option
        if not isinstance(cfg['outfile'], type(None)):
            app_log.warning(f"upd_data: {upd_data}")
            test_report(cfg['outfile'][0], float(tm_secs),
                        test_case, req_id,
                        ("PASS" if test_res == AFC_OK else "FAIL"),
                        upd_data)

    app_log.info(f"Total testcases runtime : {round(accum_secs, 1)} secs")

    if not isinstance(cfg['outfile'], type(None)):
        send_email(cfg)

    return all_test_res


def prep_and_run_tests(cfg, reqs, resps, ids, test_cases):
    """
    Run tests
    reqs: {testcaseid: [request_json_str]}
    resps: {testcaseid: [response_json_str, response_hash]}
    test_cases: [testcaseids]
    """
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()")

    test_res = AFC_OK
    results_comparator = TestResultComparator(precision=cfg['precision'] or 0)

    # calculate max number of tests to run
    max_nbr_tests = len(test_cases)
    if not isinstance(cfg['tests2run'], type(None)):
        max_nbr_tests = int("".join(cfg['tests2run']))

    while (max_nbr_tests != 0):
        app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}() "
                      f"Number of tests to run: {max_nbr_tests}")
        if max_nbr_tests < len(test_cases):
            del test_cases[-(len(test_cases) - max_nbr_tests):]

        # if stress mode execute testcases in parallel and concurrent
        if cfg['stress'] == 1:
            app_log.debug(f"{inspect.stack()[0][3]}() max {max_nbr_tests}"
                          f" len {len(test_cases)}")
            inputs = [(cfg, reqs, resps, results_comparator, ids, [test])
                      for test in test_cases]
            with Pool(max_nbr_tests) as my_pool:
                results = my_pool.starmap(_run_tests, inputs)
            if not any(r == 0 for r in results):
                test_res = AFC_ERR
        else:
            res = _run_tests(cfg, reqs, resps,
                             results_comparator, ids, test_cases)
            if res != AFC_OK:
                test_res = res
        # when required to run more tests than there are testcases
        # start run run from the beginning
        max_nbr_tests -= len(test_cases)
    return test_res


def _convert_reqs_n_resps_to_dict(cfg):
    """
       Fetch test vectors and responses from the DB
       Convert to dictionaries indexed by id or original index from table
       Prepare list of new indexes
    """
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()")

    if not os.path.isfile(cfg['db_filename']):
        app_log.error('Missing DB file %s', cfg['db_filename'])
        return AFC_ERR

    con = sqlite3.connect(cfg['db_filename'])
    cur = con.cursor()
    cur.execute('SELECT * FROM %s' % TBL_REQS_NAME)
    found_reqs = cur.fetchall()
    cur.execute('SELECT * FROM %s' % TBL_RESPS_NAME)
    found_resps = cur.fetchall()
    con.close()

    # reformat the reqs_dict and resp_dict accordingly
    if not isinstance(cfg['testcase_ids'], type(None)):
        app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}() by id")
        reqs_dict = {
            row[0]: [json.loads(row[1])]
            for row in found_reqs
        }
        resp_dict = {
            row[0]: [row[1], row[2]]
            for row in found_resps
        }
        ids_dict = {
            row[0]: [row[0], req_index + 1]
            for req_index, row in enumerate(found_reqs)
        }
        test_indx = list(map(str.strip, cfg.pop("testcase_ids").split(',')))
        cfg.pop("testcase_indexes")
    else:
        app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}() by index")
        reqs_dict = {
            str(req_index + 1): [json.loads(row[1])]
            for req_index, row in enumerate(found_reqs)
        }
        resp_dict = {
            str(resp_index + 1): [row[1], row[2]]
            for resp_index, row in enumerate(found_resps)
        }
        ids_dict = {
            str(req_index + 1): [row[0], req_index + 1]
            for req_index, row in enumerate(found_reqs)
        }
        if not isinstance(cfg['testcase_indexes'], type(None)):
            test_indx = list(map(str.strip,
                                 cfg.pop("testcase_indexes").split(',')))
            cfg.pop("testcase_ids")
        else:
            test_indx = [
                str(item) for item in list(range(1, len(reqs_dict) + 1))
            ]

    # build list of indexes, omitting non-existing elements
    test_cases = list()
    for i in range(0, len(test_indx)):
        if reqs_dict.get(test_indx[i]) is None:
            app_log.debug(f"Missing value for index {test_indx[i]}")
            continue
        test_cases.append(test_indx[i])
    app_log.debug(f"{inspect.stack()[0][3]}() Final list of indexes. "
                  f"{test_cases}")

    return reqs_dict, resp_dict, ids_dict, test_cases


def run_test(cfg):
    """Fetch test vectors from the DB and run tests"""
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}() "
                  f"{cfg['tests']}, {cfg['url_path']}")

    reqs_dict, resp_dict, ids, test_cases = _convert_reqs_n_resps_to_dict(cfg)
    return prep_and_run_tests(cfg, reqs_dict, resp_dict, ids, test_cases)


def stress_run(cfg):
    """
    Get test vectors from the database and run tests
    in parallel
    """
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}() "
                  f"{cfg['tests']}, {cfg['url_path']}")
    cfg['stress'] = 1

    return run_test(cfg)


def _run_cert_tests(cfg):
    """
    Run tests
    """
    app_log.debug(
        f"({os.getpid()}) {inspect.stack()[0][3]}() {cfg['url_path']}")

    test_res = AFC_OK

    try:
        if isinstance(cfg['cli_cert'], type(None)):
            rawresp = requests.get(cfg['url_path'],
                                   verify="".join(cfg['ca_cert']))
        else:
            rawresp = requests.get(cfg['url_path'],
                                   cert=("".join(cfg['cli_cert']),
                                         "".join(cfg['cli_key'])),
                                   verify="".join(cfg['ca_cert']))
    except Exception as e:
        app_log.error(f"({os.getpid()}) {inspect.stack()[0][3]}() {e}")
        test_res = AFC_ERR
    except OSError as os_err:
        proc = psutil.Process()
        app_log.error(f"({os.getpid()}) {inspect.stack()[0][3]}() "
                      f"{os_err} - {proc.open_files()}")
        test_res = AFC_ERR
    else:
        if rawresp.status_code != 200:
            test_res = AFC_ERR

    return test_res


def run_cert_tests(cfg):
    """
    Run certificate tests
    """
    app_log.debug(f"({os.getpid()}) {inspect.stack()[0][3]}()")

    test_res = AFC_OK

    # calculate max number of tests to run
    if isinstance(cfg['tests2run'], type(None)):
        app_log.error(f"Missing number of tests to run.")
        return AFC_ERR

    max_nbr_tests = int("".join(cfg['tests2run']))

    before_ts = time.monotonic()
    while (max_nbr_tests != 0):
        # if stress mode execute testcases in parallel and concurrent
        inputs = [(cfg)]
        with Pool(max_nbr_tests) as my_pool:
            results = my_pool.map(_run_cert_tests, inputs)
        if not any(r == 0 for r in results):
            test_res = AFC_ERR
        max_nbr_tests -= 1
    tm_secs = time.monotonic() - before_ts
    app_log.info(f"Tests {max_nbr_tests} done at {tm_secs:.3f} s")

    return test_res


log_level_map = {
    'debug': logging.DEBUG,    # 10
    'info': logging.INFO,      # 20
    'warn': logging.WARNING,   # 30
    'err': logging.ERROR,      # 40
    'crit': logging.CRITICAL,  # 50
}


def set_log_level(opt):
    app_log.setLevel(log_level_map[opt])
    return log_level_map[opt]


def get_version(cfg):
    """Get AFC test utility version"""
    app_log.info('AFC Test utility version %s', __version__)
    app_log.info('AFC Test db hash %s', get_md5(AFC_TEST_DB_FILENAME))


def pre_hook(cfg):
    """Execute provided functionality prior to main command"""
    app_log.debug(f"{inspect.stack()[0][3]}()")
    return subprocess.call(cfg['prefix_cmd'])


def parse_run_test_args(cfg):
    """Parse arguments for command 'run_test'"""
    app_log.debug(f"{inspect.stack()[0][3]}()")
    if isinstance(cfg['addr'], list):
        # check if provided required certification files
        if (cfg['prot'] != AFC_PROT_NAME):
            # update URL if not the default protocol
            cfg['url_path'] = cfg['prot'] + '://'

        cfg['url_path'] += str(cfg['addr'][0]) + ':' + str(cfg['port'])
        cfg['base_url'] = cfg['url_path'] + '/'
        if cfg['webui'] is False:
            cfg['url_path'] += AFC_URL_SUFFIX + AFC_REQ_NAME
        else:
            cfg['url_path'] += AFC_URL_SUFFIX + AFC_WEBUI_REQ_NAME
    return AFC_OK


def parse_run_cert_args(cfg):
    """Parse arguments for command 'run_cert'"""
    app_log.debug(f"{inspect.stack()[0][3]}()")
    if ((not isinstance(cfg['addr'], list)) or
            isinstance(cfg['ca_cert'], type(None))):
        app_log.error(f"{inspect.stack()[0][3]}() Missing arguments")
        return AFC_ERR

    cfg['url_path'] = cfg['prot'] + '://' + str(cfg['addr'][0]) +\
        ':' + str(cfg['port']) + '/'
    return AFC_OK


# available commands to execute in alphabetical order
execution_map = {
    'add_reqs': [add_reqs, parse_run_test_args],
    'ins_reqs': [insert_reqs, parse_run_test_args],
    'ext_reqs': [extend_reqs, parse_run_test_args],
    'ins_devs': [insert_devs, parse_run_test_args],
    'cmp_cfg': [compare_afc_config, parse_run_test_args],
    'dry_run': [dry_run_test, parse_run_test_args],
    'dump_db': [dump_database, parse_run_test_args],
    'get_nbr_testcases': [get_nbr_testcases, parse_run_test_args],
    'exp_adm_cfg': [export_admin_config, parse_run_test_args],
    'parse_tests': [parse_tests, parse_run_test_args],
    'reacq': [start_acquisition, parse_run_test_args],
    'run': [run_test, parse_run_test_args],
    'run_cert': [run_cert_tests, parse_run_cert_args],
    'stress': [stress_run, parse_run_test_args],
    'ver': [get_version, parse_run_test_args],
}


def make_arg_parser():
    """Define command line options"""
    args_parser = argparse.ArgumentParser(
        epilog=__doc__.strip(),
        formatter_class=argparse.RawTextHelpFormatter)

    args_parser.add_argument('--addr', nargs=1, type=str,
                             help="<address> - set AFC Server address.\n")
    args_parser.add_argument('--prot', nargs='?', choices=['https', 'http'],
                             default='https',
                             help="<http|https> - set connection protocol "
                             "(default=https).\n")
    args_parser.add_argument('--port', nargs='?', default='443',
                             type=int,
                             help="<port> - set connection port "
                             "(default=443).\n")
    args_parser.add_argument('--conn_type', nargs='?',
                             choices=['sync', 'async'], default='sync',
                             help="<sync|async> - set connection to be "
                             "synchronous or asynchronous (default=sync).\n")
    args_parser.add_argument('--conn_tm', nargs='?', default=None, type=int,
                             help="<secs> - set timeout for asynchronous "
                             "connection (default=None). \n")
    args_parser.add_argument('--verif', action='store_true',
                             help="<verif> - skip SSL verification "
                             "on post request.\n")
    args_parser.add_argument('--outfile', nargs=1, type=str,
                             help="<filename> - set filename for output "
                             "of tests results.\n")
    args_parser.add_argument(
        '--outpath',
        nargs=1,
        type=str,
        help="<filepath> - set path to output filename for "
        "results output.\n")
    args_parser.add_argument('--infile', nargs=1, type=str,
                             help="<filename> - set filename as a source "
                             "for test requests.\n")
    args_parser.add_argument('--debug', action='store_true',
                             help="during a request make files "
                             "with details for debugging.\n")
    args_parser.add_argument('--elaborated_debug', action='store_true',
                             help="during a request make files "
                             "with even more details for debugging.\n")
    args_parser.add_argument('--gui', action='store_true',
                             help="during a request make files "
                             "with details for debugging.\n")
    args_parser.add_argument('--webui', action='store_true',
                             help="during a request make files\n")
    args_parser.add_argument('--log', type=set_log_level,
                             default='info', dest='log_level',
                             help="<info|debug|warn|err|crit> - set "
                             "logging level (default=info).\n")
    args_parser.add_argument('--testcase_indexes', nargs='?',
                             help="<N1,...> - set single or group of tests "
                             "to run.\n")
    args_parser.add_argument(
        '--testcase_ids',
        nargs='?',
        help="<N1,...> - set single or group of test case ids "
        "to run.\n")
    args_parser.add_argument('--table', nargs=1, type=str,
                             help="<wfa|req|resp|ap|cfg|user> - set "
                             "database table name.\n")
    args_parser.add_argument('--idx', nargs='?',
                             type=int,
                             help="<index> - set table record index.\n")
    args_parser.add_argument('--test_id',
                             default='all',
                             help="WFA test identifier, for example "
                             "srs, urs, fsp, ibp, sip, etc (default=all).\n")
    args_parser.add_argument(
        "--precision",
        metavar="PRECISION_DB",
        type=float,
        help="Maximum allowed deviation of power limits from "
        "reference values in dB. 0 means exact match is "
        "required. Default is to use hash-based exact match "
        "comparison")
    args_parser.add_argument('--cache', action='store_true',
                             help="enable cache during a request, otherwise "
                             "disabled.\n")
    args_parser.add_argument(
        '--tests2run',
        nargs=1,
        type=str,
        help="<nbr> - the total number of tests to run.\n")
    args_parser.add_argument(
        '--ca_cert',
        nargs=1,
        type=str,
        help="<filename> - set CA certificate filename to "
        "verify the remote server.\n")
    args_parser.add_argument(
        '--cli_cert',
        nargs=1,
        type=str,
        help="<filename> - set client certificate filename.\n")
    args_parser.add_argument(
        '--cli_key',
        nargs=1,
        type=str,
        help="<filename> - set client private key filename.\n")
    args_parser.add_argument('--dev_desc', action='store_true',
                             help="parse only device descriptors values.\n")
    args_parser.add_argument(
        '--prefix_cmd',
        nargs='*',
        type=str,
        help="hook to call currently provided command before "
        "main command specified by --cmd option.\n")
    args_parser.add_argument('--email_from', type=str,
                             help="<email address> - set sender email.\n")
    args_parser.add_argument('--email_to', type=str,
                             help="<email address> - set receiver email.\n")
    args_parser.add_argument(
        '--email_cc',
        type=str,
        help="<email address> - set receiver of cc email.\n")
    args_parser.add_argument('--email_pwd', type=str,
                             help="<filename> - set sender email password.\n")

    args_parser.add_argument(
        '--cmd',
        choices=execution_map.keys(),
        nargs='?',
        help="run - run test from DB and compare.\n"
        "dry_run - run test from file and show response.\n"
        "exp_adm_cfg - export admin config into a file.\n"
        "add_reqs - run test from provided file and insert with response into "
        "the databsse.\n"
        "ins_reqs - insert test vectors from provided file into the test db.\n"
        "ins_devs - insert device descriptors from provided file "
        "into the test db.\n"
        "dump_db - dump tables from the database.\n"
        "get_nbr_testcases - return number of testcases.\n"
        "parse_tests - parse WFA provided tests into a files.\n"
        "reacq - reacquision every test from the database and insert new "
        "responses.\n"
        "cmp_cfg - compare AFC config from the DB to provided from a file.\n"
        "stress - run tests in stress mode.\n"
        "ver - get version.\n")

    return args_parser


def prepare_args(parser, cfg):
    """Prepare required parameters"""
    app_log.debug(f"{inspect.stack()[0][3]}() {parser.parse_args()}")
    cfg.update(vars(parser.parse_args()))

    # check if test indexes and test ids are given
    if cfg["testcase_indexes"] and cfg["testcase_ids"]:
        # reject the request
        app_log.error('Please use either "--testcase_indexes"'
                      ' or "--testcase_ids" but not both')
        return AFC_ERR

    return execution_map[cfg['cmd']][1](cfg)


def main_execution(cfg):
    """Call all requested commands"""
    app_log.debug(f"{inspect.stack()[0][3]}()")
    if (isinstance(cfg['prefix_cmd'], list) and
            (pre_hook(cfg) == AFC_ERR)):
        return AFC_ERR
    if isinstance(cfg['cmd'], type(None)):
        parser.print_help()
        return AFC_ERR
    return execution_map[cfg['cmd']][0](cfg)


def main():
    """Main function of the utility"""
    app_log.setLevel(logging.INFO)
    console_log = logging.StreamHandler()
    console_log.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
    app_log.addHandler(console_log)

    res = AFC_OK
    parser = make_arg_parser()
    test_cfg = TestCfg()
    if prepare_args(parser, test_cfg) == AFC_ERR:
        # error in preparing arguments
        res = AFC_ERR
    else:
        res = main_execution(test_cfg)
    sys.exit(res)


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(1)


# Local Variables:
# mode: Python
# indent-tabs-mode: nil
# python-indent: 4
# End:
#
# vim: sw=4:et:tw=80:cc=+1
